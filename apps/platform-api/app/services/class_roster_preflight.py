"""Privacy-safe, read-only preflight for the full Suzhou class roster."""

from __future__ import annotations

import hashlib
from collections import Counter, defaultdict
from io import BytesIO
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from app.core.privacy import phone_hash
from app.db import fetch_all
from app.services.direct_class_preflight import (
    DEVELOPMENT_CENTERS,
    DIRECT_CLASSES,
)


REQUIRED_COLUMNS = (
    "是否在册",
    "所在分中心",
    "所属班级",
    "所属小组",
    "手机号码",
)
GROUP_NOTE_ONLY_CLASSES = {"先锋班", "神仙班"}
GROUP_NOTE_ONLY_VALUES = {"目前不读书"}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _counter(counter: Counter[str], key: str) -> list[dict[str, Any]]:
    return [
        {key: name, "count": count}
        for name, count in sorted(counter.items())
    ]


def read_workbook(
    content: bytes,
    sheet_name: str = "2026 新在册表",
) -> tuple[list[dict[str, str]], dict[str, Any]]:
    """Read only hashed matching keys and organization attributes."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{sheet_name}")
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values, ())]
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"缺少必需列：{', '.join(missing)}")

    rows: list[dict[str, str]] = []
    issues: Counter[str] = Counter()
    note_only_group_count = 0
    for value_row in values:
        if not any(_text(value) for value in value_row):
            continue
        source = dict(zip(headers, value_row))
        if _text(source.get("是否在册")) != "在册":
            continue
        center_name = _text(source.get("所在分中心"))
        class_name = _text(source.get("所属班级"))
        group_name = _text(source.get("所属小组"))
        raw_phone = _text(source.get("手机号码"))
        hashed_phone = ""
        if not class_name:
            issues["MISSING_CLASS"] += 1
        if center_name not in DEVELOPMENT_CENTERS:
            issues["UNKNOWN_CENTER"] += 1
        if not raw_phone:
            issues["MISSING_PHONE"] += 1
        else:
            try:
                hashed_phone = phone_hash(raw_phone)
            except ValueError:
                issues["INVALID_PHONE"] += 1
        if group_name and (
            class_name in GROUP_NOTE_ONLY_CLASSES
            or group_name in GROUP_NOTE_ONLY_VALUES
        ):
            note_only_group_count += 1
        rows.append(
            {
                "center_name": center_name,
                "class_name": class_name,
                "group_name": group_name,
                "phone_hash": hashed_phone,
            }
        )
    return rows, {
        "sheet_name": sheet_name,
        "source_issues": _counter(issues, "code"),
        "note_only_group_count": note_only_group_count,
    }


def _production_members(phone_hashes: set[str]) -> list[dict[str, Any]]:
    if not phone_hashes:
        return []
    result: list[dict[str, Any]] = []
    ordered = sorted(phone_hashes)
    for offset in range(0, len(ordered), 200):
        batch = ordered[offset : offset + 200]
        placeholders = ",".join("?" for _ in batch)
        result.extend(
            fetch_all(
                "SELECT id, phone_hash, status, class_name, group_name, "
                "org_unit_id, development_org_unit_id "
                f"FROM members WHERE phone_hash IN ({placeholders})",
                tuple(batch),
            )
        )
    return result


def _production_relations(member_ids: set[int]) -> list[dict[str, Any]]:
    if not member_ids:
        return []
    result: list[dict[str, Any]] = []
    ordered = sorted(member_ids)
    for offset in range(0, len(ordered), 200):
        batch = ordered[offset : offset + 200]
        placeholders = ",".join("?" for _ in batch)
        result.extend(
            fetch_all(
                "SELECT member_id, org_unit_id, relation_type "
                "FROM member_org_relations "
                f"WHERE member_id IN ({placeholders}) "
                "AND relation_type IN "
                "('PRIMARY_REGION','DEVELOPMENT_RELATION',"
                "'STUDY_CLASS','STUDY_GROUP')",
                tuple(batch),
            )
        )
    return result


def _valid_group(class_name: str, group_name: str) -> bool:
    return bool(
        class_name
        and group_name
        and class_name not in GROUP_NOTE_ONLY_CLASSES
        and group_name not in GROUP_NOTE_ONLY_VALUES
    )


def build_preflight(
    source_rows: Iterable[Mapping[str, str]],
    *,
    org_units: Iterable[Mapping[str, Any]],
    members: Iterable[Mapping[str, Any]],
    relations: Iterable[Mapping[str, Any]],
    source_name: str,
    source_sha256: str,
    sheet_name: str,
    source_issues: Iterable[Mapping[str, Any]] = (),
    note_only_group_count: int = 0,
) -> dict[str, Any]:
    """Compare a full roster with production without returning identifiers."""
    rows = list(source_rows)
    units = list(org_units)
    members_by_phone: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for member in members:
        hashed_phone = _text(member.get("phone_hash"))
        if hashed_phone:
            members_by_phone[hashed_phone].append(member)

    relation_set = {
        (
            int(relation["member_id"]),
            _text(relation["relation_type"]),
            _text(relation["org_unit_id"]),
        )
        for relation in relations
    }
    root_matches = [
        unit
        for unit in units
        if unit.get("unit_code") == "SZ_ROOT" and unit.get("is_active", 1)
    ]
    root_id = _text(root_matches[0]["id"]) if len(root_matches) == 1 else ""
    center_matches = {
        name: [
            unit
            for unit in units
            if unit.get("name") == name
            and unit.get("unit_type") == "REGIONAL_CENTER"
            and unit.get("is_active", 1)
        ]
        for name in DEVELOPMENT_CENTERS
    }
    center_ids = {
        name: _text(candidates[0]["id"])
        for name, candidates in center_matches.items()
        if len(candidates) == 1
    }

    by_center = Counter(row["center_name"] for row in rows)
    by_class = Counter(row["class_name"] for row in rows if row["class_name"])
    class_centers: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row["class_name"] and row["center_name"]:
            class_centers[row["class_name"]].add(row["center_name"])
    issues = Counter(
        {
            _text(item.get("code")) or "UNKNOWN": int(item.get("count") or 0)
            for item in source_issues
        }
    )
    if len(root_matches) != 1:
        issues["ROOT_NOT_UNIQUE"] += 1
    for candidates in center_matches.values():
        if len(candidates) != 1:
            issues["CENTER_NOT_UNIQUE"] += 1

    class_status: list[dict[str, Any]] = []
    class_org_ids: dict[str, str] = {}
    for class_name in sorted(by_class):
        is_direct = class_name in DIRECT_CLASSES
        source_centers = class_centers[class_name]
        if is_direct:
            expected_parent_id = root_id
            expected_parent = "SZ_ROOT"
        elif len(source_centers) == 1:
            expected_parent = next(iter(source_centers))
            expected_parent_id = center_ids.get(expected_parent, "")
        else:
            expected_parent = "需人工确认"
            expected_parent_id = ""
            issues["CLASS_CENTER_AMBIGUOUS"] += 1
        candidates = [
            unit
            for unit in units
            if unit.get("name") == class_name
            and unit.get("unit_type") == "CLASS"
            and unit.get("is_active", 1)
        ]
        correctly_parented = [
            unit
            for unit in candidates
            if expected_parent_id
            and _text(unit.get("parent_id")) == expected_parent_id
        ]
        if len(correctly_parented) == 1:
            action = "REUSE"
            class_org_ids[class_name] = _text(correctly_parented[0]["id"])
        elif not expected_parent_id or len(candidates) > 1:
            action = "REVIEW"
        else:
            action = "CREATE_OR_RESOLVE"
        class_status.append(
            {
                "class_name": class_name,
                "member_count": by_class[class_name],
                "scope": "DIRECT" if is_direct else "ORDINARY",
                "expected_parent": expected_parent,
                "active_class_matches": len(candidates),
                "correct_parent_matches": len(correctly_parented),
                "action": action,
            }
        )

    group_counts = Counter(
        (row["class_name"], row["group_name"])
        for row in rows
        if _valid_group(row["class_name"], row["group_name"])
    )
    group_actions = Counter()
    group_org_ids: dict[tuple[str, str], str] = {}
    for (class_name, group_name), _ in sorted(group_counts.items()):
        parent_id = class_org_ids.get(class_name, "")
        candidates = [
            unit
            for unit in units
            if unit.get("name") == group_name
            and unit.get("unit_type") == "GROUP"
            and unit.get("is_active", 1)
            and parent_id
            and _text(unit.get("parent_id")) == parent_id
        ]
        if len(candidates) == 1:
            action = "REUSE"
            group_org_ids[(class_name, group_name)] = _text(candidates[0]["id"])
        elif not parent_id or len(candidates) > 1:
            action = "REVIEW"
        else:
            action = "CREATE_OR_RESOLVE"
        group_actions[action] += 1

    source_phone_counts = Counter(
        row["phone_hash"] for row in rows if row.get("phone_hash")
    )
    matching = Counter()
    reconciliation = Counter()
    no_match_by_class = Counter()
    for row in rows:
        hashed_phone = row.get("phone_hash", "")
        if not hashed_phone:
            matching["MANUAL_REVIEW"] += 1
            continue
        if source_phone_counts[hashed_phone] != 1:
            matching["MANUAL_REVIEW"] += 1
            issues["DUPLICATE_SOURCE_PHONE"] += 1
            continue
        candidates = members_by_phone.get(hashed_phone, [])
        if not candidates:
            matching["NO_PRODUCTION_MATCH"] += 1
            no_match_by_class[row["class_name"] or "(未分班)"] += 1
            continue
        if len(candidates) != 1:
            matching["MANUAL_REVIEW"] += 1
            issues["DUPLICATE_PRODUCTION_PHONE"] += 1
            continue
        member = candidates[0]
        if _text(member.get("status")) != "ACTIVE":
            matching["INACTIVE_PRODUCTION_MATCH"] += 1
            continue
        matching["UNIQUE_ACTIVE_MATCH"] += 1
        member_id = int(member["id"])
        expected_center_id = center_ids.get(row["center_name"], "")
        expected_class_id = class_org_ids.get(row["class_name"], "")
        expected_group_id = group_org_ids.get(
            (row["class_name"], row["group_name"]), ""
        )
        if _text(member.get("class_name")) != row["class_name"]:
            reconciliation["CLASS_TEXT"] += 1
        expected_group_text = (
            row["group_name"]
            if _valid_group(row["class_name"], row["group_name"])
            else ""
        )
        if _text(member.get("group_name")) != expected_group_text:
            reconciliation["GROUP_TEXT"] += 1
        if expected_center_id:
            if _text(member.get("org_unit_id")) != expected_center_id:
                reconciliation["PRIMARY_REGION_FIELD"] += 1
            if _text(member.get("development_org_unit_id")) != expected_center_id:
                reconciliation["DEVELOPMENT_FIELD"] += 1
            if (
                member_id,
                "PRIMARY_REGION",
                expected_center_id,
            ) not in relation_set:
                reconciliation["PRIMARY_REGION_RELATION"] += 1
            if (
                member_id,
                "DEVELOPMENT_RELATION",
                expected_center_id,
            ) not in relation_set:
                reconciliation["DEVELOPMENT_RELATION"] += 1
        if expected_class_id and (
            member_id,
            "STUDY_CLASS",
            expected_class_id,
        ) not in relation_set:
            reconciliation["STUDY_CLASS"] += 1
        if expected_group_id and (
            member_id,
            "STUDY_GROUP",
            expected_group_id,
        ) not in relation_set:
            reconciliation["STUDY_GROUP"] += 1

    direct_count = sum(
        count for name, count in by_class.items() if name in DIRECT_CLASSES
    )
    missing_class_count = sum(1 for row in rows if not row["class_name"])
    return {
        "mode": "READ_ONLY_FULL_CLASS_ROSTER_PREFLIGHT",
        "automatic_production_write_allowed": False,
        "source_name": source_name,
        "source_sha256": source_sha256,
        "source": {
            "sheet_name": sheet_name,
            "active_member_count": len(rows),
            "with_class_count": len(rows) - missing_class_count,
            "missing_class_count": missing_class_count,
            "ordinary_class_member_count": len(rows)
            - missing_class_count
            - direct_count,
            "direct_class_member_count": direct_count,
            "ordinary_class_count": sum(
                1 for name in by_class if name not in DIRECT_CLASSES
            ),
            "direct_class_count": sum(
                1 for name in by_class if name in DIRECT_CLASSES
            ),
            "valid_group_pair_count": len(group_counts),
            "ordinary_group_pair_count": sum(
                1 for class_name, _ in group_counts if class_name not in DIRECT_CLASSES
            ),
            "direct_group_pair_count": sum(
                1 for class_name, _ in group_counts if class_name in DIRECT_CLASSES
            ),
            "note_only_group_count": note_only_group_count,
            "by_center": _counter(by_center, "center_name"),
            "by_class": _counter(by_class, "class_name"),
        },
        "organization": {
            "root_match_count": len(root_matches),
            "development_center_match_counts": [
                {"center_name": name, "match_count": len(candidates)}
                for name, candidates in sorted(center_matches.items())
            ],
            "class_status": class_status,
            "class_action_summary": _counter(
                Counter(item["action"] for item in class_status), "action"
            ),
            "group_action_summary": _counter(group_actions, "action"),
        },
        "matching": {
            "summary": _counter(matching, "status"),
            "no_production_match_by_class": _counter(
                no_match_by_class, "class_name"
            ),
            "fields_or_relations_needing_reconciliation": _counter(
                reconciliation, "field"
            ),
        },
        "issues": _counter(issues, "code"),
        "write_gates": [
            "本预检只返回聚合计数，不返回姓名、手机号、成员编号或组织 ID。",
            "18 名未分班学长不得猜测班级，也不进入班级签到默认名单。",
            "手机号缺失、无效、重复或生产匹配不唯一的记录必须人工复核。",
            "普通班必须唯一对应一个发展分中心；直属四班父级保持 SZ_ROOT。",
            "先锋班、神仙班和“目前不读书”的小组文本只保留备注，不创建小组关系。",
            "本接口不开放写入；迁移必须另行生成确认包、事务快照和回滚清单。",
        ],
    }


def preview_production_workbook(
    content: bytes,
    source_name: str,
    sheet_name: str = "2026 新在册表",
) -> dict[str, Any]:
    rows, source_meta = read_workbook(content, sheet_name)
    phone_hashes = {
        row["phone_hash"] for row in rows if row.get("phone_hash")
    }
    members = _production_members(phone_hashes)
    preview = build_preflight(
        rows,
        org_units=fetch_all(
            "SELECT id, unit_code, name, unit_type, parent_id, is_active "
            "FROM org_units WHERE is_active=1"
        ),
        members=members,
        relations=_production_relations(
            {int(member["id"]) for member in members}
        ),
        source_name=source_name,
        source_sha256=hashlib.sha256(content).hexdigest(),
        sheet_name=source_meta["sheet_name"],
        source_issues=source_meta["source_issues"],
        note_only_group_count=source_meta["note_only_group_count"],
    )
    return preview
