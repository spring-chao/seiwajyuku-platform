"""Guarded, transactional import for the confirmed Suzhou direct-class workbook."""
from __future__ import annotations

import hashlib
import json
from collections import Counter
from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from app.core.privacy import protected_phone
from app.db import execute, transaction
from app.services.audit import write_audit
from app.services.direct_class_preflight import (
    DEVELOPMENT_CENTERS, DIRECT_CLASSES, read_workbook, preview_production_workbook,
)

CONFIRMED_SOURCE_SHA256 = "8906040ceb28fd2fae08834bb2c891869c744d0e5ba11bb898d1dce2ea2c3252"
IMPORT_TYPE = "DIRECT_CLASS_20260729"
SOURCE_TYPE = "DIRECT_CLASS_20260729"
NORMAL_GROUP_CLASSES = {"黄埔一班", "黄埔二班"}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _rows(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["2026 新在册表"]
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values, ())]
    required = {"是否在册", "姓名", "手机号码", "所在分中心", "所属班级", "所属小组"}
    missing = required - set(headers)
    if missing:
        raise ValueError("工作簿缺少受控导入字段")
    rows: list[dict[str, str]] = []
    for values_row in values:
        row = {_text(key): _text(value) for key, value in zip(headers, values_row) if _text(key)}
        if row.get("是否在册") == "在册" and row.get("所属班级") in DIRECT_CLASSES:
            rows.append(row)
    return rows


def _upsert_relation(connection, member_id: int, org_id: str, relation_type: str, now: str) -> None:
    existing = execute(connection, "SELECT id FROM member_org_relations WHERE member_id=? AND org_unit_id=? AND relation_type=?", (member_id, org_id, relation_type)).fetchone()
    if existing:
        execute(connection, "UPDATE member_org_relations SET is_primary=1, valid_from=NULL, valid_until=NULL, source_type=?, updated_at=? WHERE id=?", (SOURCE_TYPE, now, existing["id"]))
    else:
        execute(connection, "INSERT INTO member_org_relations(member_id, org_unit_id, relation_type, is_primary, source_type, created_at, updated_at) VALUES (?, ?, ?, 1, ?, ?, ?)", (member_id, org_id, relation_type, SOURCE_TYPE, now, now))


def apply_confirmed_import(content: bytes, source_name: str, actor_user_id: int) -> dict[str, Any]:
    """Apply only the user-confirmed workbook. Any invariant failure rolls back the batch."""
    source_sha256 = hashlib.sha256(content).hexdigest()
    if source_sha256 != CONFIRMED_SOURCE_SHA256:
        raise ValueError("工作簿指纹与已确认版本不一致，已停止写入")
    preview = preview_production_workbook(content, source_name, "2026 新在册表")
    matching = {item["status"]: item["count"] for item in preview["matching"]["summary"]}
    if preview["issues"] or matching.get("MANUAL_REVIEW", 0) or matching.get("UNIQUE_PRODUCTION_MATCH", 0) != 115 or matching.get("NO_PRODUCTION_MATCH", 0) != 8:
        raise ValueError("生产预检不再满足已确认门禁，已停止写入")
    source_rows = _rows(content)
    if len(source_rows) != 123:
        raise ValueError("在册直属学员计数变化，已停止写入")
    now = datetime.now(UTC).isoformat()
    with transaction() as connection:
        cursor = execute(connection, "INSERT INTO import_batches(import_type, source_name, source_sha256, status, preview_json, created_by, created_at) VALUES (?, ?, ?, 'RUNNING', ?, ?, ?)", (IMPORT_TYPE, source_name, source_sha256, json.dumps({"count": 123, "preview": preview["matching"]}, ensure_ascii=False), actor_user_id, now))
        batch_id = cursor.lastrowid
        units = execute(connection, "SELECT id, unit_code, name, unit_type, parent_id FROM org_units WHERE is_active=1").fetchall()
        root = [unit for unit in units if unit["unit_code"] == "SZ_ROOT"]
        if len(root) != 1:
            raise ValueError("苏州塾根节点不唯一")
        root_id = root[0]["id"]
        centers = {unit["name"]: unit["id"] for unit in units if unit["unit_type"] == "REGIONAL_CENTER" and unit["name"] in DEVELOPMENT_CENTERS}
        if len(centers) != len(DEVELOPMENT_CENTERS):
            raise ValueError("六个发展分中心解析失败")
        class_ids: dict[str, str] = {}
        for index, name in enumerate(DIRECT_CLASSES, 1):
            matches = [unit for unit in units if unit["name"] == name and unit["unit_type"] == "CLASS"]
            if len(matches) > 1:
                raise ValueError("班级名称存在重复组织，已停止写入")
            if matches:
                if matches[0]["parent_id"] != root_id:
                    raise ValueError("直属班级组织归属异常，已停止写入")
                class_ids[name] = matches[0]["id"]
            else:
                unit_id = str(uuid4())
                execute(connection, "INSERT INTO org_units(id, unit_code, name, unit_type, parent_id, is_active, created_at, updated_at) VALUES (?, ?, ?, 'CLASS', ?, 1, ?, ?)", (unit_id, f"SZ_DIRECT_{index}", name, root_id, now, now))
                class_ids[name] = unit_id
        groups = {
            (row["所属班级"], row["所属小组"])
            for row in source_rows
            if row["所属班级"] in NORMAL_GROUP_CLASSES
            and row["所属小组"]
            and row["所属小组"] != "目前不读书"
        }
        if len(groups) != 11:
            raise ValueError("普通班小组候选不等于 11，已停止写入")
        group_ids: dict[tuple[str, str], str] = {}
        for class_name, group_name in sorted(groups):
            existing = execute(connection, "SELECT id FROM org_units WHERE is_active=1 AND unit_type='GROUP' AND parent_id=? AND name=?", (class_ids[class_name], group_name)).fetchall()
            if len(existing) > 1:
                raise ValueError("直属普通班小组重复，已停止写入")
            if existing:
                group_ids[(class_name, group_name)] = existing[0]["id"]
            else:
                group_id = str(uuid4())
                execute(connection, "INSERT INTO org_units(id, unit_code, name, unit_type, parent_id, is_active, created_at, updated_at) VALUES (?, ?, ?, 'GROUP', ?, 1, ?, ?)", (group_id, f"SZ_DIRECT_GROUP_{len(group_ids)+1}", group_name, class_ids[class_name], now, now))
                group_ids[(class_name, group_name)] = group_id
        created = updated = note_count = relation_count = 0
        for row in source_rows:
            phone = protected_phone(row["手机号码"])
            matches = execute(connection, "SELECT id, notes FROM members WHERE phone_hash=?", (phone["phone_hash"],)).fetchall()
            if len(matches) > 1:
                raise ValueError("生产手机号匹配不唯一，已停止写入")
            center_id = centers.get(row["所在分中心"])
            if not center_id:
                raise ValueError("发展分中心无效，已停止写入")
            class_name, group_name = row["所属班级"], row["所属小组"]
            note = ""
            if (class_name == "先锋班" and group_name) or group_name == "目前不读书":
                note = f"原所属小组：{group_name}"
            if matches:
                member_id = matches[0]["id"]
                old_notes = _text(matches[0]["notes"])
                new_notes = old_notes if not note or note in old_notes else (old_notes + "\n" if old_notes else "") + note
                execute(connection, "UPDATE members SET org_unit_id=?, development_org_unit_id=?, class_name=?, group_name=?, notes=?, updated_at=? WHERE id=?", (center_id, center_id, class_name, group_name if (class_name in NORMAL_GROUP_CLASSES and group_name != "目前不读书") else None, new_notes or None, now, member_id))
                updated += 1
                note_count += bool(note)
            else:
                cursor = execute(connection, "INSERT INTO members(member_code, name, org_unit_id, development_org_unit_id, status, phone_ciphertext, phone_hash, phone_last4, phone_masked, class_name, group_name, notes, created_at, updated_at) VALUES (?, ?, ?, ?, 'ACTIVE', ?, ?, ?, ?, ?, ?, ?, ?, ?)", (f"DC-{uuid4().hex[:12].upper()}", row["姓名"], center_id, center_id, phone["phone_ciphertext"], phone["phone_hash"], phone["phone_last4"], phone["phone_masked"], class_name, group_name if (class_name in NORMAL_GROUP_CLASSES and group_name != "目前不读书") else None, note or None, now, now))
                member_id = cursor.lastrowid
                created += 1
                note_count += bool(note)
            for relation_type, org_id in (("PRIMARY_REGION", center_id), ("DEVELOPMENT_RELATION", center_id), ("STUDY_CLASS", class_ids[class_name])):
                _upsert_relation(connection, member_id, org_id, relation_type, now); relation_count += 1
            if class_name in NORMAL_GROUP_CLASSES and group_name and group_name != "目前不读书":
                _upsert_relation(connection, member_id, group_ids[(class_name, group_name)], "STUDY_GROUP", now); relation_count += 1
        if (created, updated, relation_count, note_count) != (8, 115, 430, 4):
            raise ValueError("导入后计数与确认包不一致，事务已回滚")
        execute(connection, "UPDATE import_batches SET status='APPLIED', applied_at=? WHERE id=?", (now, batch_id))
        write_audit(connection, actor_user_id=actor_user_id, action="direct_class_import.apply", resource_type="import_batch", resource_id=str(batch_id), after={"source_sha256": source_sha256, "created": created, "updated": updated, "relations": relation_count, "notes": note_count})
    return {"batch_id": batch_id, "created": created, "updated": updated, "relations": relation_count, "notes": note_count, "source_sha256": source_sha256}
