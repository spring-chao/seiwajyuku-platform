"""Privacy-safe, read-only production preflight for Suzhou direct classes."""

from __future__ import annotations

import hashlib
from collections import Counter, defaultdict
from io import BytesIO
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from app.core.privacy import phone_hash
from app.db import fetch_all


DIRECT_CLASSES = ("黄埔一班", "黄埔二班", "先锋班", "神仙班")
DEVELOPMENT_CENTERS = (
    "园区分中心",
    "昆山分中心",
    "吴江分中心",
    "新吴分中心",
    "张家港分中心",
    "姑苏相城分中心",
)
REQUIRED_COLUMNS = ("是否在册", "所在分中心", "所属班级", "手机号码")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _counter(counter: Counter[str], key: str) -> list[dict[str, Any]]:
    return [{key: name, "count": count} for name, count in sorted(counter.items())]


def read_workbook(content: bytes, sheet_name: str | None = None) -> tuple[list[dict[str, str]], dict[str, Any]]:
    """Return only hashed matching keys; source names and phones never leave this function."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    selected = sheet_name or workbook.sheetnames[0]
    if selected not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{selected}")
    sheet = workbook[selected]
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values, ())]
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"缺少必需列：{', '.join(missing)}")

    rows: list[dict[str, str]] = []
    issues: Counter[str] = Counter()
    for value_row in values:
        if not any(_text(value) for value in value_row):
            continue
        source = dict(zip(headers, value_row))
        if _text(source.get("是否在册")) != "在册":
            continue
        class_name = _text(source.get("所属班级"))
        if class_name not in DIRECT_CLASSES:
            continue
        center_name = _text(source.get("所在分中心"))
        raw_phone = _text(source.get("手机号码"))
        hashed_phone = ""
        if not raw_phone:
            issues["MISSING_PHONE"] += 1
        else:
            try:
                hashed_phone = phone_hash(raw_phone)
            except ValueError:
                issues["INVALID_PHONE"] += 1
        rows.append({
            "class_name": class_name,
            "center_name": center_name,
            "phone_hash": hashed_phone,
        })
    return rows, {
        "sheet_name": selected,
        "source_row_count": len(rows),
        "source_issues": _counter(issues, "code"),
    }


def _production_members(phone_hashes: set[str]) -> list[dict[str, Any]]:
    if not phone_hashes:
        return []
    result: list[dict[str, Any]] = []
    ordered = sorted(phone_hashes)
    for offset in range(0, len(ordered), 200):
        batch = ordered[offset: offset + 200]
        placeholders = ",".join("?" for _ in batch)
        result.extend(fetch_all(
            "SELECT phone_hash, class_name, org_unit_id, development_org_unit_id "
            f"FROM members WHERE phone_hash IN ({placeholders})",
            tuple(batch),
        ))
    return result


def build_preflight(
    source_rows: Iterable[Mapping[str, str]],
    *,
    org_units: Iterable[Mapping[str, Any]],
    members: Iterable[Mapping[str, Any]],
    source_name: str,
    source_sha256: str,
    source_issues: Iterable[Mapping[str, Any]] = (),
) -> dict[str, Any]:
    """Compare hashed source records with production records without emitting identifiers."""
    rows = list(source_rows)
    units = list(org_units)
    members_by_phone: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for member in members:
        if member.get("phone_hash"):
            members_by_phone[str(member["phone_hash"])].append(member)

    root_matches = [unit for unit in units if unit.get("unit_code") == "SZ_ROOT" and unit.get("is_active", 1)]
    root_id = root_matches[0]["id"] if len(root_matches) == 1 else None
    center_matches = {
        name: [
            unit for unit in units
            if unit.get("name") == name and unit.get("unit_type") == "REGIONAL_CENTER" and unit.get("is_active", 1)
        ]
        for name in DEVELOPMENT_CENTERS
    }
    center_ids = {
        name: candidates[0]["id"]
        for name, candidates in center_matches.items()
        if len(candidates) == 1
    }
    class_status: list[dict[str, Any]] = []
    for class_name in DIRECT_CLASSES:
        candidates = [
            unit for unit in units
            if unit.get("name") == class_name and unit.get("unit_type") == "CLASS" and unit.get("is_active", 1)
        ]
        correctly_parented = [unit for unit in candidates if unit.get("parent_id") == root_id]
        class_status.append({
            "class_name": class_name,
            "active_class_matches": len(candidates),
            "correct_parent_matches": len(correctly_parented),
            "action": "REUSE" if len(correctly_parented) == 1 else "CREATE_OR_RESOLVE",
        })

    by_class = Counter(row["class_name"] for row in rows)
    source_phone_counts = Counter(row["phone_hash"] for row in rows if row.get("phone_hash"))
    matching = Counter()
    unmatched_by_class = Counter()
    profile_updates_by_reason = Counter()
    issues = Counter(item.get("code", "UNKNOWN") for item in source_issues)
    if len(root_matches) != 1:
        issues["ROOT_NOT_UNIQUE"] += 1
    for center_name, candidates in center_matches.items():
        if len(candidates) != 1:
            issues["CENTER_NOT_UNIQUE"] += 1
    for row in rows:
        hashed_phone = row.get("phone_hash", "")
        if not hashed_phone:
            matching["MANUAL_REVIEW"] += 1
            continue
        if source_phone_counts[hashed_phone] != 1:
            matching["MANUAL_REVIEW"] += 1
            issues["DUPLICATE_SOURCE_PHONE"] += 1
            continue
        candidates = members_by_phone.get(hashed_phone, [])
        if not candidates:
            matching["NO_PRODUCTION_MATCH"] += 1
            unmatched_by_class[row["class_name"]] += 1
            continue
        if len(candidates) != 1:
            matching["MANUAL_REVIEW"] += 1
            issues["DUPLICATE_PRODUCTION_PHONE"] += 1
            continue
        matching["UNIQUE_PRODUCTION_MATCH"] += 1
        member = candidates[0]
        expected_center_id = center_ids.get(row["center_name"])
        if not expected_center_id:
            profile_updates_by_reason["CENTER_UNRESOLVED"] += 1
            continue
        if _text(member.get("class_name")) != row["class_name"]:
            profile_updates_by_reason["CLASS_TEXT"] += 1
        if member.get("org_unit_id") != expected_center_id:
            profile_updates_by_reason["PRIMARY_REGION"] += 1
        if member.get("development_org_unit_id") != expected_center_id:
            profile_updates_by_reason["DEVELOPMENT_RELATION"] += 1

    existing_direct = Counter(
        _text(member.get("class_name"))
        for member in members
        if _text(member.get("class_name")) in DIRECT_CLASSES
    )
    # Include direct-class population from production even where it is not in the workbook match set.
    return {
        "mode": "READ_ONLY_PRODUCTION_PREFLIGHT",
        "automatic_production_write_allowed": False,
        "source_name": source_name,
        "source_sha256": source_sha256,
        "source": {
            "active_direct_member_count": len(rows),
            "by_class": _counter(by_class, "class_name"),
        },
        "organization": {
            "root_unit_code": "SZ_ROOT",
            "root_match_count": len(root_matches),
            "development_center_match_counts": [
                {"center_name": name, "match_count": len(candidates)}
                for name, candidates in sorted(center_matches.items())
            ],
            "direct_class_status": class_status,
        },
        "matching": {
            "summary": _counter(matching, "status"),
            "no_production_match_by_class": _counter(unmatched_by_class, "class_name"),
            "matched_profile_fields_needing_reconciliation": _counter(profile_updates_by_reason, "field"),
        },
        "issues": _counter(issues, "code"),
        "write_gates": [
            "预检只返回聚合计数，不返回姓名、手机号、成员编号或组织 ID。",
            "任何 MANUAL_REVIEW、未唯一解析组织或源文件指纹变化都必须停止生产写入。",
            "NO_PRODUCTION_MATCH 只能进入受确认的新建批次；不得推断或覆盖既有归属。",
            "写入前必须另行生成事务快照、审计批次与回滚清单，并取得当次明确确认。",
        ],
    }


def preview_production_workbook(content: bytes, source_name: str, sheet_name: str | None = None) -> dict[str, Any]:
    rows, source_meta = read_workbook(content, sheet_name)
    phone_hashes = {row["phone_hash"] for row in rows if row.get("phone_hash")}
    org_units = fetch_all(
        "SELECT id, unit_code, name, unit_type, parent_id, is_active FROM org_units WHERE is_active=1"
    )
    # A separate aggregate query captures existing direct-class records which do not
    # appear in the source workbook; no row-level identifiers are returned.
    production_direct_counts = fetch_all(
        "SELECT class_name, COUNT(*) AS count FROM members "
        "WHERE class_name IN (?, ?, ?, ?) GROUP BY class_name",
        DIRECT_CLASSES,
    )
    preview = build_preflight(
        rows,
        org_units=org_units,
        members=_production_members(phone_hashes),
        source_name=source_name,
        source_sha256=hashlib.sha256(content).hexdigest(),
        source_issues=source_meta["source_issues"],
    )
    preview["production_existing_direct_class_records"] = [
        {"class_name": _text(item["class_name"]), "count": int(item["count"])}
        for item in sorted(production_direct_counts, key=lambda item: _text(item["class_name"]))
    ]
    preview["source"]["sheet_name"] = source_meta["sheet_name"]
    return preview
