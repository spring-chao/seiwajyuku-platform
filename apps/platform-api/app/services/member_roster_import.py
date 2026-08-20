"""Guarded, privacy-safe supplement import for the member master workbook.

The workbook is treated as a source for missing member/profile facts. Existing
non-empty profile fields and existing organization assignments are never
overwritten automatically. Phone numbers are hashed/encrypted in memory and
are never returned in preview or import responses.
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, UTC
from io import BytesIO
from typing import Any, Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.core.privacy import protected_phone, encrypt_text, decrypt_text
from app.core.settings import get_settings
from app.db import execute, fetch_all, fetch_one, transaction
from app.services.audit import write_audit
from app.services.iam import accessible_org_ids, user_context
from app.services.organization_policy import is_valid_member_class_parent


IMPORT_TYPE = "MEMBER_PROFILE_SUPPLEMENT"
SOURCE_TYPE = "MEMBER_ROSTER_SUPPLEMENT"
STATUS_MAP = {"在册": "ACTIVE", "流失": "INACTIVE", "暂停": "SUSPENDED"}
SPECIAL_GROUP_VALUES = {"目前不读书"}
FIELD_MAP = {
    "职务": "position",
    "公司名称": "company_name",
    "隶属区": "district",
    "公司地址": "company_address",
    "班组委名称": "class_committee_name",
    "推荐人": "referrer",
    "推荐人所属分中心": "referrer_center",
    "行业分类": "industry_category",
    "所属行业": "industry",
    "公司产品": "company_products",
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _date_value(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date().isoformat()
            if isinstance(parsed, date):
                return parsed.isoformat()
        except (TypeError, ValueError, OverflowError):
            return None
    text = _text(value)
    for pattern in (r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?$", r"^(\d{4})[-/.年](\d{1,2})月?$"):
        match = re.match(pattern, text)
        if match:
            try:
                if len(match.groups()) == 3:
                    return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
                return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-01"
            except ValueError:
                return None
    return text if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text) else None


def _month_value(value: Any) -> str | None:
    parsed = _date_value(value)
    return parsed[:7] if parsed else None


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", _text(value).replace(",", ""))
    return float(match.group(0)) if match else None


def _employee_count(value: Any) -> int | None:
    parsed = _number(value)
    return int(parsed) if parsed is not None and parsed >= 0 else None


def _gender(value: Any) -> str | None:
    value = _text(value)
    return {"男": "MALE", "女": "FEMALE", "其他": "UNSPECIFIED"}.get(value, value or None)


def _counter_rows(counter: Counter[str], key: str) -> list[dict[str, Any]]:
    return [{key: name, "count": count} for name, count in sorted(counter.items())]


def _read_rows(content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values, ())]
    required = {"姓名", "是否在册", "手机号码", "所在分中心"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"工作表缺少必需列：{', '.join(missing)}")
    rows: list[dict[str, Any]] = []
    issues: Counter[str] = Counter()
    for row_number, value_row in enumerate(values, 2):
        if not any(_text(value) for value in value_row):
            continue
        source = {header: value_row[index] for index, header in enumerate(headers) if header}
        status_text = _text(source.get("是否在册"))
        if status_text not in STATUS_MAP:
            issues["UNKNOWN_STATUS"] += 1
            continue
        rows.append({"source_row": row_number, **source, "status": STATUS_MAP[status_text]})
    return rows, {
        "sheet_name": sheet.title,
        "row_count": len(rows),
        "status_counts": Counter(row["status"] for row in rows),
        "issues": issues,
    }


def _source_fields(row: dict[str, Any]) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    for source_key, target_key in FIELD_MAP.items():
        value = _text(row.get(source_key))
        if value:
            fields[target_key] = value
    for source_key, target_key in (("生日时间", "birthday"), ("入塾日期", "join_date"), ("学习时间", "study_start_date")):
        value = _date_value(row.get(source_key))
        if value:
            fields[target_key] = value
    renewal_month = _month_value(row.get("26年续费时间"))
    if renewal_month and renewal_month.startswith("2026-"):
        fields["renewal_month"] = renewal_month
    employee_count = _employee_count(row.get("员工人数"))
    if employee_count is not None:
        fields["employee_count"] = employee_count
    membership_years = _number(row.get("入塾年限"))
    if membership_years is not None and 0 <= membership_years <= 100:
        fields["membership_years"] = round(membership_years, 1)
    gender = _gender(row.get("性别"))
    if gender:
        fields["gender"] = gender
    note = _text(row.get("备注"))
    if note:
        fields["notes"] = note
    sales = _text(row.get("销售收入"))
    if sales:
        fields["annual_sales"] = sales
    return fields


def _field_non_empty(rows: Iterable[dict[str, Any]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for row in rows:
        fields = _source_fields(row)
        counter.update(fields.keys())
    return counter


def _active_units() -> list[dict[str, Any]]:
    return fetch_all(
        "SELECT id, unit_code, name, unit_type, parent_id, is_active "
        "FROM org_units WHERE is_active=1"
    )


def _resolve_relations(row: dict[str, Any], *, units: list[dict[str, Any]], center_id: str) -> dict[str, str | None]:
    class_name = _text(row.get("所属班级"))
    group_name = _text(row.get("所属小组"))
    root_ids = {str(unit["id"]) for unit in units if unit.get("unit_code") == "SZ_ROOT"}
    class_candidates = [
        unit for unit in units
        if class_name and unit.get("name") == class_name
        and unit.get("unit_type") in {"CLASS", "SPECIAL_COHORT"}
        and (str(unit.get("parent_id")) == str(center_id) or str(unit.get("parent_id")) in root_ids)
        and is_valid_member_class_parent(
            class_name=class_name,
            parent_id=unit.get("parent_id"),
            member_center_id=center_id,
        )
    ]
    class_id = str(class_candidates[0]["id"]) if len(class_candidates) == 1 else None
    class_type = class_candidates[0].get("unit_type") if len(class_candidates) == 1 else None
    group_id = None
    if class_id and group_name and group_name not in SPECIAL_GROUP_VALUES:
        groups = [
            unit for unit in units
            if unit.get("unit_type") == "GROUP"
            and str(unit.get("parent_id")) == class_id
            and unit.get("name") == group_name
        ]
        if len(groups) == 1:
            group_id = str(groups[0]["id"])
    return {"class_id": class_id, "class_type": class_type, "group_id": group_id}


def _member_fields_for_response(fields: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in fields.items() if key not in {"annual_sales", "phone"}}


def _masked_phone(phone: dict[str, Any] | None) -> str | None:
    """Return a reviewer-safe phone hint without retaining a raw number."""
    last4 = _text((phone or {}).get("phone_last4"))
    return f"****{last4}" if last4 else None


def _financial_data(member: dict[str, Any] | None) -> dict[str, Any]:
    """Read encrypted financial data only for an already-authorized internal decision."""
    ciphertext = (member or {}).get("enterprise_financial_ciphertext")
    if not ciphertext:
        return {}
    try:
        return json.loads(decrypt_text(ciphertext))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}


def _manual_review_item(item: dict[str, Any]) -> dict[str, Any]:
    """Reviewer identity is minimal: name, source row and masked phone only."""
    row = item["row"]
    return {
        "source_row": row["source_row"],
        "name": _text(row.get("姓名")),
        "phone_masked": _masked_phone(item.get("phone")),
        "center_name": _text(row.get("所在分中心")) or None,
        "class_name": _text(row.get("所属班级")) or None,
        "group_name": _text(row.get("所属小组")) or None,
        "reasons": sorted(item["reasons"]),
    }


def _build_plan(rows: list[dict[str, Any]], *, actor_user_id: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    allowed = accessible_org_ids(actor_user_id)
    actor = user_context(actor_user_id)
    allow_sensitive = bool(actor and "members:enterprise_view" in actor["permissions"])
    units = _active_units()
    centers_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for unit in units:
        if unit.get("unit_type") == "REGIONAL_CENTER":
            centers_by_name[_text(unit.get("name"))].append(unit)
    members = fetch_all(
        "SELECT id, name, status, phone_hash, phone_last4, org_unit_id, "
        "development_org_unit_id, company_name, gender, district, company_address, "
        "class_name, class_committee_name, group_name, birthday, join_date, "
        "study_start_date, membership_years, membership_years_overridden, renewal_month, "
        "position, referrer, referrer_center, industry_category, industry, "
        "company_products, employee_count, notes, company_size, enterprise_financial_ciphertext "
        "FROM members"
    )
    by_phone: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for member in members:
        if member.get("phone_hash"):
            by_phone[str(member["phone_hash"])].append(member)
        by_name[_text(member.get("name"))].append(member)
    # Count all source phones before classifying rows.  If a duplicate appears,
    # every occurrence must go to manual review; classifying the first row as
    # ready before seeing the later duplicate would make the result order
    # dependent and could create a partial, unsafe match.
    source_phone_counts: Counter[str] = Counter()
    for row in rows:
        raw_phone = _text(row.get("手机号码"))
        if not raw_phone:
            continue
        try:
            source_phone_counts[protected_phone(raw_phone)["phone_hash"]] += 1
        except ValueError:
            continue
    prepared: list[dict[str, Any]] = []
    for row in rows:
        raw_phone = _text(row.get("手机号码"))
        phone = None
        if raw_phone:
            try:
                phone = protected_phone(raw_phone)
            except ValueError:
                pass
        center_name = _text(row.get("所在分中心"))
        center_candidates = centers_by_name.get(center_name, [])
        center_id = str(center_candidates[0]["id"]) if len(center_candidates) == 1 else None
        source_fields = _source_fields(row)
        matches = by_phone.get(phone["phone_hash"], []) if phone else []
        status = "MANUAL_REVIEW"
        member = None
        reasons: list[str] = []
        if not phone:
            reasons.append("MISSING_OR_INVALID_PHONE")
        elif source_phone_counts[phone["phone_hash"]] != 1:
            reasons.append("DUPLICATE_SOURCE_PHONE")
        elif len(matches) > 1:
            reasons.append("DUPLICATE_PRODUCTION_PHONE")
        elif len(matches) == 1:
            member = matches[0]
            if _text(member.get("name")) != _text(row.get("姓名")):
                reasons.append("NAME_PHONE_CONFLICT")
            if center_id is None:
                reasons.append("CENTER_NOT_UNIQUE")
            elif allowed is not None and center_id not in allowed and str(member.get("org_unit_id")) not in allowed:
                reasons.append("ORG_SCOPE")
            else:
                status = "READY_EXISTING"
        else:
            if center_id is None:
                reasons.append("CENTER_NOT_UNIQUE")
            elif allowed is not None and center_id not in allowed:
                reasons.append("ORG_SCOPE")
            elif by_name.get(_text(row.get("姓名"))):
                reasons.append("NAME_EXISTS_PHONE_UNMATCHED")
            else:
                status = "READY_NEW"
        relation_ids = _resolve_relations(row, units=units, center_id=center_id) if center_id else {"class_id": None, "class_type": None, "group_id": None}
        prepared.append({
            "row": row,
            "phone": phone,
            "member": member,
            "center_id": center_id,
            "source_fields": source_fields,
            "relation_ids": relation_ids,
            "status": status,
            "reasons": reasons,
        })

    summary: Counter[str] = Counter(item["status"] for item in prepared)
    field_fill: Counter[str] = Counter()
    relation_ready = Counter()
    sensitive_count = 0
    for item in prepared:
        if item["status"] not in {"READY_EXISTING", "READY_NEW"}:
            continue
        fields = item["source_fields"]
        if fields.get("annual_sales"):
            sensitive_count += 1
        if item["status"] == "READY_NEW":
            field_fill.update(
                key for key in fields if key != "annual_sales" or allow_sensitive
            )
        else:
            member = item["member"]
            current_financial = _financial_data(member) if allow_sensitive else {}
            field_fill.update(
                key
                for key, value in fields.items()
                if value not in (None, "")
                and not (current_financial.get("annual_sales") if key == "annual_sales" else member.get(key))
                and (key != "annual_sales" or allow_sensitive)
            )
        relation_ready["class"] += bool(item["relation_ids"].get("class_id"))
        relation_ready["group"] += bool(item["relation_ids"].get("group_id"))
    issue_counter = Counter(reason for item in prepared for reason in item["reasons"])
    return prepared, {
        "matching": {
            "summary": _counter_rows(summary, "status"),
            "new_member_count": summary["READY_NEW"],
            "existing_member_count": summary["READY_EXISTING"],
            "manual_review_count": summary["MANUAL_REVIEW"],
            "field_fill_counts": _counter_rows(field_fill, "field"),
        },
        "organization": {
            "class_relation_ready_count": relation_ready["class"],
            "group_relation_ready_count": relation_ready["group"],
        },
        "sensitive": {
            "annual_sales_source_count": sensitive_count,
            "requires_enterprise_permission": sensitive_count > 0,
            "enterprise_financial_write_allowed": allow_sensitive,
            "annual_sales_ready_count": field_fill["annual_sales"] if allow_sensitive else None,
        },
        "issues": _counter_rows(issue_counter, "code"),
        "manual_review_items": [
            _manual_review_item(item)
            for item in prepared
            if item["status"] == "MANUAL_REVIEW"
        ],
    }


def preview_member_roster(content: bytes, source_name: str, actor_user_id: int) -> dict[str, Any]:
    rows, meta = _read_rows(content)
    _, plan = _build_plan(rows, actor_user_id=actor_user_id)
    field_counts = _field_non_empty(rows)
    return {
        "mode": "READ_ONLY_MEMBER_PROFILE_SUPPLEMENT",
        "automatic_production_write_allowed": False,
        "source_name": source_name,
        "source_sha256": hashlib.sha256(content).hexdigest(),
        "source": {
            "sheet_name": meta["sheet_name"],
            "row_count": meta["row_count"],
            "status_counts": _counter_rows(meta["status_counts"], "status"),
            "field_non_empty_counts": _counter_rows(field_counts, "field"),
        },
        **plan,
        "write_gates": [
            "只补充空白学员资料，不覆盖已有非空字段、现有状态或现有分中心归属。",
            "手机号必须唯一匹配；缺失、无效、重复或姓名冲突的记录进入人工复核。",
            "班级和小组只在正式组织唯一匹配时补关系，无法唯一匹配不猜测。",
            "销售收入属于企业敏感资料，只有具备 members:enterprise_view 的账号才会写入加密字段。",
            "预检不落库；正式导入必须再次上传同一 SHA-256 文件并填写确认文字。",
        ],
    }


def _next_member_code(connection, source_row: dict[str, Any]) -> str:
    source_code = re.sub(r"[^0-9A-Za-z_-]", "", _text(source_row.get("序号")))
    base = f"IMPORT-2026-{source_code}" if source_code else f"IMPORT-2026-{uuid4().hex[:12].upper()}"
    candidate = base[:128]
    if not execute(connection, "SELECT id FROM members WHERE member_code=?", (candidate,)).fetchone():
        return candidate
    return f"{base[:112]}-{uuid4().hex[:12].upper()}"


def _upsert_relation(connection, member_id: int, org_id: str, relation_type: str, now: str) -> bool:
    existing = execute(
        connection,
        "SELECT id FROM member_org_relations WHERE member_id=? AND org_unit_id=? AND relation_type=? LIMIT 1",
        (member_id, org_id, relation_type),
    ).fetchone()
    if existing:
        return False
    execute(
        connection,
        "INSERT INTO member_org_relations(member_id, org_unit_id, relation_type, is_primary, source_type, created_at, updated_at) "
        "VALUES (?, ?, ?, 1, ?, ?, ?)",
        (member_id, org_id, relation_type, SOURCE_TYPE, now, now),
    )
    return True


def _apply_fields(current: dict[str, Any] | None, fields: dict[str, Any], *, allow_sensitive: bool) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in fields.items():
        if key == "annual_sales" and not allow_sensitive:
            continue
        if value in (None, ""):
            continue
        if current is None or current.get(key) in (None, ""):
            result[key] = value
    return result


def apply_member_roster(content: bytes, source_name: str, actor_user_id: int, confirmation_text: str) -> dict[str, Any]:
    settings = get_settings()
    if settings.is_production and not settings.allow_production_mutations:
        raise PermissionError("生产写入开关未开启")
    if confirmation_text != "确认补充导入学员主档":
        raise ValueError("确认文字不匹配，已禁止学员主档导入")
    user = user_context(actor_user_id)
    if not user or "members:manage" not in user["permissions"]:
        raise PermissionError("当前角色不能导入学员主档")
    source_sha256 = hashlib.sha256(content).hexdigest()
    existing_batch = fetch_one(
        "SELECT id, status FROM import_batches WHERE import_type=? AND source_sha256=? ORDER BY id DESC LIMIT 1",
        (IMPORT_TYPE, source_sha256),
    )
    if existing_batch and existing_batch["status"] == "APPLIED":
        return {"batch_id": existing_batch["id"], "status": "ALREADY_APPLIED", "source_sha256": source_sha256}
    rows, meta = _read_rows(content)
    prepared, plan = _build_plan(rows, actor_user_id=actor_user_id)
    if not plan["matching"]["existing_member_count"] and not plan["matching"]["new_member_count"]:
        raise ValueError("预检没有可安全导入的记录")
    allow_sensitive = "members:enterprise_view" in user["permissions"]
    if plan["sensitive"]["requires_enterprise_permission"] and not allow_sensitive:
        raise PermissionError("当前账号缺少企业敏感资料权限，无法导入销售收入")
    now = datetime.now(UTC).isoformat()
    created = updated = field_count = relation_count = annual_sales_applied = 0
    with transaction() as connection:
        cursor = execute(
            connection,
            "INSERT INTO import_batches(import_type, source_name, source_sha256, status, preview_json, created_by, created_at) "
            "VALUES (?, ?, ?, 'RUNNING', ?, ?, ?)",
            (IMPORT_TYPE, source_name, source_sha256, json.dumps(plan, ensure_ascii=False), actor_user_id, now),
        )
        batch_id = cursor.lastrowid
        for item in prepared:
            if item["status"] == "MANUAL_REVIEW":
                # These rows are listed for human reconciliation and must never
                # reach the existing-member update path.
                continue
            row = item["row"]
            fields = dict(item["source_fields"])
            if not allow_sensitive:
                fields.pop("annual_sales", None)
            if item["status"] == "READY_NEW":
                phone = item["phone"]
                if not phone:
                    raise ValueError("预检与导入阶段手机号状态不一致")
                member_code = _next_member_code(connection, row)
                center_id = item["center_id"]
                class_name = _text(row.get("所属班级")) or None
                group_name = _text(row.get("所属小组")) or None
                financial_data = {"annual_sales": fields.pop("annual_sales")} if fields.get("annual_sales") else {}
                financial_ciphertext = encrypt_text(json.dumps(financial_data, ensure_ascii=False)) if financial_data else None
                annual_sales_applied += int(bool(financial_data.get("annual_sales")))
                columns = {
                    "member_code": member_code,
                    "name": _text(row.get("姓名")),
                    "org_unit_id": center_id,
                    "development_org_unit_id": center_id,
                    "status": row["status"],
                    **phone,
                    "class_name": class_name,
                    "group_name": group_name if group_name not in SPECIAL_GROUP_VALUES else None,
                    **fields,
                    "enterprise_financial_ciphertext": financial_ciphertext,
                    "created_at": now,
                    "updated_at": now,
                }
                columns.setdefault("membership_years_overridden", 1 if "membership_years" in columns else 0)
                columns.setdefault("renewal_month_overridden", 1 if "renewal_month" in columns else 0)
                names = ", ".join(columns)
                placeholders = ", ".join("?" for _ in columns)
                cursor = execute(connection, f"INSERT INTO members({names}) VALUES ({placeholders})", tuple(columns.values()))
                member_id = cursor.lastrowid
                created += 1
                field_count += len(fields)
                relation_count += int(_upsert_relation(connection, member_id, center_id, "PRIMARY_REGION", now))
                relation_count += int(_upsert_relation(connection, member_id, center_id, "DEVELOPMENT_RELATION", now))
                class_id = item["relation_ids"].get("class_id")
                group_id = item["relation_ids"].get("group_id")
                if class_id:
                    relation_type = "SPECIAL_COHORT" if item["relation_ids"].get("class_type") == "SPECIAL_COHORT" else "STUDY_CLASS"
                    relation_count += int(_upsert_relation(connection, member_id, class_id, relation_type, now))
                if group_id:
                    relation_count += int(_upsert_relation(connection, member_id, group_id, "STUDY_GROUP", now))
                write_audit(connection, actor_user_id=actor_user_id, action="members.roster_import.create", resource_type="member", resource_id=str(member_id), org_unit_id=center_id, after={"source_batch_id": batch_id, "changed_fields": sorted(_member_fields_for_response(fields))})
                continue
            current = fetch_one("SELECT * FROM members WHERE id=?", (item["member"]["id"],))
            if not current:
                raise ValueError("导入期间学员记录发生变化，已停止写入")
            current_financial = _financial_data(current) if allow_sensitive else {}
            current["annual_sales"] = current_financial.get("annual_sales")
            changes = _apply_fields(current, fields, allow_sensitive=allow_sensitive)
            center_id = item["center_id"]
            if center_id and not current.get("development_org_unit_id"):
                changes["development_org_unit_id"] = center_id
            if "annual_sales" in changes:
                financial_data = current_financial
                financial_data["annual_sales"] = changes.pop("annual_sales")
                changes["enterprise_financial_ciphertext"] = encrypt_text(json.dumps(financial_data, ensure_ascii=False))
                annual_sales_applied += 1
            if current.get("join_date") in (None, "") and changes.get("join_date") and "membership_years" not in changes:
                changes["membership_years_overridden"] = 0
            if "renewal_month" in changes:
                changes["renewal_month_overridden"] = 1
            if changes:
                changes["updated_at"] = now
                assignments = ", ".join(f"{key}=?" for key in changes)
                execute(connection, f"UPDATE members SET {assignments} WHERE id=?", (*changes.values(), current["id"]))
                updated += 1
                field_count += len([key for key in changes if key != "updated_at"])
                before = {key: current.get(key) for key in changes if key != "updated_at" and key != "enterprise_financial_ciphertext"}
                after = {key: changes[key] for key in before}
                execute(connection, "INSERT INTO member_change_history(member_id, change_type, before_json, after_json, changed_by, changed_at) VALUES (?, 'PROFILE_IMPORT', ?, ?, ?, ?)", (current["id"], json.dumps(before, ensure_ascii=False, default=str), json.dumps(after, ensure_ascii=False, default=str), actor_user_id, now))
            if center_id and not current.get("development_org_unit_id"):
                relation_count += int(_upsert_relation(connection, current["id"], center_id, "DEVELOPMENT_RELATION", now))
            if item["relation_ids"].get("class_id"):
                relation_type = "SPECIAL_COHORT" if item["relation_ids"].get("class_type") == "SPECIAL_COHORT" else "STUDY_CLASS"
                relation_count += int(_upsert_relation(connection, current["id"], item["relation_ids"]["class_id"], relation_type, now))
            if item["relation_ids"].get("group_id"):
                relation_count += int(_upsert_relation(connection, current["id"], item["relation_ids"]["group_id"], "STUDY_GROUP", now))
            if changes or item["relation_ids"].get("class_id") or item["relation_ids"].get("group_id"):
                write_audit(connection, actor_user_id=actor_user_id, action="members.roster_import.update", resource_type="member", resource_id=str(current["id"]), org_unit_id=current["org_unit_id"], after={"source_batch_id": batch_id, "changed_fields": sorted(_member_fields_for_response(changes))})
        execute(connection, "UPDATE import_batches SET status='APPLIED', applied_at=? WHERE id=?", (now, batch_id))
        write_audit(connection, actor_user_id=actor_user_id, action="members.roster_import.apply", resource_type="import_batch", resource_id=str(batch_id), after={"source_sha256": source_sha256, "created": created, "updated": updated, "fields": field_count, "relations": relation_count})
    return {
        "batch_id": batch_id,
        "status": "APPLIED",
        "created": created,
        "updated": updated,
        "fields": field_count,
        "relations": relation_count,
        "annual_sales_applied": annual_sales_applied,
        "skipped_manual_review": plan["matching"]["manual_review_count"],
        "source_sha256": source_sha256,
    }
