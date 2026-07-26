from __future__ import annotations

import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.db import execute, fetch_one, transaction
from app.services.audit import write_audit


CENTER_SHEETS = {
    "园区MP": ("org-yuanqu", "YQ", "园区分中心"),
    "姑苏相城MP": ("org-gusu", "GSXC", "姑苏相城分中心"),
    "昆山MP": ("org-kunshan", "KS", "昆山分中心"),
    "新吴MP": ("org-xinwu", "XW", "新吴分中心"),
    "吴江MP": ("org-wujiang", "WJ", "吴江分中心"),
    "张家港MP": ("org-zhangjiagang", "ZJG", "张家港分中心"),
}
METRICS = [
    ("reading_checkin_rate", "读书打卡率", "学习践行", "PERCENT", "WEIGHTED_AVG", "AUTO"),
    ("reflection_share_rate", "检视分享率", "学习践行", "PERCENT", "AVG", "MIXED"),
    ("credits_per_member", "人均学分", "学习践行", "SCORE", "AVG", "MIXED"),
    ("teaching_goal_rate", "教学目标达成率", "学习践行", "PERCENT", "AVG", "MANUAL"),
    ("class_meeting_rate", "班会参与率", "运营管理", "PERCENT", "WEIGHTED_AVG", "AUTO"),
    ("group_meeting_rate", "小组学习会参与率", "运营管理", "PERCENT", "WEIGHTED_AVG", "AUTO"),
    ("staff_training_rate", "班主任辅导员培训会参与率", "运营管理", "PERCENT", "WEIGHTED_AVG", "AUTO"),
    ("volunteer_points_per_member", "人均志工分", "运营管理", "SCORE", "AVG", "MIXED"),
    ("active_member_count", "在册塾生数", "发展成果", "PERSON", "ENDING_BALANCE", "MIXED"),
    ("new_member_count", "新增学员数", "发展成果", "PERSON", "SUM", "MIXED"),
    ("renewal_rate", "续费率", "发展成果", "PERCENT", "AVG", "MIXED"),
    ("new_member_avg_score", "新增学员平均分数", "发展成果", "SCORE", "AVG", "MANUAL"),
    ("board_attendance_rate", "理事会参与率", "监督管理", "PERCENT", "WEIGHTED_AVG", "AUTO"),
    ("profit_margin", "利润率", "监督管理", "PERCENT", "AVG", "MANUAL"),
]
METRIC_BY_NAME = {item[1]: item for item in METRICS}
CENTER_TARGET_AGGREGATION = {
    "active_member_count": "SUM",
    "new_member_count": "SUM",
    "renewal_rate": "AVG",
}
MONTH_START_COLUMN = {month: 8 + (month - 1) * 4 for month in range(1, 13)}


def _state(value: Any, month: int | None = None) -> tuple[float | None, str]:
    if value == "/":
        return None, "NOT_APPLICABLE"
    if value in (None, ""):
        return None, "NO_DATA"
    if isinstance(value, bool):
        return None, "NO_DATA"
    numeric = float(value)
    return numeric, "ZERO_IS_VALID" if numeric == 0 else "VALUE"


def _source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def preview_workbook(path: str | Path, months: range | list[int] = range(1, 6)) -> dict[str, Any]:
    source = Path(path)
    workbook = load_workbook(source, data_only=True, read_only=True)
    selected_months = sorted({int(month) for month in months if 1 <= int(month) <= 12})
    rows: list[dict[str, Any]] = []
    issues: list[dict[str, str]] = []
    center_targets: dict[str, dict[str, float]] = {}
    try:
        sheet_map = {name.strip(): name for name in workbook.sheetnames}
        missing = [sheet for sheet in [*CENTER_SHEETS, "苏州塾MP"] if sheet not in sheet_map]
        if missing:
            raise ValueError(f"工作簿缺少工作表: {', '.join(missing)}")
        for sheet_name, (org_id, _, center_name) in CENTER_SHEETS.items():
            sheet = workbook[sheet_map[sheet_name]]
            center_targets[org_id] = {}
            for row_no in range(8, 22):
                metric_name = str(sheet.cell(row_no, 3).value or "").strip()
                metric = METRIC_BY_NAME.get(metric_name)
                if not metric:
                    issues.append({
                        "code": "UNKNOWN_METRIC",
                        "severity": "ERROR",
                        "location": f"{sheet_name}!C{row_no}",
                        "message": f"未识别指标：{metric_name}",
                    })
                    continue
                metric_key = metric[0]
                annual, annual_state = _state(sheet.cell(row_no, 5).value)
                rows.append({
                    "org_unit_id": org_id,
                    "org_name": center_name,
                    "metric_key": metric_key,
                    "record_type": "ANNUAL_TARGET",
                    "period_no": 1,
                    "value_kind": "MP",
                    "numeric_value": annual,
                    "value_state": annual_state,
                    "source_reference": f"{source.name}|{sheet_name}|E{row_no}",
                })
                if annual is not None:
                    center_targets[org_id][metric_key] = annual
                for month in selected_months:
                    start = MONTH_START_COLUMN[month]
                    for offset, kind in ((0, "MP"), (1, "FORECAST"), (2, "ACTUAL")):
                        value, value_state = _state(sheet.cell(row_no, start + offset).value, month)
                        rows.append({
                            "org_unit_id": org_id,
                            "org_name": center_name,
                            "metric_key": metric_key,
                            "record_type": "PERIOD_VALUE",
                            "period_no": month,
                            "value_kind": kind,
                            "numeric_value": value,
                            "value_state": value_state,
                            "source_reference": (
                                f"{source.name}|{sheet_name}|"
                                f"{sheet.cell(row_no, start + offset).coordinate}"
                            ),
                        })
                        if metric_key == "new_member_count" and kind == "ACTUAL" and value is not None and value < 0:
                            issues.append({
                                "code": "NEGATIVE_NEW_MEMBER",
                                "severity": "WARNING",
                                "location": f"{sheet_name}!{sheet.cell(row_no, start + offset).coordinate}",
                                "message": "新增学员实绩为负数；按原值保留，需确认净增/冲销口径。",
                            })

        root = workbook[sheet_map["苏州塾MP"]]
        root_targets = {
            "active_member_count": float(root.cell(4, 3).value),
            "new_member_count": float(root.cell(4, 4).value),
            "renewal_rate": float(root.cell(4, 7).value),
        }
        for metric_key, value in root_targets.items():
            rows.append({
                "org_unit_id": "org-suzhou",
                "org_name": "苏州塾",
                "metric_key": metric_key,
                "record_type": "ANNUAL_TARGET",
                "period_no": 1,
                "value_kind": "MP",
                "numeric_value": value,
                "value_state": "ZERO_IS_VALID" if value == 0 else "VALUE",
                "source_reference": f"{source.name}|苏州塾MP|"
                f"{'C4' if metric_key == 'active_member_count' else 'D4' if metric_key == 'new_member_count' else 'G4'}",
            })

        reconciliation = []
        for metric_key, root_value in root_targets.items():
            child_values = [
                targets[metric_key] for targets in center_targets.values() if metric_key in targets
            ]
            method = CENTER_TARGET_AGGREGATION[metric_key]
            child_total = (
                sum(child_values)
                if method == "SUM"
                else sum(child_values) / len(child_values)
            )
            difference = root_value - child_total
            reconciliation.append({
                "metric_key": metric_key,
                "aggregation": method,
                "root_target": root_value,
                "child_target_total": child_total,
                "difference": difference,
                "requires_reason": abs(difference) > 1e-9,
            })
            if abs(difference) > 1e-9:
                issues.append({
                    "code": "TARGET_VARIANCE",
                    "severity": "WARNING",
                    "location": "苏州塾MP",
                    "message": f"{metric_key} 总目标与六中心分解相差 {difference:g}，必须保留差额原因。",
                })
        return {
            "source_name": source.name,
            "source_sha256": _source_hash(source),
            "sheet_names": workbook.sheetnames,
            "months": selected_months,
            "rows": rows,
            "issues": issues,
            "reconciliation": reconciliation,
            "summary": {
                "row_count": len(rows),
                "issue_count": len(issues),
                "center_count": len(CENTER_SHEETS),
                "metric_count": len(METRICS),
            },
        }
    finally:
        workbook.close()


def save_preview(preview: dict[str, Any], actor_user_id: int | None) -> int:
    now = datetime.now(UTC).isoformat()
    with transaction() as connection:
        cursor = execute(
            connection,
            "INSERT INTO import_batches(import_type, source_name, source_sha256, status, preview_json, "
            "created_by, created_at) VALUES ('ANNUAL_MP', ?, ?, 'PREVIEWED', ?, ?, ?)",
            (
                preview["source_name"],
                preview["source_sha256"],
                json.dumps(preview, ensure_ascii=False),
                actor_user_id,
                now,
            ),
        )
        batch_id = cursor.lastrowid
        for issue in preview["issues"]:
            execute(
                connection,
                "INSERT INTO data_quality_issues(batch_id, issue_code, severity, location, message, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (
                    batch_id, issue["code"], issue["severity"], issue["location"], issue["message"], now
                ),
            )
        return batch_id


def _upsert_orgs(connection, now: str) -> None:
    orgs = [
        ("org-suzhou", "SZ_ROOT", "苏州塾", "ROOT", None),
        *[
            (org_id, code, name, "REGIONAL_CENTER", "org-suzhou")
            for org_id, code, name in CENTER_SHEETS.values()
        ],
        ("org-huangpu", "HP_CLASS", "黄埔班", "CLASS", "org-suzhou"),
    ]
    for org in orgs:
        existing = execute(connection, "SELECT id FROM org_units WHERE id=?", (org[0],)).fetchone()
        if not existing:
            execute(
                connection,
                "INSERT INTO org_units(id, unit_code, name, unit_type, parent_id, is_active, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, 1, ?, ?)",
                org + (now, now),
            )


def _upsert_target(connection, *, plan_id: int, version_id: int, row: dict, now: str) -> None:
    existing = execute(
        connection,
        "SELECT id FROM org_metric_targets WHERE annual_plan_id=? AND org_unit_id=? "
        "AND metric_version_id=?",
        (plan_id, row["org_unit_id"], version_id),
    ).fetchone()
    if existing:
        target_id = existing["id"] if hasattr(existing, "keys") else existing[0]
        execute(
            connection,
            "UPDATE org_metric_targets SET annual_target=?, value_state=?, balance_mode='ALLOW_VARIANCE', "
            "source_reference=?, updated_at=? WHERE id=?",
            (
                row["numeric_value"], row["value_state"], row["source_reference"], now, target_id
            ),
        )
    else:
        execute(
            connection,
            "INSERT INTO org_metric_targets(annual_plan_id, org_unit_id, metric_version_id, annual_target, "
            "value_state, balance_mode, source_reference, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, 'ALLOW_VARIANCE', ?, ?, ?)",
            (
                plan_id, row["org_unit_id"], version_id, row["numeric_value"],
                row["value_state"], row["source_reference"], now, now,
            ),
        )


def _upsert_period(
    connection, *, plan_id: int, version_id: int, row: dict, actor_user_id: int, batch_id: int, now: str
) -> None:
    existing = execute(
        connection,
        "SELECT id FROM metric_period_values WHERE annual_plan_id=? AND org_unit_id=? "
        "AND metric_version_id=? AND period_type='MONTH' AND period_no=? AND value_kind=?",
        (plan_id, row["org_unit_id"], version_id, row["period_no"], row["value_kind"]),
    ).fetchone()
    detail = json.dumps({"batch_id": batch_id}, ensure_ascii=False)
    if existing:
        value_id = existing["id"] if hasattr(existing, "keys") else existing[0]
        execute(
            connection,
            "UPDATE metric_period_values SET numeric_value=?, value_state=?, source_type='IMPORT', "
            "source_reference=?, calculation_detail_json=?, is_manual_override=0, updated_by=?, "
            "updated_at=? WHERE id=?",
            (
                row["numeric_value"], row["value_state"], row["source_reference"], detail,
                actor_user_id, now, value_id,
            ),
        )
    else:
        execute(
            connection,
            "INSERT INTO metric_period_values(annual_plan_id, org_unit_id, metric_version_id, period_type, "
            "period_no, value_kind, numeric_value, value_state, source_type, source_reference, "
            "calculation_detail_json, is_manual_override, updated_by, updated_at) "
            "VALUES (?, ?, ?, 'MONTH', ?, ?, ?, ?, 'IMPORT', ?, ?, 0, ?, ?)",
            (
                plan_id, row["org_unit_id"], version_id, row["period_no"], row["value_kind"],
                row["numeric_value"], row["value_state"], row["source_reference"], detail,
                actor_user_id, now,
            ),
        )


def apply_preview(batch_id: int, actor_user_id: int) -> dict[str, Any]:
    batch = fetch_one("SELECT * FROM import_batches WHERE id=?", (batch_id,))
    if not batch or batch["status"] != "PREVIEWED":
        raise ValueError("导入批次不存在或已处理")
    preview = json.loads(batch["preview_json"])
    if any(issue["severity"] == "ERROR" for issue in preview["issues"]):
        raise ValueError("预览含阻断错误，不能应用")
    now = datetime.now(UTC).isoformat()
    with transaction() as connection:
        _upsert_orgs(connection, now)
        plan = execute(
            connection, "SELECT id FROM annual_plans WHERE year=2026 AND version=1"
        ).fetchone()
        if plan:
            plan_id = plan["id"] if hasattr(plan, "keys") else plan[0]
        else:
            plan_id = execute(
                connection,
                "INSERT INTO annual_plans(year, version, policy_text, status, write_enabled, created_at, updated_at) "
                "VALUES (2026, 1, '以人为中心，以数据发现问题，以服务促进成长', 'DRAFT', 0, ?, ?)",
                (now, now),
            ).lastrowid
        version_ids: dict[str, int] = {}
        for order, metric in enumerate(METRICS, 1):
            key, name, category, unit, aggregation, source_type = metric
            definition = execute(
                connection, "SELECT id FROM metric_definitions WHERE metric_key=?", (key,)
            ).fetchone()
            if definition:
                definition_id = definition["id"] if hasattr(definition, "keys") else definition[0]
            else:
                definition_id = execute(
                    connection,
                    "INSERT INTO metric_definitions(metric_key, name, category, default_unit, created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (key, name, category, unit, now, now),
                ).lastrowid
            version = execute(
                connection,
                "SELECT id FROM metric_versions WHERE metric_definition_id=? AND year=2026 AND version=1",
                (definition_id,),
            ).fetchone()
            if version:
                version_id = version["id"] if hasattr(version, "keys") else version[0]
            else:
                period_type = "ENDING_BALANCE" if aggregation == "ENDING_BALANCE" else "INCREMENT"
                version_id = execute(
                    connection,
                    "INSERT INTO metric_versions(metric_definition_id, year, version, aggregation_type, "
                    "period_value_type, unit, data_source_type, null_policy, formula_text, status, created_at) "
                    "VALUES (?, 2026, 1, ?, ?, ?, ?, 'EXPLICIT_STATE', ?, 'DRAFT', ?)",
                    (
                        definition_id, aggregation, period_type, unit, source_type,
                        "人工填写值，不由系统推导" if key == "profit_margin" else None, now,
                    ),
                ).lastrowid
            version_ids[key] = version_id
            exists = execute(
                connection,
                "SELECT id FROM plan_metrics WHERE annual_plan_id=? AND metric_version_id=?",
                (plan_id, version_id),
            ).fetchone()
            if not exists:
                execute(
                    connection,
                    "INSERT INTO plan_metrics(annual_plan_id, metric_version_id, display_order, applicable_unit_types) "
                    "VALUES (?, ?, ?, 'ROOT,REGIONAL_CENTER')",
                    (plan_id, version_id, order),
                )

        for row in preview["rows"]:
            version_id = version_ids[row["metric_key"]]
            if row["record_type"] == "ANNUAL_TARGET":
                _upsert_target(
                    connection, plan_id=plan_id, version_id=version_id, row=row, now=now
                )
            else:
                _upsert_period(
                    connection,
                    plan_id=plan_id,
                    version_id=version_id,
                    row=row,
                    actor_user_id=actor_user_id,
                    batch_id=batch_id,
                    now=now,
                )
        execute(
            connection,
            "UPDATE import_batches SET status='APPLIED_READ_ONLY', applied_at=? WHERE id=?",
            (now, batch_id),
        )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="imports.mp.apply_read_only",
            resource_type="import_batch",
            resource_id=str(batch_id),
            after={"annual_plan_id": plan_id, "write_enabled": False},
        )
    return {
        "annual_plan_id": plan_id,
        "status": "DRAFT",
        "write_enabled": False,
        "imported_rows": len(preview["rows"]),
        "issues": preview["issues"],
        "reconciliation": preview["reconciliation"],
    }
