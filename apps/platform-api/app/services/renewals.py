from __future__ import annotations

import hashlib
import json
import re
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.privacy import encrypt_text, phone_hash
from app.db import execute, fetch_all, fetch_one, transaction
from app.services.audit import write_audit
from app.services.iam import accessible_org_ids, user_context
from app.services.member_memories import verified_member_memories


CENTER_IDS = {
    "姑苏相城分中心": "org-gusu", "昆山分中心": "org-kunshan", "吴江分中心": "org-wujiang",
    "新吴分中心": "org-xinwu", "园区分中心": "org-yuanqu", "张家港分中心": "org-zhangjiagang",
}

MASTER_HEADER_ALIASES = {
    "姓名": ("姓名", "名字", "name"),
    "手机号码": ("手机号码", "手机号", "phone"),
    "所在分中心": ("所在分中心", "所属分中心", "center"),
}
IMPORTABLE_MATCH_STATUSES = frozenset(
    {"MASTER_PHONE_EXACT", "MASTER_NAME_CENTER_EXACT", "MATCHED"}
)
RENEWAL_STATUSES = frozenset(
    {
        "PENDING_FIRST_CONTACT",
        "CONTACTED_WAITING_REPLY",
        "IN_COMMUNICATION",
        "RENEWED",
        "NOT_RENEWING",
        "DEFERRED",
        "EXITED",
    }
)
CLOSED_RENEWAL_STATUSES = frozenset({"RENEWED", "NOT_RENEWING", "EXITED"})
RENEWAL_STAGE_LABELS = {
    "PREPARE": "日常维护",
    "OBSERVE_3": "观3",
    "RENEW_2": "续2",
    "FOLLOW_1": "追1",
    "DUE_NOW": "到期冲刺",
    "RECOVERY": "挽回/复盘",
    "CLOSED": "已闭环",
}
TODAY_ACTION_STAGE_CODES = frozenset(
    {"OBSERVE_3", "RENEW_2", "FOLLOW_1", "DUE_NOW", "RECOVERY"}
)
TODAY_ACTION_REASON_CODES = frozenset(
    {
        "FOLLOWUP_OVERDUE",
        "FOLLOWUP_TODAY",
        "SUPPORT_NEEDED",
        "STAGE_UNTOUCHED",
        "NEXT_STEP_MISSING",
    }
)
TODAY_ACTION_REASON_RANK = {
    "FOLLOWUP_OVERDUE": 0,
    "FOLLOWUP_TODAY": 1,
    "SUPPORT_NEEDED": 2,
    "STAGE_UNTOUCHED": 3,
    "NEXT_STEP_MISSING": 4,
}
TODAY_ACTION_STAGE_RANK = {
    "RECOVERY": 0,
    "DUE_NOW": 1,
    "FOLLOW_1": 2,
    "RENEW_2": 3,
    "OBSERVE_3": 4,
}
PREVIEW_ROW_FIELDS = (
    "row_no",
    "name",
    "center_name",
    "class_name",
    "due_month",
    "match_status",
    "issue_code",
    "proposed_status",
    "history_note",
    "assistance_note",
)

# Renewal attribution is a member-management fact.  A member's development
# relation is the authoritative center for renewal and regional reporting;
# the primary member relation is the safe fallback for older profiles that do
# not yet have a development relation.  The workbook's center is only used
# during the one-time matching/import flow and is never used for daily reads.
MEMBER_RENEWAL_ORG_SQL = (
    "COALESCE(NULLIF(m.development_org_unit_id, ''), m.org_unit_id)"
)
MEMBER_CLASS_NAME_SQL = (
    "(SELECT ou.name FROM member_org_relations mor "
    "JOIN org_units ou ON ou.id=mor.org_unit_id "
    "WHERE mor.member_id=m.id AND mor.relation_type IN ('STUDY_CLASS','SPECIAL_COHORT') "
    "AND mor.valid_until IS NULL AND ou.is_active=1 "
    "ORDER BY mor.is_primary DESC, mor.id DESC LIMIT 1)"
)
MEMBER_GROUP_NAME_SQL = (
    "(SELECT ou.name FROM member_org_relations mor "
    "JOIN org_units ou ON ou.id=mor.org_unit_id "
    "WHERE mor.member_id=m.id AND mor.relation_type='STUDY_GROUP' "
    "AND mor.valid_until IS NULL AND ou.is_active=1 "
    "ORDER BY mor.is_primary DESC, mor.id DESC LIMIT 1)"
)


def _cycle_with_member_scope(cycle_id: int) -> dict[str, Any] | None:
    return fetch_one(
        "SELECT c.*, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS member_org_unit_id "
        "FROM renewal_cycles c JOIN members m ON m.id=c.member_id WHERE c.id=?",
        (cycle_id,),
    )


def list_assignees(actor_user_id: int, org_unit_id: str | None = None) -> list[dict[str, Any]]:
    """Return active users who can manage renewals in the requested org scope."""
    actor_allowed = accessible_org_ids(actor_user_id)
    if org_unit_id and actor_allowed is not None and org_unit_id not in actor_allowed:
        raise PermissionError("组织不在当前用户授权范围内")
    rows = fetch_all(
        "SELECT id, username, display_name FROM app_users WHERE is_active=1 ORDER BY display_name, id"
    )
    result: list[dict[str, Any]] = []
    for row in rows:
        context = user_context(row["id"])
        if not context or "renewals:manage" not in context["permissions"]:
            continue
        if org_unit_id:
            assignee_allowed = accessible_org_ids(row["id"])
            if assignee_allowed is not None and org_unit_id not in assignee_allowed:
                continue
        result.append(row)
    return result


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _month(value: Any) -> int | None:
    text = str(value or "").strip().replace("月", "")
    return int(text) if text.isdigit() and 1 <= int(text) <= 12 else None


def _member_renewal_month(value: Any) -> int | None:
    """Return the recurring month maintained in the member master profile."""
    match = re.fullmatch(r"\d{4}-(\d{2})", str(value or "").strip())
    if not match:
        return None
    month = int(match.group(1))
    return month if 1 <= month <= 12 else None


def _initial_cycle_status(due_month: int, renewal_year: int, now: datetime) -> str:
    """Treat months before the current month as already renewed for this year."""
    if renewal_year == now.year and due_month < now.month:
        return "RENEWED"
    return "PENDING_FIRST_CONTACT"


def determine_calendar_stage(
    renewal_year: int,
    due_month: int,
    *,
    as_of: date | datetime | None = None,
) -> dict[str, Any]:
    """Determine the time-based renewal stage without considering terminal status."""
    if not 1 <= int(due_month) <= 12:
        raise ValueError("续费月份必须在1至12之间")
    current = as_of or datetime.now(UTC)
    current_date = current.date() if isinstance(current, datetime) else current
    months_until_due = (
        int(renewal_year) * 12 + int(due_month)
        - (current_date.year * 12 + current_date.month)
    )
    if months_until_due > 3:
        code = "PREPARE"
    elif months_until_due == 3:
        code = "OBSERVE_3"
    elif months_until_due == 2:
        code = "RENEW_2"
    elif months_until_due == 1:
        code = "FOLLOW_1"
    elif months_until_due == 0:
        code = "DUE_NOW"
    else:
        code = "RECOVERY"
    return {
        "code": code,
        "label": RENEWAL_STAGE_LABELS[code],
        "months_until_due": months_until_due,
        "as_of_month": f"{current_date.year:04d}-{current_date.month:02d}",
        "source": "CALENDAR_RULE",
    }


def determine_renewal_stage(
    renewal_year: int,
    due_month: int,
    status: str,
    *,
    as_of: date | datetime | None = None,
) -> dict[str, Any]:
    """Determine the operational phase from calendar months and cycle status."""
    normalised_status = str(status or "").upper()
    if normalised_status in CLOSED_RENEWAL_STATUSES:
        current = as_of or datetime.now(UTC)
        current_date = current.date() if isinstance(current, datetime) else current
        calendar = determine_calendar_stage(renewal_year, due_month, as_of=current_date)
        return {
            **calendar,
            "code": "CLOSED",
            "label": RENEWAL_STAGE_LABELS["CLOSED"],
        }
    return determine_calendar_stage(renewal_year, due_month, as_of=as_of)


def _completed_membership_years(
    join_date: Any,
    stored_years: Any,
    overridden: Any,
    *,
    as_of: date,
) -> int | None:
    if overridden and stored_years is not None:
        try:
            return max(0, int(float(stored_years)))
        except (TypeError, ValueError):
            return None
    if not join_date:
        return None
    try:
        joined = date.fromisoformat(str(join_date)[:10])
    except ValueError:
        return None
    years = as_of.year - joined.year
    if (as_of.month, as_of.day) < (joined.month, joined.day):
        years -= 1
    return max(0, years)


_PHONE_IN_TEXT_RE = re.compile(r"(?<!\d)1[3-9](?:[\s-]?\d){9}(?!\d)")
_PRECISE_AMOUNT_RE = re.compile(
    r"(?<!\d)(?:人民币\s*|RMB\s*|[￥¥]\s*)?"
    r"(?:\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?|"
    r"[零〇一二三四五六七八九十百千万亿两壹贰叁肆伍陆柒捌玖拾佰仟萬]+)"
    r"\s*(?:亿元|万元|元|亿|万)(?![\d元])",
    re.IGNORECASE,
)


def _redact_action_text(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = _PHONE_IN_TEXT_RE.sub("[手机号已脱敏]", text)
    return _PRECISE_AMOUNT_RE.sub("[明确金额已脱敏]", text)


def _calendar_date(value: Any) -> date | None:
    """Read the calendar date prefix used by follow-up date fields."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _renewal_action_strategy(
    stage_code: str,
    *,
    latest_followup: dict[str, Any] | None,
) -> dict[str, Any]:
    strategies = {
        "PREPARE": {
            "goal": "保持日常陪伴，不进入续费动作",
            "channel": "WECHAT",
            "reason": "距离续费超过3个月，适合自然保持联系，不提前制造续费压力。",
            "questions": ["最近学习和生活状态怎么样？", "近期有没有我们可以提供支持的地方？"],
            "do_not": ["不以续费为本次联系主题。", "不发送付款信息。"],
        },
        "OBSERVE_3": {
            "goal": "重新建立连接、了解近况",
            "channel": "WECHAT",
            "reason": "距离续费还有3个月，当前属于观3阶段，应先关爱和了解，不以完成续费为目标。",
            "questions": [
                "最近企业和生活状态怎么样？",
                "今年学习过程中，哪些内容对经营帮助比较大？",
                "最近有没有什么我们能提供支持的地方？",
            ],
            "do_not": ["不直接询问是否续费。", "不直接发送付款信息。"],
        },
        "RENEW_2": {
            "goal": "回顾同行价值、自然确认续费意愿",
            "channel": "PHONE" if latest_followup else "WECHAT",
            "reason": "距离续费还有2个月，适合先回顾真实获得，再自然了解后续同行意愿。",
            "questions": [
                "今年哪些学习或同行经历对您最有帮助？",
                "接下来的学习中，您最希望获得哪方面支持？",
                "关于继续同行，您现在还有哪些考虑？",
            ],
            "do_not": ["不跳过价值回顾直接催促决定。", "不把未确认意愿表述成已承诺。"],
        },
        "FOLLOW_1": {
            "goal": "明确意向、识别障碍并协助解决",
            "channel": "PHONE",
            "reason": "距离续费还有1个月，需要基于既有沟通明确意向和实际障碍。",
            "questions": [
                "关于下一年度继续同行，您目前的考虑是什么？",
                "现在最需要我们协助解决的障碍是什么？",
                "下一次确认安排在什么时间比较合适？",
            ],
            "do_not": ["不反复施压。", "不替学长推断困难或作出决定。"],
        },
        "DUE_NOW": {
            "goal": "责任到人，确认决定与下一步安排",
            "channel": "PHONE",
            "reason": "已进入续费月份，应由责任人确认当前决定、障碍和具体下一步。",
            "questions": [
                "目前关于继续同行的决定是否已经明确？",
                "还有什么事项需要我们协调支持？",
                "下一步由谁在什么时间跟进最合适？",
            ],
            "do_not": ["不在没有确认的情况下代替学长做决定。", "不遗漏责任人和下一步时间。"],
        },
        "RECOVERY": {
            "goal": "确认延期、不续、困难或失联原因并完成复盘",
            "channel": "PHONE",
            "reason": "续费月份已过且周期未闭环，需要尊重事实地确认结果和后续关系安排。",
            "questions": [
                "目前未完成续费的主要原因是什么？",
                "是希望延期沟通，还是已经有明确决定？",
                "后续我们以什么方式保持联系最合适？",
            ],
            "do_not": ["不把所有未完成都归因于价格。", "不在明确退出后继续重复打扰。"],
        },
        "CLOSED": {
            "goal": "保持正常关系，不重复发起续费动作",
            "channel": "NONE",
            "reason": "当前续费周期已闭环，不应再次触发本周期续费联系。",
            "questions": [],
            "do_not": ["不重复发起本周期续费提醒。"],
        },
    }
    strategy = dict(strategies[stage_code])
    if latest_followup and latest_followup.get("needs_support") and stage_code not in {
        "PREPARE",
        "OBSERVE_3",
        "CLOSED",
    }:
        strategy["channel"] = "MEETING"
        strategy["reason"] += " 最近一次跟进标记需要协助，建议责任人与负责人协同。"
        strategy["coordination_recommended"] = True
    else:
        strategy["coordination_recommended"] = False
    return strategy


def maybe_create_historical_cycle(
    connection: Any,
    *,
    member_id: int,
    actor_user_id: int,
    member_status: str,
    renewal_month: Any,
    org_unit_id: str,
    renewal_year: int | None = None,
    now: datetime | None = None,
) -> int | None:
    """Reconcile one audited historical cycle after member maintenance.

    This is deliberately a single-member helper. It runs inside the member
    transaction, so maintaining a past renewal month automatically closes the
    corresponding current-year cycle without introducing a bulk write on a
    read-only coverage request.
    """
    if str(member_status or "").upper() != "ACTIVE":
        return None
    due_month = _member_renewal_month(renewal_month)
    current = now or datetime.now(UTC)
    target_year = renewal_year or current.year
    if not due_month or _initial_cycle_status(due_month, target_year, current) != "RENEWED":
        return None
    existing = execute(
        connection,
        "SELECT id, status, due_month, completed_at FROM renewal_cycles "
        "WHERE member_id=? AND renewal_year=?",
        (member_id, target_year),
    ).fetchone()
    if existing:
        existing_status = str(existing["status"] or "").upper()
        if existing_status == "RENEWED":
            return int(existing["id"])
        # Do not overwrite an explicit negative, paused, or exited decision.
        # Open follow-up states are safe to close because maintaining a past
        # renewal month is the operator's explicit confirmation that this
        # year's renewal has already happened.
        if existing_status not in {
            "PENDING_FIRST_CONTACT",
            "CONTACTED_WAITING_REPLY",
            "IN_COMMUNICATION",
        }:
            return None
        completed_at = current.isoformat()
        execute(
            connection,
            "UPDATE renewal_cycles SET org_unit_id=?, due_month=?, status='RENEWED', "
            "completed_at=?, updated_at=? WHERE id=?",
            (org_unit_id, due_month, completed_at, completed_at, existing["id"]),
        )
        execute(
            connection,
            "INSERT INTO renewal_status_history(renewal_cycle_id, from_status, to_status, "
            "reason, changed_by, created_at) VALUES (?, ?, 'RENEWED', ?, ?, ?)",
            (
                existing["id"],
                existing_status,
                "学员管理维护历史续费月份，已有周期自动标记为已续费",
                actor_user_id,
                completed_at,
            ),
        )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.cycle.auto_complete_historical",
            resource_type="renewal_cycle",
            resource_id=str(existing["id"]),
            org_unit_id=org_unit_id,
            purpose="学员管理维护历史续费月份，自动完成已有当前年度周期",
            before={
                "status": existing_status,
                "due_month": existing["due_month"],
                "completed_at": existing["completed_at"],
            },
            after={
                "status": "RENEWED",
                "due_month": due_month,
                "completed_at": completed_at,
                "source": "member_renewal_month_maintenance",
            },
        )
        return int(existing["id"])
    created_at = current.isoformat()
    cycle_id = execute(
        connection,
        "INSERT INTO renewal_cycles(member_id, renewal_year, org_unit_id, due_month, "
        "status, completed_at, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (
            member_id,
            target_year,
            org_unit_id,
            due_month,
            "RENEWED",
            created_at,
            created_at,
            created_at,
        ),
    ).lastrowid
    execute(
        connection,
        "INSERT INTO renewal_status_history(renewal_cycle_id, from_status, to_status, "
        "reason, changed_by, created_at) VALUES (?, NULL, 'RENEWED', ?, ?, ?)",
        (
            cycle_id,
            "学员管理维护续费月份，历史月份自动标记为已续费",
            actor_user_id,
            created_at,
        ),
    )
    write_audit(
        connection,
        actor_user_id=actor_user_id,
        action="renewals.cycle.auto_create_historical",
        resource_type="renewal_cycle",
        resource_id=str(cycle_id),
        org_unit_id=org_unit_id,
        purpose="学员管理维护历史续费月份，自动补齐当前年度已续费周期",
        after={
            "member_id": member_id,
            "renewal_year": target_year,
            "due_month": due_month,
            "status": "RENEWED",
            "source": "member_renewal_month_maintenance",
        },
    )
    return int(cycle_id)


def _status(note: Any) -> str:
    text = str(note or "").strip()
    if not text:
        return "PENDING_FIRST_CONTACT"
    if "已续费" in text:
        return "RENEWED"
    if "退出" in text:
        return "EXITED"
    if "不续费" in text:
        return "NOT_RENEWING"
    if "暂停" in text or "休学" in text:
        return "DEFERRED"
    if "未接" in text or "未回复" in text:
        return "CONTACTED_WAITING_REPLY"
    if "邀请" in text or "提醒" in text:
        return "CONTACTED_WAITING_REPLY"
    return "IN_COMMUNICATION"


def _clean(value: Any) -> str:
    return str(value or "").strip().replace("\n", "").replace(" ", "")


def _master_columns(headers: list[str]) -> dict[str, int]:
    normalized = {_clean(name).lower(): index for index, name in enumerate(headers)}
    columns: dict[str, int] = {}
    for canonical, aliases in MASTER_HEADER_ALIASES.items():
        for alias in aliases:
            index = normalized.get(_clean(alias).lower())
            if index is not None:
                columns[canonical] = index
                break
    return columns


def _phone(value: Any) -> str:
    return "".join(char for char in _clean(value) if char.isdigit())[-11:]


def _master_index(path: Path) -> tuple[dict[str, list[dict]], dict[tuple[str, str], list[dict]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "2026 新在册表" not in workbook.sheetnames:
            raise ValueError("主档案缺少“2026 新在册表”工作表")
        sheet = workbook["2026 新在册表"]
        headers = [_clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        col = _master_columns(headers)
        required = {"姓名", "手机号码", "所在分中心"}
        if not required.issubset(col):
            missing = "、".join(sorted(required - set(col)))
            raise ValueError(f"主档案缺少必要列：{missing}")
        by_phone: dict[str, list[dict]] = {}; by_name_center: dict[tuple[str, str], list[dict]] = {}
        for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            record = {
                "姓名": values[col["姓名"]],
                "手机号码": values[col["手机号码"]],
                "所在分中心": values[col["所在分中心"]],
                "source_row": row_no,
            }
            phone = _phone(record.get("手机号码")); key = (_clean(record.get("姓名")), _clean(record.get("所在分中心")))
            if phone: by_phone.setdefault(phone, []).append(record)
            if all(key): by_name_center.setdefault(key, []).append(record)
        return by_phone, by_name_center
    finally:
        workbook.close()


def _linked_member_id(
    phone: str,
    name: str,
    org_id: str | None,
    by_phone_hash: dict[str, list[dict[str, Any]]],
    by_name_org: dict[tuple[str, str], list[dict[str, Any]]],
) -> int | None:
    """Resolve a preview row to one existing production member, if unique."""
    if phone:
        try:
            candidates = by_phone_hash.get(phone_hash(phone), [])
        except ValueError:
            candidates = []
        if len(candidates) == 1:
            return int(candidates[0]["id"])
        if org_id:
            scoped = [row for row in candidates if row["org_unit_id"] == org_id]
            if len(scoped) == 1:
                return int(scoped[0]["id"])
    if name and org_id:
        candidates = by_name_org.get((name, org_id), [])
        if len(candidates) == 1:
            return int(candidates[0]["id"])
    return None


def preview_workbook(path: str | Path, master_path: str | Path | None = None) -> dict[str, Any]:
    source = Path(path)
    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        if "2026年续费基数" not in workbook.sheetnames:
            raise ValueError("未找到“2026年续费基数”工作表")
        sheet = workbook["2026年续费基数"]
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required = {"名字", "所在分中心", "2025年缴费月份"}
        if not required.issubset(headers):
            raise ValueError("续费基数表缺少姓名、分中心或2025年缴费月份列")
        col = {name: index for index, name in enumerate(headers)}
        master_by_phone, master_by_name_center = _master_index(Path(master_path)) if master_path else ({}, {})
        existing = fetch_all("SELECT id, name, org_unit_id, phone_hash FROM members")
        by_phone_hash: dict[str, list[dict[str, Any]]] = {}
        by_name_org: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in existing:
            if row["phone_hash"]:
                by_phone_hash.setdefault(row["phone_hash"], []).append(row)
            by_name_org.setdefault((row["name"], row["org_unit_id"]), []).append(row)
        rows: list[dict[str, Any]] = []
        summary = {
            "total": 0,
            "matched": 0,
            "needs_review": 0,
            "invalid": 0,
            "assistance_review": 0,
            "production_linked": 0,
            "production_unlinked": 0,
            "importable": 0,
        }
        for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            name = str(values[col["名字"]] or "").strip()
            center_name = str(values[col["所在分中心"]] or "").strip()
            class_name = str(values[col["所属班级"]] or "").strip() if "所属班级" in col else ""
            # 先锋班、黄埔班仅为直属学习班级；续费发展仍按六大分中心统计。
            org_id, reporting_name = CENTER_IDS.get(center_name), center_name
            due_month = _month(values[col["2025年缴费月份"]])
            note = str(values[col["事务所跟进续费情况"]] or "").strip() if "事务所跟进续费情况" in col else ""
            assistance = str(values[col["需要协助"]] or "").strip() if "需要协助" in col else ""
            raw = {headers[i]: values[i] for i in range(len(headers)) if values[i] not in (None, "")}
            summary["total"] += 1
            org_id = CENTER_IDS.get(center_name)
            source_phone = _phone(values[col["手机号码"]]) if "手机号码" in col else ""
            master_phone = master_by_phone.get(source_phone, []) if source_phone else []
            master_name = master_by_name_center.get((_clean(name), _clean(center_name)), []) if name and center_name else []
            existing_members = by_name_org.get((name, org_id), []) if name and org_id else []
            member = existing_members[0] if len(existing_members) == 1 else None
            linked_member_id = _linked_member_id(
                source_phone, _clean(name), org_id, by_phone_hash, by_name_org
            )
            if not name or not org_id or not due_month:
                match_status, issue = "INVALID", "MISSING_REQUIRED_FIELD"
                summary["invalid"] += 1
            elif len(master_phone) == 1:
                match_status, issue, member = "MASTER_PHONE_EXACT", None, master_phone[0]
                summary["matched"] += 1
            elif len(master_phone) > 1:
                match_status, issue = "NEEDS_REVIEW", "MASTER_PHONE_DUPLICATE"
                summary["needs_review"] += 1
            elif len(master_name) == 1:
                match_status, issue, member = "MASTER_NAME_CENTER_EXACT", None, master_name[0]
                summary["matched"] += 1
            elif len(master_name) > 1:
                match_status, issue = "NEEDS_REVIEW", "MASTER_NAME_CENTER_DUPLICATE"
                summary["needs_review"] += 1
            elif member:
                match_status, issue = "MATCHED", None
                summary["matched"] += 1
            else:
                match_status, issue = "NEEDS_REVIEW", "MEMBER_NOT_MATCHED"
                summary["needs_review"] += 1
            if assistance:
                summary["assistance_review"] += 1
            rows.append({"row_no": row_no, "name": name, "org_unit_id": org_id, "center_name": reporting_name, "source_center_name": center_name, "class_name": class_name,
                         "member_id": linked_member_id or (member.get("id") if member else None),
                         "master_source_row": member.get("source_row") if member else None, "due_month": due_month,
                         "match_status": match_status, "issue_code": issue, "proposed_status": _status(note),
                         "history_note": note, "assistance_note": assistance, "raw": raw})
        summary["production_linked"] = sum(
            row["member_id"] is not None for row in rows
        )
        summary["importable"] = sum(
            row["member_id"] is not None
            and row["match_status"] in IMPORTABLE_MATCH_STATUSES
            for row in rows
        )
        summary["production_unlinked"] = sum(
            row["member_id"] is None
            and row["match_status"] in IMPORTABLE_MATCH_STATUSES
            for row in rows
        )
        return {"source_name": source.name, "source_sha256": _hash(source), "summary": summary, "rows": rows}
    finally:
        workbook.close()


def preview_result_view(preview: dict[str, Any]) -> dict[str, Any]:
    """Return the complete review queues without exposing raw workbook rows."""
    rows = [
        {field: row.get(field) for field in PREVIEW_ROW_FIELDS}
        for row in preview["rows"]
    ]
    review_rows = [
        row for row in rows if row["match_status"] in {"NEEDS_REVIEW", "INVALID"}
    ]
    assistance_rows = [row for row in rows if row.get("assistance_note")]
    matched_samples = [
        row for row in rows if row["match_status"] in IMPORTABLE_MATCH_STATUSES
    ][:20]
    issue_summary: dict[str, int] = {}
    for row in review_rows:
        code = row.get("issue_code") or "UNKNOWN"
        issue_summary[code] = issue_summary.get(code, 0) + 1
    return {
        "summary": preview["summary"],
        "review_rows": review_rows,
        "assistance_rows": assistance_rows,
        "matched_samples": matched_samples,
        "issue_summary": issue_summary,
    }


def save_preview(preview: dict[str, Any], actor_user_id: int) -> int:
    now = datetime.now(UTC).isoformat()
    encrypted_preview = encrypt_text(
        json.dumps(preview, ensure_ascii=False, default=str)
    )
    redacted_preview = {
        "redacted": True,
        "source_sha256": preview["source_sha256"],
        "summary": preview["summary"],
        "row_count": len(preview["rows"]),
    }
    with transaction() as connection:
        batch_id = execute(connection, "INSERT INTO renewal_import_batches(source_name, source_sha256, status, preview_json, created_by, created_at) VALUES (?, ?, 'PREVIEWED', ?, ?, ?)",
                           (preview["source_name"], preview["source_sha256"], json.dumps(redacted_preview, ensure_ascii=False), actor_user_id, now)).lastrowid
        execute(
            connection,
            "UPDATE renewal_import_batches SET preview_ciphertext=? WHERE id=?",
            (encrypted_preview, batch_id),
        )
        for row in preview["rows"]:
            raw_json = json.dumps(row["raw"], ensure_ascii=False, default=str)
            history_note = row.get("history_note") or None
            assistance_note = row.get("assistance_note") or None
            execute(
                connection,
                "INSERT INTO renewal_import_staging("
                "batch_id,row_no,match_status,member_id,org_unit_id,due_month,proposed_status,"
                "history_note,assistance_note,raw_json,issue_code,created_at,"
                "history_note_ciphertext,assistance_note_ciphertext,raw_json_ciphertext) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, NULL, NULL, '{}', ?, ?, ?, ?, ?)",
                (
                    batch_id, row["row_no"], row["match_status"], row["member_id"],
                    row["org_unit_id"], row["due_month"], row["proposed_status"],
                    row["issue_code"], now,
                    encrypt_text(history_note) if history_note else None,
                    encrypt_text(assistance_note) if assistance_note else None,
                    encrypt_text(raw_json),
                ),
            )
        return batch_id


def apply_preview(batch_id: int, actor_user_id: int, renewal_year: int, confirmation: str) -> dict[str, int]:
    if confirmation != "确认正式导入续费周期":
        raise PermissionError("确认文字不匹配，已禁止正式导入")
    batch = fetch_one(
        "SELECT id, status, source_name FROM renewal_import_batches WHERE id=?",
        (batch_id,),
    )
    if not batch:
        raise ValueError("续费预检批次不存在")
    if batch["status"] != "PREVIEWED":
        raise ValueError("该批次已处理，不能重复正式导入")
    allowed = accessible_org_ids(actor_user_id)
    staged_total = fetch_one(
        "SELECT COUNT(*) AS count FROM renewal_import_staging WHERE batch_id=?",
        (batch_id,),
    )["count"]
    placeholders = ",".join("?" for _ in IMPORTABLE_MATCH_STATUSES)
    rows = fetch_all(
        "SELECT s.id, s.member_id, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, "
        "s.due_month, s.proposed_status FROM renewal_import_staging s "
        "JOIN members m ON m.id=s.member_id "
        f"WHERE s.batch_id=? AND s.member_id IS NOT NULL AND s.org_unit_id IS NOT NULL "
        f"AND s.match_status IN ({placeholders})",
        (batch_id, *sorted(IMPORTABLE_MATCH_STATUSES)),
    )
    if allowed is not None:
        rows = [row for row in rows if row["org_unit_id"] in allowed]
    if not rows:
        raise ValueError("没有已关联生产学员且通过匹配门禁的可导入记录")
    # A renewal workbook can contain more than one eligible line for the same
    # production member (for example, a duplicated source line or a corrected
    # due month).  renewal_cycles deliberately has a unique member/year key;
    # importing the raw staging rows would therefore abort the whole batch on
    # the second line.  Keep the first staging line deterministically and count
    # later lines as skipped.  The precheck remains read-only and the batch
    # audit records the deduplication count for manual follow-up.
    unique_rows: list[dict[str, Any]] = []
    seen_member_ids: set[int] = set()
    duplicate_skipped = 0
    for row in rows:
        member_id = int(row["member_id"])
        if member_id in seen_member_ids:
            duplicate_skipped += 1
            continue
        seen_member_ids.add(member_id)
        unique_rows.append(row)
    rows = unique_rows
    member_ids = sorted(seen_member_ids)
    member_placeholders = ",".join("?" for _ in member_ids)
    existing_count = fetch_one(
        f"SELECT COUNT(*) AS count FROM renewal_cycles WHERE renewal_year=? "
        f"AND member_id IN ({member_placeholders})",
        (renewal_year, *member_ids),
    )["count"]
    if existing_count:
        raise ValueError(
            "目标年度已存在续费周期，首次整批导入已停止；请先生成差异确认包"
        )
    now = datetime.now(UTC).isoformat()
    created = 0
    with transaction() as connection:
        for row in rows:
            cycle_id = execute(
                connection,
                "INSERT INTO renewal_cycles(member_id, renewal_year, org_unit_id, due_month, "
                "status, source_batch_id, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    row["member_id"], renewal_year, row["org_unit_id"], row["due_month"],
                    row["proposed_status"], batch_id, now, now,
                ),
            ).lastrowid
            execute(
                connection,
                "INSERT INTO renewal_status_history(renewal_cycle_id, from_status, to_status, "
                "reason, changed_by, created_at) VALUES (?, NULL, ?, ?, ?, ?)",
                (
                    cycle_id,
                    row["proposed_status"],
                    f"续费名单正式导入（批次 #{batch_id}）",
                    actor_user_id,
                    now,
                ),
            )
            created += 1
        execute(
            connection,
            "UPDATE renewal_import_batches SET status='APPLIED', applied_at=? WHERE id=?",
            (now, batch_id),
        )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.import.apply",
            resource_type="renewal_import_batch",
            resource_id=str(batch_id),
            purpose="续费名单正式导入",
            after={
                "renewal_year": renewal_year,
                "created": created,
                "updated": 0,
                "duplicate_staging_rows_skipped": duplicate_skipped,
            },
        )
    return {
        "created": created,
        "updated": 0,
        "skipped": staged_total - len(rows),
    }


def rollback_import(
    batch_id: int,
    actor_user_id: int,
    confirmation: str,
) -> dict[str, int]:
    if confirmation != "确认回滚续费导入批次":
        raise PermissionError("确认文字不匹配，已禁止回滚")
    batch = fetch_one(
        "SELECT id, status, applied_at FROM renewal_import_batches WHERE id=?",
        (batch_id,),
    )
    if not batch:
        raise ValueError("续费导入批次不存在")
    if batch["status"] != "APPLIED":
        raise ValueError("只有已正式导入且未回滚的批次可以回滚")
    cycles = fetch_all(
        "SELECT c.id, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, "
        "c.created_at, c.updated_at FROM renewal_cycles c "
        "JOIN members m ON m.id=c.member_id "
        "WHERE c.source_batch_id=? ORDER BY c.id",
        (batch_id,),
    )
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and any(
        cycle["org_unit_id"] not in allowed for cycle in cycles
    ):
        raise PermissionError("批次包含当前账号授权范围外的续费周期")
    if any(cycle["created_at"] != cycle["updated_at"] for cycle in cycles):
        raise ValueError("批次中的续费周期已被修改，必须先人工生成联合回滚清单")
    cycle_ids = [cycle["id"] for cycle in cycles]
    if cycle_ids:
        placeholders = ",".join("?" for _ in cycle_ids)
        followup_count = fetch_one(
            f"SELECT COUNT(*) AS count FROM renewal_followups "
            f"WHERE renewal_cycle_id IN ({placeholders})",
            tuple(cycle_ids),
        )["count"]
        if followup_count:
            raise ValueError("批次中的续费周期已有跟进记录，禁止自动回滚")
    now = datetime.now(UTC).isoformat()
    with transaction() as connection:
        if cycle_ids:
            placeholders = ",".join("?" for _ in cycle_ids)
            execute(
                connection,
                f"DELETE FROM renewal_status_history WHERE renewal_cycle_id IN ({placeholders})",
                tuple(cycle_ids),
            )
            execute(
                connection,
                f"DELETE FROM renewal_cycles WHERE id IN ({placeholders})",
                tuple(cycle_ids),
            )
        execute(
            connection,
            "UPDATE renewal_import_batches SET status='ROLLED_BACK' WHERE id=?",
            (batch_id,),
        )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.import.rollback",
            resource_type="renewal_import_batch",
            resource_id=str(batch_id),
            purpose="续费名单整批回滚",
            before={"status": "APPLIED", "cycle_count": len(cycle_ids)},
            after={"status": "ROLLED_BACK", "rolled_back_at": now},
        )
    return {"deleted_cycles": len(cycle_ids)}


def list_cycles(
    user_id: int,
    year: int = 2026,
    status: str | None = None,
    *,
    org_unit_id: str | None = None,
    due_month: int | None = None,
    member_name: str | None = None,
    renewal_status: str = "UNRENEWED",
    include_past: bool = False,
) -> list[dict[str, Any]]:
    """List renewal cycles with scoped, privacy-safe operational filters.

    The default view is intentionally limited to the current year's remaining
    months and cycles that are not marked RENEWED. Callers can select a month
    explicitly or opt into all months for historical review.
    """
    conditions = ["c.renewal_year=?"]
    params: list[Any] = [year]
    if status:
        conditions.append("c.status=?")
        params.append(status)
    else:
        if renewal_status == "RENEWED":
            conditions.append("c.status='RENEWED'")
        elif renewal_status == "UNRENEWED":
            conditions.append("c.status<>'RENEWED'")
        elif renewal_status != "ALL":
            raise ValueError("是否续费筛选值无效")
    if org_unit_id:
        conditions.append(f"{MEMBER_RENEWAL_ORG_SQL}=?")
        params.append(org_unit_id)
    if due_month is not None:
        if not 1 <= due_month <= 12:
            raise ValueError("月份必须在1至12之间")
        conditions.append("c.due_month=?")
        params.append(due_month)
    elif not include_past and year == datetime.now(UTC).year:
        conditions.append("c.due_month>=? AND c.due_month<=12")
        params.append(datetime.now(UTC).month)
    if member_name and member_name.strip():
        conditions.append("m.name LIKE ?")
        params.append(f"%{member_name.strip()}%")
    rows = fetch_all(
        "SELECT c.id, c.member_id, m.member_code, m.name AS member_name, c.renewal_year, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, o.name AS org_name, "
        "c.org_unit_id AS imported_org_unit_id, imported_org.name AS imported_org_name, "
        "m.org_unit_id AS member_org_unit_id, m.development_org_unit_id AS member_development_org_unit_id, "
        f"{MEMBER_CLASS_NAME_SQL} AS member_class_name, "
        f"{MEMBER_GROUP_NAME_SQL} AS member_group_name, "
        "c.due_month, c.phase, c.status, c.result, "
        "c.assigned_user_id, u.display_name AS assigned_user_name, c.completed_at, c.updated_at "
        "FROM renewal_cycles c JOIN members m ON m.id=c.member_id "
        f"JOIN org_units o ON o.id={MEMBER_RENEWAL_ORG_SQL} "
        "LEFT JOIN org_units imported_org ON imported_org.id=c.org_unit_id "
        "LEFT JOIN app_users u ON u.id=c.assigned_user_id WHERE "
        + " AND ".join(conditions)
        + " ORDER BY c.due_month, c.id",
        tuple(params),
    )
    allowed = accessible_org_ids(user_id)
    if allowed is not None:
        rows = [row for row in rows if row["org_unit_id"] in allowed]
    for row in rows:
        row["stage"] = determine_renewal_stage(
            row["renewal_year"], row["due_month"], row["status"]
        )
    return rows


def list_today_actions(
    user_id: int,
    year: int = 2026,
    *,
    org_unit_id: str | None = None,
    stage: str | None = None,
    reason: str | None = None,
    as_of: date | datetime | None = None,
) -> dict[str, Any]:
    """Return the explainable renewal actions that need attention today.

    This list deliberately contains only facts needed to choose a person and
    an action.  The heavier action-card endpoint remains the single source for
    message references and is opened only after an operator chooses a row.
    """
    if not 2020 <= int(year) <= 2100:
        raise ValueError("续费年度无效")
    current = as_of or datetime.now(UTC)
    current_date = current.date() if isinstance(current, datetime) else current
    requested_stage = str(stage or "").strip().upper() or None
    if requested_stage and requested_stage not in RENEWAL_STAGE_LABELS:
        raise ValueError("今日行动阶段筛选值无效")
    requested_reason = str(reason or "").strip().upper() or None
    if requested_reason and requested_reason not in TODAY_ACTION_REASON_CODES:
        raise ValueError("今日行动原因筛选值无效")

    allowed = accessible_org_ids(user_id)
    if org_unit_id and allowed is not None and org_unit_id not in allowed:
        raise PermissionError("组织不在当前用户授权范围内")

    conditions = ["c.renewal_year=?"]
    params: list[Any] = [year]
    if org_unit_id:
        conditions.append(f"{MEMBER_RENEWAL_ORG_SQL}=?")
        params.append(org_unit_id)
    rows = fetch_all(
        "SELECT c.id, c.member_id, c.renewal_year, c.due_month, c.status, "
        "c.assigned_user_id, u.display_name AS assigned_user_name, "
        "m.name AS member_name, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, o.name AS org_name, "
        f"{MEMBER_CLASS_NAME_SQL} AS class_name, "
        f"{MEMBER_GROUP_NAME_SQL} AS group_name "
        "FROM renewal_cycles c JOIN members m ON m.id=c.member_id "
        f"JOIN org_units o ON o.id={MEMBER_RENEWAL_ORG_SQL} "
        "LEFT JOIN app_users u ON u.id=c.assigned_user_id WHERE "
        + " AND ".join(conditions),
        tuple(params),
    )
    if allowed is not None:
        rows = [row for row in rows if row["org_unit_id"] in allowed]
    if not rows:
        return {
            "year": year,
            "as_of": current_date.isoformat(),
            "summary": {
                "total": 0,
                "overdue_count": 0,
                "today_count": 0,
                "support_needed_count": 0,
                "stage_untouched_count": 0,
                "next_step_missing_count": 0,
                "stage_counts": {},
            },
            "items": [],
        }

    cycle_ids = [int(row["id"]) for row in rows]
    placeholders = ",".join("?" for _ in cycle_ids)
    followup_rows = fetch_all(
        "SELECT id, renewal_cycle_id, followed_at, channel, intention, "
        "needs_support, next_action, next_followup_at "
        f"FROM renewal_followups WHERE renewal_cycle_id IN ({placeholders}) "
        "ORDER BY followed_at DESC, id DESC",
        tuple(cycle_ids),
    )
    followups_by_cycle: dict[int, list[dict[str, Any]]] = {}
    for followup in followup_rows:
        followups_by_cycle.setdefault(int(followup["renewal_cycle_id"]), []).append(
            followup
        )

    items: list[dict[str, Any]] = []
    for row in rows:
        cycle_id = int(row["id"])
        cycle_stage = determine_renewal_stage(
            row["renewal_year"],
            row["due_month"],
            row["status"],
            as_of=current_date,
        )
        stage_code = cycle_stage["code"]
        if stage_code not in TODAY_ACTION_STAGE_CODES:
            continue
        if requested_stage and stage_code != requested_stage:
            continue

        cycle_followups = followups_by_cycle.get(cycle_id, [])
        cycle_followups.sort(
            key=lambda value: (
                _calendar_date(value.get("followed_at")) or date.min,
                str(value.get("followed_at") or ""),
                int(value.get("id") or 0),
            ),
            reverse=True,
        )
        latest = cycle_followups[0] if cycle_followups else None
        current_month_followups = [
            value
            for value in cycle_followups
            if (
                (followed_on := _calendar_date(value.get("followed_at")))
                and followed_on.year == current_date.year
                and followed_on.month == current_date.month
            )
        ]
        latest_current_month = (
            current_month_followups[0] if current_month_followups else None
        )

        reason_items: list[dict[str, Any]] = []

        def add_reason(code: str, label: str, **details: Any) -> None:
            reason_items.append({"code": code, "label": label, **details})

        next_followup_date = _calendar_date(
            latest.get("next_followup_at") if latest else None
        )
        if next_followup_date and next_followup_date < current_date:
            days_overdue = (current_date - next_followup_date).days
            add_reason(
                "FOLLOWUP_OVERDUE",
                f"约定跟进已逾期{days_overdue}天",
                days_overdue=days_overdue,
            )
        elif next_followup_date and next_followup_date == current_date:
            add_reason("FOLLOWUP_TODAY", "约定今天联系")

        if latest and bool(latest.get("needs_support")):
            add_reason("SUPPORT_NEEDED", "最近一次沟通标记需要协助")

        if not current_month_followups:
            add_reason(
                "STAGE_UNTOUCHED",
                f"进入{cycle_stage['label']}，本阶段尚未联系",
            )

        if (
            stage_code in {"RENEW_2", "FOLLOW_1", "DUE_NOW", "RECOVERY"}
            and latest_current_month
            and not str(latest_current_month.get("next_action") or "").strip()
            and not str(latest_current_month.get("next_followup_at") or "").strip()
        ):
            add_reason("NEXT_STEP_MISSING", "已沟通，但下一步尚未明确")

        if requested_reason and not any(
            item["code"] == requested_reason for item in reason_items
        ):
            continue
        if not reason_items:
            continue

        reason_items.sort(key=lambda item: TODAY_ACTION_REASON_RANK[item["code"]])
        safe_next_action = _redact_action_text(
            latest.get("next_action") if latest else None
        )
        item = {
            "cycle_id": cycle_id,
            "member_id": row["member_id"],
            "member_name": row["member_name"],
            "org_unit_id": row["org_unit_id"],
            "org_name": row["org_name"],
            "class_name": row["class_name"],
            "group_name": row["group_name"],
            "renewal_year": row["renewal_year"],
            "due_month": row["due_month"],
            "status": row["status"],
            "stage": stage_code,
            "stage_label": cycle_stage["label"],
            "assigned_user_id": row["assigned_user_id"],
            "assigned_user_name": row["assigned_user_name"],
            "latest_followup_at": latest.get("followed_at") if latest else None,
            "latest_channel": latest.get("channel") if latest else None,
            "intention": latest.get("intention") if latest else None,
            "needs_support": bool(latest.get("needs_support")) if latest else False,
            "next_action": safe_next_action,
            "next_followup_at": latest.get("next_followup_at") if latest else None,
            "primary_reason": reason_items[0]["code"],
            "reasons": reason_items,
            "reason_codes": [item["code"] for item in reason_items],
        }
        items.append(item)

    def action_sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        next_date = _calendar_date(item.get("next_followup_at"))
        return (
            TODAY_ACTION_REASON_RANK[item["primary_reason"]],
            TODAY_ACTION_STAGE_RANK[item["stage"]],
            next_date or date.max,
            str(item.get("member_name") or ""),
            int(item["cycle_id"]),
        )

    items.sort(key=action_sort_key)
    summary = {
        "total": len(items),
        "overdue_count": sum(
            1 for item in items if "FOLLOWUP_OVERDUE" in item["reason_codes"]
        ),
        "today_count": sum(
            1 for item in items if "FOLLOWUP_TODAY" in item["reason_codes"]
        ),
        "support_needed_count": sum(
            1 for item in items if "SUPPORT_NEEDED" in item["reason_codes"]
        ),
        "stage_untouched_count": sum(
            1 for item in items if "STAGE_UNTOUCHED" in item["reason_codes"]
        ),
        "next_step_missing_count": sum(
            1 for item in items if "NEXT_STEP_MISSING" in item["reason_codes"]
        ),
        "stage_counts": {},
    }
    for item in items:
        summary["stage_counts"][item["stage"]] = (
            summary["stage_counts"].get(item["stage"], 0) + 1
        )
    return {
        "year": year,
        "as_of": current_date.isoformat(),
        "summary": summary,
        "items": items,
    }


def list_cycle_coverage(
    user_id: int,
    year: int = 2026,
    *,
    org_unit_id: str | None = None,
    member_name: str | None = None,
    include_synced: bool = False,
    actionable_only: bool = True,
    limit: int = 200,
) -> dict[str, Any]:
    """Compare member master data with annual renewal-cycle coverage.

    This is intentionally read-only. Missing cycles remain visible instead of
    being silently excluded from the operations page; creating a cycle is a
    separate, audited action.
    """
    if not 1 <= limit <= 500:
        raise ValueError("同步检查条数必须在1至500之间")
    conditions = ["1=1"]
    params: list[Any] = [year]
    if org_unit_id:
        conditions.append(f"{MEMBER_RENEWAL_ORG_SQL}=?")
        params.append(org_unit_id)
    if member_name and member_name.strip():
        conditions.append("m.name LIKE ?")
        params.append(f"%{member_name.strip()}%")
    rows = fetch_all(
        "SELECT m.id AS member_id, m.member_code, m.name AS member_name, "
        "m.status AS member_status, m.renewal_month, "
        f"{MEMBER_CLASS_NAME_SQL} AS member_class_name, "
        f"{MEMBER_GROUP_NAME_SQL} AS member_group_name, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, o.name AS org_name, "
        "c.id AS cycle_id, c.due_month, c.status AS cycle_status, c.updated_at "
        "FROM members m "
        f"JOIN org_units o ON o.id={MEMBER_RENEWAL_ORG_SQL} "
        "LEFT JOIN renewal_cycles c ON c.member_id=m.id AND c.renewal_year=? "
        "WHERE " + " AND ".join(conditions) + " ORDER BY o.name, m.name, m.id",
        tuple(params),
    )
    allowed = accessible_org_ids(user_id)
    if allowed is not None:
        rows = [row for row in rows if row["org_unit_id"] in allowed]

    decorated: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        member_status = str(item["member_status"] or "").upper()
        member_active = member_status == "ACTIVE"
        recurring_month = _member_renewal_month(item.get("renewal_month"))
        if item.get("cycle_id"):
            if member_active:
                sync_status = "SYNCED"
            elif member_status == "SUSPENDED":
                sync_status = "SYNCED_SUSPENDED"
            else:
                sync_status = "SYNCED_INACTIVE"
        elif not member_active:
            sync_status = "SUSPENDED" if member_status == "SUSPENDED" else "INACTIVE"
        elif recurring_month:
            sync_status = "READY_TO_CREATE"
            item["due_month"] = recurring_month
        else:
            sync_status = "MISSING_RENEWAL_MONTH"
        item["sync_status"] = sync_status
        item["can_create_cycle"] = sync_status == "READY_TO_CREATE"
        decorated.append(item)

    active_rows = [
        item
        for item in decorated
        if str(item["member_status"] or "").upper() == "ACTIVE"
    ]
    summary = {
        "member_total": len(decorated),
        "active_member_total": len(active_rows),
        "cycle_total": sum(1 for item in decorated if item.get("cycle_id")),
        "ready_to_create_count": sum(
            1 for item in decorated if item["sync_status"] == "READY_TO_CREATE"
        ),
        "missing_renewal_month_count": sum(
            1
            for item in decorated
            if item["sync_status"] == "MISSING_RENEWAL_MONTH"
        ),
        "inactive_member_count": sum(
            1
            for item in decorated
            if item["sync_status"] in {"INACTIVE", "SYNCED_INACTIVE"}
        ),
        "suspended_member_count": sum(
            1
            for item in decorated
            if item["sync_status"] in {"SUSPENDED", "SYNCED_SUSPENDED"}
        ),
    }
    visible = (
        decorated
        if include_synced
        else [item for item in decorated if item["sync_status"] != "SYNCED"]
    )
    if actionable_only:
        visible = [
            item
            for item in visible
            if item["can_create_cycle"]
            or item["sync_status"] == "MISSING_RENEWAL_MONTH"
        ]
    return {
        "year": year,
        "summary": summary,
        "rows": visible[:limit],
        "truncated": len(visible) > limit,
    }


def create_cycle_from_member(
    member_id: int,
    actor_user_id: int,
    *,
    renewal_year: int,
    confirmation: str,
) -> int:
    """Create one missing annual cycle from confirmed member-master fields."""
    if confirmation != "确认从学员主档建立续费周期":
        raise PermissionError("确认文字不匹配，已禁止建立续费周期")
    member = fetch_one(
        "SELECT m.id, m.status, m.renewal_month, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id "
        "FROM members m WHERE m.id=?",
        (member_id,),
    )
    if not member:
        raise ValueError("学员不存在")
    if str(member["status"] or "").upper() != "ACTIVE":
        raise ValueError("只有在册学员可以建立新的续费周期")
    due_month = _member_renewal_month(member.get("renewal_month"))
    if not due_month:
        raise ValueError("请先在学员管理补充有效的续费月份")
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and member["org_unit_id"] not in allowed:
        raise PermissionError("学员不在当前账号的续费组织范围内")
    if fetch_one(
        "SELECT id FROM renewal_cycles WHERE member_id=? AND renewal_year=?",
        (member_id, renewal_year),
    ):
        raise ValueError("该学员本年度续费周期已存在")
    now_dt = datetime.now(UTC)
    now = now_dt.isoformat()
    initial_status = _initial_cycle_status(due_month, renewal_year, now_dt)
    completed_at = now if initial_status == "RENEWED" else None
    status_reason = (
        "由学员管理续费月份建立，历史月份自动标记为已续费"
        if initial_status == "RENEWED"
        else "由学员管理续费月份建立"
    )
    with transaction() as connection:
        cycle_id = execute(
            connection,
            "INSERT INTO renewal_cycles(member_id, renewal_year, org_unit_id, due_month, "
            "status, completed_at, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                member_id,
                renewal_year,
                member["org_unit_id"],
                due_month,
                initial_status,
                completed_at,
                now,
                now,
            ),
        ).lastrowid
        execute(
            connection,
            "INSERT INTO renewal_status_history(renewal_cycle_id, from_status, to_status, "
            "reason, changed_by, created_at) VALUES (?, NULL, "
            "?, ?, ?, ?)",
            (cycle_id, initial_status, status_reason, actor_user_id, now),
        )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.cycle.create_from_member",
            resource_type="renewal_cycle",
            resource_id=str(cycle_id),
            org_unit_id=member["org_unit_id"],
            purpose="从学员管理主档补齐单个续费周期",
            after={
                "member_id": member_id,
                "renewal_year": renewal_year,
                "due_month": due_month,
                "status": initial_status,
            },
        )
    return int(cycle_id)


def update_cycle(
    cycle_id: int,
    actor_user_id: int,
    *,
    status: str | None = None,
    phase: str | None = None,
    result: str | None = None,
    assigned_user_id: int | None = None,
) -> None:
    if status is not None:
        status = status.strip().upper()
        if status not in RENEWAL_STATUSES:
            raise ValueError("续费状态无效")
    if phase is not None:
        phase = phase.strip()
        if len(phase) > 32:
            raise ValueError("续费阶段不能超过32个字符")
    if result is not None:
        result = result.strip()
        if len(result) > 64:
            raise ValueError("续费结果不能超过64个字符")
    cycle = _cycle_with_member_scope(cycle_id)
    if not cycle:
        raise ValueError("续费周期不存在")
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and cycle["member_org_unit_id"] not in allowed:
        raise PermissionError("续费周期不在组织授权范围内")
    if assigned_user_id is not None:
        assignee = fetch_one("SELECT id, is_active FROM app_users WHERE id=?", (assigned_user_id,))
        if not assignee or not assignee["is_active"]:
            raise ValueError("责任人账号当前不可用")
        assignee_context = user_context(assigned_user_id)
        if not assignee_context or "renewals:manage" not in assignee_context["permissions"]:
            raise ValueError("责任人当前没有续费运营权限")
        assignee_allowed = accessible_org_ids(assigned_user_id)
        if assignee_allowed is not None and cycle["member_org_unit_id"] not in assignee_allowed:
            raise ValueError("责任人不在续费归属组织范围内")
    fields = {key: value for key, value in {
        "status": status, "phase": phase, "result": result,
        "assigned_user_id": assigned_user_id,
    }.items() if value is not None}
    if not fields:
        raise ValueError("至少提供一项续费周期变更")
    now = datetime.now(UTC).isoformat()
    fields["updated_at"] = now
    if status:
        fields["completed_at"] = (
            now if status in {"RENEWED", "NOT_RENEWING", "EXITED"} else None
        )
    with transaction() as connection:
        assignments = ", ".join(f"{key}=?" for key in fields)
        execute(connection, f"UPDATE renewal_cycles SET {assignments} WHERE id=?", (*fields.values(), cycle_id))
        if status and status != cycle["status"]:
            execute(
                connection,
                "INSERT INTO renewal_status_history(renewal_cycle_id, from_status, to_status, "
                "changed_by, created_at) VALUES (?, ?, ?, ?, ?)",
                (cycle_id, cycle["status"], status, actor_user_id, now),
            )
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.cycle.update",
            resource_type="renewal_cycle",
            resource_id=str(cycle_id),
            org_unit_id=cycle["member_org_unit_id"],
            after=fields,
        )


def add_followup(
    cycle_id: int,
    actor_user_id: int,
    *,
    channel: str,
    summary: str,
    intention: str | None = None,
    needs_support: bool = False,
    next_action: str | None = None,
    next_followup_at: str | None = None,
) -> int:
    cycle = _cycle_with_member_scope(cycle_id)
    if not cycle:
        raise ValueError("续费周期不存在")
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and cycle["member_org_unit_id"] not in allowed:
        raise PermissionError("续费周期不在组织授权范围内")
    if channel.strip().upper() not in {"PHONE", "WECHAT", "MEETING", "VISIT", "OTHER"}:
        raise ValueError("不支持的联系渠道")
    if len(summary.strip()) < 4:
        raise ValueError("跟进摘要至少填写4个字符")
    now = datetime.now(UTC).isoformat()
    with transaction() as connection:
        followup_id = execute(
            connection,
            "INSERT INTO renewal_followups(renewal_cycle_id, followed_at, followed_by, channel, "
            "summary, intention, needs_support, next_action, next_followup_at, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                cycle_id, now, actor_user_id, channel.strip().upper(), summary.strip(),
                intention, 1 if needs_support else 0, next_action, next_followup_at, now,
            ),
        ).lastrowid
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.followup.create",
            resource_type="renewal_cycle",
            resource_id=str(cycle_id),
            org_unit_id=cycle["member_org_unit_id"],
            after={"followup_id": followup_id, "channel": channel.strip().upper()},
        )
        return followup_id


def list_followups(cycle_id: int, actor_user_id: int) -> list[dict[str, Any]]:
    cycle = _cycle_with_member_scope(cycle_id)
    if not cycle:
        raise ValueError("续费周期不存在")
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and cycle["member_org_unit_id"] not in allowed:
        raise PermissionError("续费周期不在组织授权范围内")
    return fetch_all(
        "SELECT id, followed_at, followed_by, channel, summary, intention, needs_support, "
        "next_action, next_followup_at FROM renewal_followups WHERE renewal_cycle_id=? "
        "ORDER BY followed_at DESC, id DESC",
        (cycle_id,),
    )


def get_action_card(
    cycle_id: int,
    actor_user_id: int,
    *,
    as_of: date | datetime | None = None,
) -> dict[str, Any]:
    """Build a read-only, rules-based action card for one renewal cycle."""
    current = as_of or datetime.now(UTC)
    current_date = current.date() if isinstance(current, datetime) else current
    cycle = fetch_one(
        "SELECT c.id, c.member_id, c.renewal_year, c.due_month, c.status, "
        "c.result, c.assigned_user_id, u.display_name AS assigned_user_name, "
        "m.name AS member_name, m.join_date, m.study_start_date, m.membership_years, "
        "m.membership_years_overridden, "
        f"{MEMBER_RENEWAL_ORG_SQL} AS member_org_unit_id, o.name AS org_name, "
        f"{MEMBER_CLASS_NAME_SQL} AS class_name, "
        f"{MEMBER_GROUP_NAME_SQL} AS group_name "
        "FROM renewal_cycles c JOIN members m ON m.id=c.member_id "
        f"JOIN org_units o ON o.id={MEMBER_RENEWAL_ORG_SQL} "
        "LEFT JOIN app_users u ON u.id=c.assigned_user_id WHERE c.id=?",
        (cycle_id,),
    )
    if not cycle:
        raise ValueError("续费周期不存在")
    allowed = accessible_org_ids(actor_user_id)
    if allowed is not None and cycle["member_org_unit_id"] not in allowed:
        raise PermissionError("续费周期不在组织授权范围内")

    latest = fetch_one(
        "SELECT f.id, f.followed_at, f.channel, f.summary, f.intention, "
        "f.needs_support, f.next_action, f.next_followup_at, "
        "u.display_name AS followed_by_name "
        "FROM renewal_followups f LEFT JOIN app_users u ON u.id=f.followed_by "
        "WHERE f.renewal_cycle_id=? ORDER BY f.followed_at DESC, f.id DESC LIMIT 1",
        (cycle_id,),
    )
    if latest:
        latest = dict(latest)
        latest["summary"] = _redact_action_text(latest.get("summary"))
        latest["next_action"] = _redact_action_text(latest.get("next_action"))
        latest["needs_support"] = bool(latest.get("needs_support"))

    stage = determine_renewal_stage(
        cycle["renewal_year"], cycle["due_month"], cycle["status"], as_of=current_date
    )
    memories = verified_member_memories(cycle["member_id"], limit=4, as_of=current_date)
    strategy = _renewal_action_strategy(stage["code"], latest_followup=latest)
    salutation = (
        cycle["member_name"]
        if str(cycle["member_name"]).endswith(("学长", "学姐"))
        else f"{cycle['member_name']}学长"
    )
    memory = memories[0] if memories else None
    if stage["code"] == "CLOSED":
        wechat_reference = None
        phone_opening_reference = None
    else:
        memory_wechat = (
            f"前段时间想到您参加过的{memory['title']}，也想听听您最近有没有新的感受。"
            if memory
            else "有一段时间没和您细聊了，也一直惦记着您的近况。"
        )
        recent_contact = "上次联系后，最近还顺利吗？" if latest else "最近还顺利吗？"
        if stage["code"] in {"PREPARE", "OBSERVE_3"}:
            wechat_reference = (
                f"{salutation}好，{memory_wechat}{recent_contact}"
                "最近企业、学习或生活上如果有我们可以支持的地方，也请随时告诉我。"
            )
            phone_opening_reference = (
                f"{salutation}好，我是盛和塾的运营同仁。今天联系您主要是想关心一下近况，"
                "也听听您最近学习和经营上的感受，现在方便聊几分钟吗？"
            )
        elif stage["code"] == "RENEW_2":
            wechat_reference = (
                f"{salutation}好，{memory_wechat}{recent_contact}"
                "也想和您回顾一下这一年的学习与同行，听听您对下一阶段学习的期待。"
            )
            phone_opening_reference = (
                f"{salutation}好，我是盛和塾的运营同仁。想听听您对这一年学习和同行的真实感受，"
                "也了解一下接下来最希望获得哪些支持，现在方便聊几分钟吗？"
            )
        else:
            wechat_reference = (
                f"{salutation}好，{memory_wechat}{recent_contact}"
                "想和您确认一下下一阶段继续同行的考虑，以及有没有需要我们协调支持的地方。"
            )
            phone_opening_reference = (
                f"{salutation}好，我是盛和塾的运营同仁。今天想基于前面的沟通，"
                "确认一下您对下一阶段同行的考虑，也看看有没有需要我们协助解决的事项。"
            )

    join_date = str(cycle["join_date"])[:10] if cycle.get("join_date") else None
    membership_years = _completed_membership_years(
        cycle.get("join_date"),
        cycle.get("membership_years"),
        cycle.get("membership_years_overridden"),
        as_of=current_date,
    )
    result = {
        "cycle": {
            "id": cycle["id"],
            "renewal_year": cycle["renewal_year"],
            "due_month": cycle["due_month"],
            "status": cycle["status"],
            "result": cycle["result"],
            "assigned_user_id": cycle["assigned_user_id"],
            "assigned_user_name": cycle["assigned_user_name"],
        },
        "member": {
            "id": cycle["member_id"],
            "name": cycle["member_name"],
            "org_unit_id": cycle["member_org_unit_id"],
            "org_name": cycle["org_name"],
            "class_name": cycle["class_name"],
            "group_name": cycle["group_name"],
            "join_date": join_date,
            "study_start_date": (
                str(cycle["study_start_date"])[:10]
                if cycle.get("study_start_date")
                else None
            ),
            "membership_years": membership_years,
        },
        "stage": stage,
        "latest_followup": latest,
        "current_context": {
            "intention": latest.get("intention") if latest else None,
            "needs_support": latest.get("needs_support") if latest else False,
            "next_action": latest.get("next_action") if latest else None,
            "next_followup_at": latest.get("next_followup_at") if latest else None,
        },
        "verified_memories": memories,
        "action": {
            "goal": strategy["goal"],
            "recommended_channel": strategy["channel"],
            "recommendation_reason": strategy["reason"],
            "coordination_recommended": strategy["coordination_recommended"],
            "wechat_reference": wechat_reference,
            "phone_opening_reference": phone_opening_reference,
            "questions": strategy["questions"],
            "do_not": strategy["do_not"],
            "advice_source": "RULE_TEMPLATE_V1",
        },
        "data_quality": {
            "facts_only": True,
            "memory_count": len(memories),
            "memory_fallback_used": not bool(memories),
            "join_date_available": bool(join_date),
        },
        "policy": "阶段由后端月份规则决定；建议仅使用已验证经历，不包含手机号或明确金额，不自动修改续费数据。",
    }
    with transaction() as connection:
        write_audit(
            connection,
            actor_user_id=actor_user_id,
            action="renewals.action_card.view",
            resource_type="renewal_cycle",
            resource_id=str(cycle_id),
            org_unit_id=cycle["member_org_unit_id"],
            purpose="查看单个学长续费今日行动卡",
            after={
                "stage": stage["code"],
                "memory_count": len(memories),
                "facts_only": True,
            },
        )
    return result


def list_overview(user_id: int, year: int = 2026) -> dict[str, Any]:
    allowed = accessible_org_ids(user_id)
    rows = fetch_all(
        "SELECT "
        f"{MEMBER_RENEWAL_ORG_SQL} AS org_unit_id, o.name AS org_name, "
        "c.due_month, c.status, COUNT(*) AS count "
        "FROM renewal_cycles c JOIN members m ON m.id=c.member_id "
        f"JOIN org_units o ON o.id={MEMBER_RENEWAL_ORG_SQL} "
        "WHERE c.renewal_year=? "
        f"GROUP BY {MEMBER_RENEWAL_ORG_SQL}, o.name, c.due_month, c.status",
        (year,),
    )
    if allowed is not None:
        rows = [row for row in rows if row["org_unit_id"] in allowed]
    return {"year": year, "rows": rows}
