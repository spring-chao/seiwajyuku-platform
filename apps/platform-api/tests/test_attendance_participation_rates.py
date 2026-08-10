from __future__ import annotations

import unittest
import asyncio
from io import BytesIO
from datetime import UTC, datetime
from openpyxl import load_workbook

from app.api.attendance import (
    download_event_group_records,
    event_group_detail,
    list_activity_sessions,
    list_event_groups,
)
from app.db import execute, fetch_one, transaction
from app.migrations import run_migrations
from app.services.iam import seed_iam
from app.services.members import create_member


class AttendanceParticipationRateTests(unittest.TestCase):
    """Verify class and regional participation counts use organization relations."""

    @classmethod
    def setUpClass(cls) -> None:
        run_migrations()
        seed_iam()
        cls.admin = fetch_one("SELECT id FROM app_users WHERE username='admin'")
        now = datetime.now(UTC).isoformat()
        with transaction() as connection:
            execute(
                connection,
                "INSERT OR IGNORE INTO org_units"
                "(id, unit_code, name, unit_type, parent_id, is_active, created_at, updated_at) "
                "VALUES ('participation-rate-center', 'PARTICIPATION_RATE_CENTER', '参会率测试分中心', "
                "'REGIONAL_CENTER', 'org-suzhou', 1, ?, ?)",
                (now, now),
            )
            execute(
                connection,
                "INSERT OR IGNORE INTO org_units"
                "(id, unit_code, name, unit_type, parent_id, is_active, created_at, updated_at) "
                "VALUES ('participation-rate-class', 'PARTICIPATION_RATE_CLASS', '参会率测试班', "
                "'CLASS', 'participation-rate-center', 1, ?, ?)",
                (now, now),
            )

        cls.member_ids = []
        for code, name in (
            ("PARTICIPATION-RATE-001", "参会率测试学员一"),
            ("PARTICIPATION-RATE-002", "参会率测试学员二"),
            ("PARTICIPATION-RATE-003", "参会率测试学员三"),
        ):
            member = fetch_one("SELECT id FROM members WHERE member_code=?", (code,))
            member_id = (
                member["id"]
                if member
                else create_member(
                    cls.admin["id"],
                    member_code=code,
                    name=name,
                    org_unit_id="participation-rate-center",
                    development_org_unit_id=None,
                    phone=None,
                )
            )
            cls.member_ids.append(member_id)

        with transaction() as connection:
            now = datetime.now(UTC).isoformat()
            for member_id in cls.member_ids[:2]:
                execute(
                    connection,
                    "INSERT OR IGNORE INTO member_org_relations"
                    "(member_id, org_unit_id, relation_type, is_primary, source_type, created_at, updated_at) "
                    "VALUES (?, 'participation-rate-class', 'STUDY_CLASS', 1, 'TEST', ?, ?)",
                    (member_id, now, now),
                )

            cls.group_ids = []
            for external_id, title, activity_type, study_org_unit_id in (
                (
                    "participation-rate-class-group",
                    "参会率测试班级活动",
                    "CLASS_MEETING",
                    "participation-rate-class",
                ),
                (
                    "participation-rate-report-group",
                    "参会率测试分中心报告会",
                    "CENTER_QUARTERLY_REPORT",
                    None,
                ),
            ):
                group = execute(
                    connection,
                    "INSERT INTO attendance_event_groups"
                    "(source_key, external_group_id, org_unit_id, study_org_unit_id, title, activity_type, "
                    "event_date, status, created_at, updated_at) "
                    "VALUES ('participation-rate-test', ?, 'participation-rate-center', ?, ?, ?, '2026-08-04', 'ACTIVE', ?, ?)",
                    (external_id, study_org_unit_id, title, activity_type, now, now),
                )
                group_id = group.lastrowid
                cls.group_ids.append(group_id)
                session = execute(
                    connection,
                    "INSERT INTO attendance_sessions"
                    "(event_group_id, external_session_id, session_code, session_name, session_order, status, created_at, updated_at) "
                    "VALUES (?, ?, 'MORNING', '上午', 1, 'ACTIVE', ?, ?)",
                    (group_id, f"{external_id}-session", now, now),
                )
                session_id = session.lastrowid
                for index, member_id in enumerate(cls.member_ids[:2]):
                    execute(
                        connection,
                        "INSERT INTO attendance_records"
                        "(attendance_session_id, external_record_id, member_id, member_code_snapshot, name_snapshot, "
                        "participant_type, score_eligible, attendance_status, received_at, created_at, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, 'MEMBER', 1, ?, ?, ?, ?)",
                        (
                            session_id,
                            f"{external_id}-member-{index}",
                            member_id,
                            f"PARTICIPATION-RATE-00{index + 1}",
                            f"参会率测试学员{['一', '二'][index]}",
                            "PRESENT" if index == 0 else "ABSENT",
                            now,
                            now,
                            now,
                        ),
                    )
                if activity_type == "CLASS_MEETING":
                    execute(
                        connection,
                        "INSERT INTO attendance_records"
                        "(attendance_session_id, external_record_id, member_id, member_code_snapshot, name_snapshot, "
                        "participant_type, score_eligible, attendance_status, received_at, created_at, updated_at) "
                        "VALUES (?, ?, NULL, NULL, '参会率测试非塾生嘉宾', 'GUEST', 0, 'PRESENT', ?, ?, ?)",
                        (session_id, f"{external_id}-guest-1", now, now, now),
                    )
                if activity_type.startswith("CENTER_"):
                    execute(
                        connection,
                        "INSERT INTO attendance_records"
                        "(attendance_session_id, external_record_id, member_id, member_code_snapshot, name_snapshot, "
                        "participant_type, score_eligible, attendance_status, received_at, created_at, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, 'MEMBER', 1, 'PRESENT', ?, ?, ?)",
                        (
                            session_id,
                            f"{external_id}-member-3",
                            cls.member_ids[2],
                            "PARTICIPATION-RATE-003",
                            "参会率测试学员三",
                            now,
                            now,
                            now,
                        ),
                    )

    def test_event_groups_include_both_scopes(self) -> None:
        response = list_event_groups(
            month="2026-08", user={"id": self.admin["id"]}
        )
        rows = {row["title"]: row for row in self._detail_rows(response)}
        class_row = rows["参会率测试班级活动"]
        report_row = rows["参会率测试分中心报告会"]
        self.assertEqual(
            (class_row["class_member_count"], class_row["class_present_count"]),
            (2, 1),
        )
        self.assertEqual(
            (class_row["record_count"], class_row["present_count"]),
            (3, 2),
        )
        self.assertEqual(
            (report_row["region_member_count"], report_row["region_present_count"]),
            (3, 2),
        )

    def test_event_group_detail_includes_session_scope_counts(self) -> None:
        response = event_group_detail(
            self.group_ids[0], user={"id": self.admin["id"]}
        )
        session = response["data"]["sessions"][0]
        self.assertEqual(session["class_member_count"], 2)
        self.assertEqual(session["class_present_count"], 1)
        self.assertEqual((session["record_count"], session["present_count"]), (3, 2))

        report_response = event_group_detail(
            self.group_ids[1], user={"id": self.admin["id"]}
        )
        report_breakdown = report_response["data"]["class_breakdown"]
        self.assertEqual(len(report_breakdown), 1)
        self.assertEqual(report_breakdown[0]["class_name"], "参会率测试班")
        self.assertEqual(
            (
                report_breakdown[0]["class_member_count"],
                report_breakdown[0]["class_present_count"],
            ),
            (2, 1),
        )

    def test_activity_sessions_split_group_counts_and_export_records(self) -> None:
        now = datetime.now(UTC).isoformat()
        with transaction() as connection:
            group = execute(
                connection,
                "INSERT INTO attendance_event_groups"
                "(source_key, external_group_id, org_unit_id, study_org_unit_id, title, activity_type, "
                "event_date, status, created_at, updated_at) "
                "VALUES ('activity-session-test', 'activity-session-test-group', 'participation-rate-center', "
                "'participation-rate-class', '多场次活动', 'CLASS_MEETING', '2026-08-05', 'ACTIVE', ?, ?)",
                (now, now),
            )
            group_id = group.lastrowid
            morning = execute(
                connection,
                "INSERT INTO attendance_sessions"
                "(event_group_id, external_session_id, session_code, session_name, session_order, status, created_at, updated_at) "
                "VALUES (?, 'activity-session-test-morning', 'MORNING', '上午', 1, 'ACTIVE', ?, ?)",
                (group_id, now, now),
            )
            execute(
                connection,
                "INSERT INTO attendance_records"
                "(attendance_session_id, external_record_id, member_id, member_code_snapshot, name_snapshot, "
                "participant_type, score_eligible, attendance_status, received_at, created_at, updated_at) "
                "VALUES (?, 'activity-session-morning-member', ?, 'PARTICIPATION-RATE-001', '参会率测试学员一', 'MEMBER', 1, 'PRESENT', ?, ?, ?)",
                (morning.lastrowid, self.member_ids[0], now, now, now),
            )
            session = execute(
                connection,
                "INSERT INTO attendance_sessions"
                "(event_group_id, external_session_id, session_code, session_name, session_order, checkin_start_at, status, created_at, updated_at) "
                "VALUES (?, 'activity-session-test-konpa', 'KONPA', '空巴', 2, '2026-08-05 19:00:00', 'ACTIVE', ?, ?)",
                (group_id, now, now),
            )
            session_id = session.lastrowid
            execute(
                connection,
                "INSERT INTO attendance_records"
                "(attendance_session_id, external_record_id, member_id, member_code_snapshot, name_snapshot, "
                "participant_type, score_eligible, attendance_status, checked_at, received_at, created_at, updated_at) "
                "VALUES (?, 'activity-session-konpa-guest', NULL, NULL, '场次测试嘉宾', 'GUEST', 0, 'PRESENT', '2026-08-05 00:43:00', ?, ?, ?)",
                (session_id, now, now, now),
            )

        response = list_activity_sessions(
            month="2026-08", user={"id": self.admin["id"]}
        )
        rows = [row for row in response["data"] if row["id"] == group_id]
        self.assertEqual(len(rows), 2)
        evening = next(row for row in rows if row["session_id"] == session_id)
        self.assertEqual((evening["record_count"], evening["present_count"]), (1, 1))
        self.assertIn("空巴", evening["display_title"])

        export = download_event_group_records(
            group_id, session_id=session_id, user={"id": self.admin["id"]}
        )
        async def collect_body() -> bytes:
            return b"".join([chunk async for chunk in export.body_iterator])

        workbook = load_workbook(BytesIO(asyncio.run(collect_body())), read_only=True)
        sheet = workbook["签到明细"]
        self.assertEqual(
            [cell.value for cell in next(sheet.iter_rows())],
            ["活动", "活动日期", "场次", "姓名", "学员编号", "参与类型", "签到状态", "签到时间", "迟到", "早退", "积分"],
        )
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(list(sheet.iter_rows(values_only=True))[1][7:], ("待核对", "—", "—", "—"))

    @staticmethod
    def _detail_rows(response: dict) -> list[dict]:
        rows = response["data"]
        # Use the known test titles to identify the two rows without relying on IDs.
        return [row for row in rows if row["title"] in {"参会率测试班级活动", "参会率测试分中心报告会"}]


if __name__ == "__main__":
    unittest.main()
