from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.db import fetch_one
from app.migrations import run_migrations
from app.services.iam import seed_iam
from app.services.mp_import import (
    CENTER_SHEETS,
    METRICS,
    apply_preview,
    preview_workbook,
    save_preview,
)
from app.services.plans import (
    enable_plan_write,
    mp_dashboard,
    update_period_values,
)


class MpImportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        run_migrations()
        seed_iam()

    def _workbook(self, path: Path) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        root = workbook.create_sheet("苏州塾MP ")
        root.cell(4, 3, 1000)
        root.cell(4, 4, 381.46)
        root.cell(4, 7, 0.78)
        for index, (sheet_name, _) in enumerate(CENTER_SHEETS.items()):
            sheet = workbook.create_sheet(sheet_name)
            for offset, metric in enumerate(METRICS):
                row = 8 + offset
                sheet.cell(row, 3, metric[1])
                annual = 100 if metric[0] == "active_member_count" else (
                    50 if metric[0] == "new_member_count" else 0.8
                )
                sheet.cell(row, 5, annual)
                for month in range(1, 6):
                    start = 8 + (month - 1) * 4
                    sheet.cell(row, start, 0 if metric[0] == "teaching_goal_rate" else annual)
                    sheet.cell(row, start + 1, "/" if metric[0] == "credits_per_member" else annual)
                    actual = -1 if index == 0 and month == 2 and metric[0] == "new_member_count" else annual
                    sheet.cell(row, start + 2, actual)
        workbook.save(path)

    def test_preview_preserves_states_and_applies_read_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "mp.xlsx"
            self._workbook(path)
            preview = preview_workbook(path)
        self.assertEqual(preview["summary"]["row_count"], 1347)
        states = {row["value_state"] for row in preview["rows"]}
        self.assertIn("NOT_APPLICABLE", states)
        self.assertIn("ZERO_IS_VALID", states)
        self.assertTrue(any(issue["code"] == "NEGATIVE_NEW_MEMBER" for issue in preview["issues"]))
        admin = fetch_one("SELECT id FROM app_users WHERE username='admin'")
        batch_id = save_preview(preview, admin["id"])
        result = apply_preview(batch_id, admin["id"])
        self.assertFalse(result["write_enabled"])
        plan = fetch_one("SELECT status, write_enabled FROM annual_plans WHERE id=?", (result["annual_plan_id"],))
        self.assertEqual(plan, {"status": "DRAFT", "write_enabled": 0})
        huangpu = fetch_one("SELECT parent_id, unit_type FROM org_units WHERE id='org-huangpu'")
        self.assertEqual(huangpu, {"parent_id": "org-suzhou", "unit_type": "CLASS"})
        forecast = fetch_one(
            "SELECT id FROM metric_period_values WHERE annual_plan_id=? AND value_kind='FORECAST' LIMIT 1",
            (result["annual_plan_id"],),
        )
        with self.assertRaisesRegex(PermissionError, "当前只读"):
            update_period_values(
                plan_id=result["annual_plan_id"],
                user_id=admin["id"],
                updates=[{"id": forecast["id"], "numeric_value": 12, "value_state": "VALUE"}],
            )
        enable_plan_write(
            result["annual_plan_id"], admin["id"], "TEST-APPROVAL-ONLY"
        )
        changed = update_period_values(
            plan_id=result["annual_plan_id"],
            user_id=admin["id"],
            updates=[{"id": forecast["id"], "numeric_value": 12, "value_state": "VALUE"}],
        )
        self.assertEqual(changed, 1)
        dashboard = mp_dashboard(
            plan_id=result["annual_plan_id"], user_id=admin["id"], month=1
        )
        self.assertEqual(len(dashboard["centers"]), 6)


if __name__ == "__main__":
    unittest.main()
