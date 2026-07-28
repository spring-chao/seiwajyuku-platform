from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook


ACTIVE_VALUE = "在册"
PHONE_PATTERN = re.compile(r"(?<!\d)(1\d{10})(?!\d)")


def _text(value: Any, max_length: int | None = None) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    if not result:
        return None
    return result[:max_length] if max_length else result


def _date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    matched = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    return "-".join(matched.groups()) if matched else None


def _renewal_month(row: dict[str, Any]) -> str | None:
    for year in (2026, 2025, 2024, 2023):
        raw = _text(row.get(f"fee_{year}_month"))
        if not raw:
            continue
        matched = re.search(r"(\d{1,2})", raw)
        if matched and 1 <= int(matched.group(1)) <= 12:
            return f"{year}-{int(matched.group(1)):02d}"
    return None


def _phone_candidates(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return list(dict.fromkeys(PHONE_PATTERN.findall(str(value))))


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if 0 <= result <= 100 else None


def _source_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["学员数据库全量"]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(values)]
    rows: list[dict[str, Any]] = []
    for source_row, values_row in enumerate(values, start=2):
        if not any(value not in (None, "") for value in values_row):
            continue
        row = dict(zip(headers, values_row, strict=True))
        row["__source_row__"] = source_row
        rows.append(row)
    workbook.close()
    return rows


def _login(client: httpx.Client, username: str, password: str) -> str:
    response = client.post(
        "/api/v1/auth/login",
        json={"username": username, "password": password},
    )
    response.raise_for_status()
    return response.json()["data"]["access_token"]


def _build_payloads(
    rows: list[dict[str, Any]],
    center_ids: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    active_rows = [row for row in rows if _text(row.get("is_active")) == ACTIVE_VALUE]
    single_phone_counts = Counter(
        candidates[0]
        for row in active_rows
        if len(candidates := _phone_candidates(row.get("phone"))) == 1
    )
    payloads: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    phone_assigned = 0
    phone_withheld = 0

    for row in active_rows:
        source_row = int(row["__source_row__"])
        center_name = _text(row.get("center"))
        if not center_name or center_name not in center_ids:
            issues.append(
                {
                    "source_row": source_row,
                    "legacy_id": row.get("id"),
                    "name": _text(row.get("name")),
                    "issue": "CENTER_NOT_MAPPED",
                    "center": center_name,
                }
            )
            continue

        candidates = _phone_candidates(row.get("phone"))
        phone = (
            candidates[0]
            if len(candidates) == 1 and single_phone_counts[candidates[0]] == 1
            else None
        )
        if phone:
            phone_assigned += 1
        else:
            phone_withheld += 1

        note_parts = [
            value
            for value in (
                _text(row.get("remark")),
                _text(row.get("role_notes")),
            )
            if value
        ]
        payloads.append(
            {
                "member_code": f"LEGACY-{row['id']}",
                "name": _text(row.get("name"), 255),
                "org_unit_id": center_ids[center_name],
                "phone": phone,
                "company_name": _text(row.get("company"), 500),
                "gender": {"男": "MALE", "女": "FEMALE"}.get(
                    _text(row.get("gender"))
                ),
                "district": _text(row.get("district"), 255),
                "company_address": _text(row.get("address"), 1000),
                "class_name": _text(row.get("class_name"), 255),
                "group_name": _text(row.get("group_name"), 255),
                "birthday": _date_text(row.get("birthday")),
                "join_date": _date_text(row.get("join_date")),
                "study_start_date": _date_text(row.get("study_start")),
                "membership_years": _number(row.get("years")),
                "renewal_month": _renewal_month(row),
                "status": "ACTIVE",
                "position": _text(row.get("position"), 255),
                "referrer": _text(row.get("referrer"), 255),
                "referrer_center": _text(row.get("referrer_center"), 255),
                "industry_category": _text(row.get("industry_type"), 255),
                "industry": _text(row.get("industry"), 255),
                "company_products": _text(row.get("product"), 4000),
                "company_size": _text(row.get("scale"), 255),
                "profit_margin": _text(row.get("profit"), 64),
                "notes": _text("；".join(note_parts), 4000),
            }
        )

    summary = {
        "source_count": len(rows),
        "active_count": len(active_rows),
        "ready_count": len(payloads),
        "excluded_count": len(issues),
        "phone_assigned_count": phone_assigned,
        "phone_withheld_count": phone_withheld,
    }
    return payloads, issues, summary


def main() -> int:
    parser = argparse.ArgumentParser(
        description="通过正式 API 幂等导入在册学员；默认仅预览"
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--username", default="admin")
    parser.add_argument(
        "--password-env", default="SEIWAJYUKU_ADMIN_PASSWORD"
    )
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm-production", action="store_true")
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    if args.apply and not args.confirm_production:
        parser.error("--apply 必须同时提供 --confirm-production")
    password = os.getenv(args.password_env, "")
    if not password:
        parser.error(f"环境变量 {args.password_env} 未设置")
    if not args.workbook.is_file():
        parser.error(f"工作簿不存在: {args.workbook}")

    with httpx.Client(
        base_url=args.base_url.rstrip("/"),
        timeout=30,
        follow_redirects=True,
    ) as client:
        token = _login(client, args.username, password)
        client.headers["Authorization"] = f"Bearer {token}"
        org_response = client.get("/api/v1/org-units/tree")
        org_response.raise_for_status()
        center_ids = {
            row["name"].strip(): row["id"]
            for row in org_response.json()["data"]
            if row["unit_type"] == "REGIONAL_CENTER"
        }

        rows = _source_rows(args.workbook)
        payloads, issues, summary = _build_payloads(rows, center_ids)
        existing_response = client.get("/api/v1/members")
        existing_response.raise_for_status()
        existing_codes = {
            row["member_code"] for row in existing_response.json()["data"]
        }
        pending = [
            payload
            for payload in payloads
            if payload["member_code"] not in existing_codes
        ]
        selected = pending[args.offset :]
        if args.limit > 0:
            selected = selected[: args.limit]

        imported = 0
        failures: list[dict[str, Any]] = []
        if args.apply:
            for index, payload in enumerate(selected, start=args.offset + 1):
                try:
                    response = client.post("/api/v1/members", json=payload)
                    response.raise_for_status()
                    imported += 1
                except Exception as exc:
                    failures.append(
                        {
                            "item": index,
                            "member_code": payload["member_code"],
                            "name": payload["name"],
                            "error": str(exc)[:500],
                        }
                    )

        output = {
            **summary,
            "existing_count": len(existing_codes),
            "pending_count": len(pending),
            "selected_count": len(selected),
            "mode": "apply" if args.apply else "preview",
            "imported_count": imported,
            "failure_count": len(failures),
            "issues": issues,
            "failures": failures,
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return 2 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
