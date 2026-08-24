"""Read-only normalization and controlled import helpers for the 2026 plan.

The Excel workbooks are treated as source evidence.  Only the explicitly
listed formal cohort sheets are read; draft, pilot and explanatory sheets are
never discovered implicitly.  The JSON produced by this module is the review
boundary between source workbooks and runtime database rows.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable


PLAN_KEY = "standard-3y"
VERSION_LABEL = "2026"
COHORT_MONTHS = (1, 4, 7, 10)
TASK_TYPES = {
    "CLASS_MEETING",
    "GROUP_MEETING",
    "ONLINE_COURSE",
    "OFFLINE_COURSE",
    "READING",
    "SUPPLEMENTARY_READING",
    "ENTERPRISE_VISIT",
    "CARE",
    "KONPA",
    "REPORT_MEETING",
    "STUDY_TOUR",
    "ENTERPRISE_PRACTICE",
    "OTHER",
}

FORMAL_SHEETS: dict[int, dict[int, str]] = {
    1: {
        1: "26版1月开班",
        4: "26版4月开班",
        7: "26版7月开班",
        10: "26版10月开班",
    },
    2: {
        1: "2026版1月开班第二年",
        4: "2026版4月开班第二年",
        7: "2026版7月开班第二年",
        10: "2026版10月开班第二年",
    },
    3: {
        1: "1月开班",
        4: "4月开班",
        7: "7月开班",
        10: "10月开班",
    },
}

_MONTH_RE = re.compile(r"^\s*(\d{1,2})\s*月\s*$")
_DURATION_RE = re.compile(r"(\d+)\s*分钟")
_BRACKET_COURSE_RE = re.compile(
    r"(?P<open>[【\[《])(?P<title>[^】\]》\r\n]+)(?P<close>[】\]》])"
)
_NUMBER_SUFFIX_RE = re.compile(r"^\s*([0-9]+(?:\s*[、,，]\s*[0-9]+)*)")

# These are intentionally narrow.  A course without one of these confirmed
# rules keeps credit_points=null rather than inheriting a guessed default.
CREDIT_RULES: tuple[tuple[str, int, str], ...] = (
    ("幸福测评表讲解", 20, "幸福测评表讲解"),
    ("班级学习会发表稿编写", 20, "班级学习会发表稿编写"),
    ("六项精进实践", 20, "六项精进实践"),
    ("经营十二条实践", 20, "经营十二条实践"),
    ("整体核算表相关课程", 40, "整体核算表相关课程"),
    ("核算表分析与任务单制作", 40, "核算表分析与任务单制作"),
    ("核算表分析&任务单制作", 40, "核算表分析与任务单制作"),
    ("会计七原则实践", 20, "会计七原则实践"),
    ("京瓷如何制定年度计划", 40, "京瓷如何制定年度计划"),
)

_COMPONENT_TASKS = (
    ("企业走访", "ENTERPRISE_VISIT"),
    ("幸福关爱", "CARE"),
    ("空巴", "KONPA"),
)


class PlanValidationError(ValueError):
    """Raised when the B1 quality gate does not pass."""

    def __init__(self, report: dict[str, Any]):
        self.report = report
        errors = report.get("errors") or ["学习计划校验失败"]
        super().__init__("；".join(str(item) for item in errors))


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _one_line(value: Any) -> str:
    return re.sub(r"[ \t]+", " ", _text(value).replace("\n", " ")).strip()


def _normalized(value: Any) -> str:
    return re.sub(r"[\s\"“”‘’'（）()【】《》、，,；;：:]+", "", _text(value))


def _is_nonempty(value: Any) -> bool:
    return bool(_text(value))


def _month_header(ws: Any) -> tuple[int, list[tuple[int, int]]]:
    """Find a row containing the twelve calendar-month columns."""

    max_row = min(int(ws.max_row or 0), 12)
    max_col = min(max(int(ws.max_column or 0), 13), 80)
    best: tuple[int, list[tuple[int, int]]] | None = None
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        matches: list[tuple[int, int]] = []
        for cell in row:
            match = _MONTH_RE.match(_text(cell.value))
            if match:
                matches.append((int(cell.column), int(match.group(1))))
        if len(matches) == 12 and len({month for _, month in matches}) >= 10:
            best = (int(row[0].row), matches)
            break
    if best is None:
        raise ValueError(f"工作表 {ws.title!r} 未找到包含12个月份的表头")
    return best


def _stop_at_notes(label: str) -> bool:
    compact = _one_line(label)
    return compact == "备注" or compact.startswith("备注：") or compact == "说明"


def _section_type(section: str, raw: str) -> str:
    section_text = _one_line(section)
    raw_text = _one_line(raw)
    if all(component in raw_text for component, _ in _COMPONENT_TASKS):
        return "COMPONENTS"
    if raw_text in {"空巴", "空巴晚宴"}:
        return "KONPA"
    if raw_text in {"幸福关爱", "幸福关爱活动"}:
        return "CARE"
    if raw_text in {"企业走访", "学员企业走访"}:
        return "ENTERPRISE_VISIT"
    if "线上课程" in section_text:
        return "ONLINE_COURSE"
    if "线下课程" in section_text:
        return "OFFLINE_COURSE"
    if "专题讲座" in section_text:
        return "MIXED_COURSE"
    if "每日读书" in section_text or "每日阅读" in section_text:
        return "READING"
    if "辅读" in section_text or "辅导教材" in section_text or "辅读书目" in section_text:
        return "SUPPLEMENTARY_READING"
    if "班级学习会" in section_text:
        return "CLASS_MEETING"
    if "小组学习会" in section_text or "小组学习会" in raw_text:
        return "GROUP_MEETING"
    for component, task_type in _COMPONENT_TASKS:
        if component in section_text:
            return "COMPONENTS"
    if "报告会" in section_text or "报告会" in raw_text:
        return "REPORT_MEETING"
    if "游学" in section_text or "游学" in raw_text:
        return "STUDY_TOUR"
    if "企业导入" in section_text or "阿米巴经营" in section_text or "企业落地" in section_text:
        return "ENTERPRISE_PRACTICE"
    if "幸福关爱" in section_text:
        return "CARE"
    if "空巴" in section_text:
        return "KONPA"
    if "企业走访" in section_text or "参观企业" in section_text:
        return "ENTERPRISE_VISIT"
    return "OTHER"


def _course_fragments(raw: str) -> list[str]:
    """Split one course cell into one fragment per explicitly named course."""

    text = _text(raw)
    matches = list(_BRACKET_COURSE_RE.finditer(text))
    if matches:
        fragments: list[str] = []
        for index, match in enumerate(matches):
            end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
            segment = text[match.start() : end].strip(" \t\n/；;，,")
            base = _one_line(match.group("title"))
            suffix_text = text[match.end() : end]
            suffix_match = _NUMBER_SUFFIX_RE.match(suffix_text)
            if suffix_match:
                numbers = re.split(r"\s*[、,，]\s*", suffix_match.group(1))
                if len(numbers) > 1:
                    for number in numbers:
                        fragments.append(f"{base}{number}{segment[len(match.group(0)) + len(suffix_match.group(1)) :]}")
                    continue
            fragments.append(segment)
        return [item for item in fragments if _is_nonempty(item)]

    # Some source cells use plain line-separated titles.  Do not split on a
    # comma by default: commas also occur inside explanatory sentences.
    chunks = [
        chunk.strip(" \t\n/；;")
        for chunk in re.split(r"\n+|[；;]", text)
        if _is_nonempty(chunk)
    ]
    return chunks or [text]


def _course_title(fragment: str, task_type: str) -> str:
    title = _one_line(fragment)
    title = re.sub(r"^(?:线上课程|线下课程)\s*[：:]\s*", "", title)
    title = re.sub(r"[【\[《]([^】\]》]+)[】\]》]", r"\1", title)
    title = re.sub(r"[（(][^（）()]*\d+\s*分钟[^（）()]*[）)]", "", title)
    title = re.sub(r"[（(][^（）()]*?(?:老师|教授|总)[^（）()]*[）)]", "", title)
    title = re.sub(r"\s*(?:视频学习|视频|学习|研讨|讨论)\s*[*+＋]?\s*$", "", title)
    title = title.strip(" \t\n/；;，,。．")
    if not title:
        return "线上课程" if task_type == "ONLINE_COURSE" else "线下课程"
    return title


def _credit_rule(title: str) -> tuple[int | None, str | None]:
    normalized = _normalized(title)
    for keyword, points, canonical_key in CREDIT_RULES:
        if _normalized(keyword) in normalized:
            return points, canonical_key
    return None, None


def _task_metadata(
    *,
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    source_year: int,
    cohort_month: int,
    nominal_calendar_month: int,
    section: str,
    source_text: str,
    canonical_key: str | None,
) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "source_file": source_file,
        "source_sheet": sheet,
        "source_cell": f"{_column_name(col)}{row}",
        "source_year": source_year,
        "source_cohort_month": cohort_month,
        "nominal_calendar_month": nominal_calendar_month,
        "source_section": _one_line(section),
        "source_text": source_text,
    }
    duration = _DURATION_RE.search(source_text)
    if duration:
        metadata["duration_minutes"] = int(duration.group(1))
    if "集体" in source_text:
        metadata["group_collective_required"] = True
    if canonical_key:
        metadata["canonical_key"] = canonical_key
    return metadata


def _column_name(index: int) -> str:
    result = ""
    value = int(index)
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _make_task(
    *,
    task_type: str,
    title: str,
    raw: str,
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    source_year: int,
    cohort_month: int,
    nominal_calendar_month: int,
    section: str,
) -> dict[str, Any]:
    credit_points, canonical_key = _credit_rule(title)
    return {
        "task_type": task_type,
        "title": title,
        "description": raw,
        "credit_points": credit_points,
        "canonical_key": canonical_key,
        "is_required": "可选" not in raw,
        "sort_order": row * 100,
        "metadata": _task_metadata(
            source_file=source_file,
            sheet=sheet,
            row=row,
            col=col,
            source_year=source_year,
            cohort_month=cohort_month,
            nominal_calendar_month=nominal_calendar_month,
            section=section,
            source_text=raw,
            canonical_key=canonical_key,
        ),
    }


def _tasks_for_cell(
    *,
    section: str,
    raw: str,
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    source_year: int,
    cohort_month: int,
    nominal_calendar_month: int,
) -> list[dict[str, Any]]:
    kind = _section_type(section, raw)
    results: list[dict[str, Any]] = []

    if kind == "COMPONENTS":
        for component, task_type in _COMPONENT_TASKS:
            if component in _one_line(section) or component in _one_line(raw):
                results.append(
                    _make_task(
                        task_type=task_type,
                        title=component,
                        raw=raw,
                        source_file=source_file,
                        sheet=sheet,
                        row=row,
                        col=col,
                        source_year=source_year,
                        cohort_month=cohort_month,
                        nominal_calendar_month=nominal_calendar_month,
                        section=section,
                    )
                )
        return results

    if kind in {"ONLINE_COURSE", "OFFLINE_COURSE"}:
        task_type = kind
        for fragment in _course_fragments(raw):
            results.append(
                _make_task(
                    task_type=task_type,
                    title=_course_title(fragment, task_type),
                    raw=fragment,
                    source_file=source_file,
                    sheet=sheet,
                    row=row,
                    col=col,
                    source_year=source_year,
                    cohort_month=cohort_month,
                    nominal_calendar_month=nominal_calendar_month,
                    section=section,
                )
            )
        return results

    if kind == "MIXED_COURSE":
        mixed_matches = list(
            re.finditer(
                r"(线上课程|线下课程)\s*[：:]\s*(.*?)(?=(?:\n\s*)?(?:线上课程|线下课程)\s*[：:]|$)",
                raw,
                flags=re.IGNORECASE | re.DOTALL,
            )
        )
        if not mixed_matches:
            mixed_matches = [re.match(r"(线上课程|线下课程)\s*[：:]\s*(.*)", raw, re.DOTALL)]
        for match in mixed_matches:
            if not match:
                continue
            task_type = "ONLINE_COURSE" if match.group(1).startswith("线上") else "OFFLINE_COURSE"
            for fragment in _course_fragments(match.group(2)):
                results.append(
                    _make_task(
                        task_type=task_type,
                        title=_course_title(fragment, task_type),
                        raw=fragment,
                        source_file=source_file,
                        sheet=sheet,
                        row=row,
                        col=col,
                        source_year=source_year,
                        cohort_month=cohort_month,
                        nominal_calendar_month=nominal_calendar_month,
                        section=section,
                    )
                )
        if results:
            return results

    if kind == "GROUP_MEETING":
        title = "小组学习会"
    elif kind == "CLASS_MEETING":
        first_line = next((line.strip() for line in _text(raw).splitlines() if line.strip()), "")
        title = f"班级学习会：{first_line}" if first_line else "班级学习会"
    elif kind == "READING":
        title = "每日读书"
    elif kind == "SUPPLEMENTARY_READING":
        title = "辅读教材"
    elif kind == "OTHER":
        title = _one_line(raw) or _one_line(section) or "其他学习任务"
    else:
        title = _one_line(raw) or _one_line(section) or "学习任务"

    results.append(
        _make_task(
            task_type=kind if kind in TASK_TYPES else "OTHER",
            title=title,
            raw=_text(raw),
            source_file=source_file,
            sheet=sheet,
            row=row,
            col=col,
            source_year=source_year,
            cohort_month=cohort_month,
            nominal_calendar_month=nominal_calendar_month,
            section=section,
        )
    )
    return results


def parse_workbook(path: str | Path, source_year: int) -> dict[str, Any]:
    """Parse the four whitelisted cohort sheets for one study year."""

    from openpyxl import load_workbook

    workbook_path = Path(path)
    if source_year not in FORMAL_SHEETS:
        raise ValueError(f"不支持的学习计划年份: {source_year}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        tracks: list[dict[str, Any]] = []
        for cohort_month, sheet_name in FORMAL_SHEETS[source_year].items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"{workbook_path.name} 缺少正式工作表 {sheet_name!r}；"
                    f"不会从其他草案页自动替代"
                )
            sheet = workbook[sheet_name]
            header_row, month_columns = _month_header(sheet)
            data_rows = sheet.iter_rows(
                min_row=header_row + 2,
                max_row=int(sheet.max_row or header_row + 1),
                min_col=1,
                max_col=max(column for column, _ in month_columns),
            )
            section = ""
            cycles: list[dict[str, Any]] = []
            for cycle_slot, (column, nominal_month) in enumerate(month_columns, start=1):
                cycles.append(
                    {
                        "cycle_index": (source_year - 1) * 12 + cycle_slot,
                        "year_index": source_year,
                        "year_cycle_index": cycle_slot,
                        "nominal_calendar_month": nominal_month,
                        "label": f"第{source_year}学年第{cycle_slot}学习周期",
                        "tasks": [],
                    }
                )

            for row_number, row_cells in enumerate(data_rows, start=header_row + 2):
                label = _text(row_cells[0].value)
                if label and _stop_at_notes(label):
                    break
                if label:
                    section = label
                if not section:
                    continue
                for cycle_slot, (column, nominal_month) in enumerate(month_columns, start=1):
                    raw = _text(row_cells[column - 1].value)
                    if not raw or raw in {"——", "--", "—"}:
                        continue
                    tasks = _tasks_for_cell(
                        section=section,
                        raw=raw,
                        source_file=workbook_path.name,
                        sheet=sheet_name,
                        row=row_number,
                        col=column,
                        source_year=source_year,
                        cohort_month=cohort_month,
                        nominal_calendar_month=nominal_month,
                    )
                    for split_index, task in enumerate(tasks, start=1):
                        task["sort_order"] = row_number * 100 + split_index
                        cycles[cycle_slot - 1]["tasks"].append(task)

            tracks.append(
                {
                    "cohort_month": cohort_month,
                    "source_sheet": sheet_name,
                    "header_row": header_row,
                    "month_columns": [
                        {"column": _column_name(column), "nominal_calendar_month": month}
                        for column, month in month_columns
                    ],
                    "cycles": cycles,
                }
            )
        return {
            "source_year": source_year,
            "source_file": workbook_path.name,
            "tracks": tracks,
        }
    finally:
        workbook.close()


def build_standard_plan(
    year1_path: str | Path,
    year2_path: str | Path,
    year3_path: str | Path,
) -> dict[str, Any]:
    parsed = {
        1: parse_workbook(year1_path, 1),
        2: parse_workbook(year2_path, 2),
        3: parse_workbook(year3_path, 3),
    }
    tracks: list[dict[str, Any]] = []
    for cohort_month in COHORT_MONTHS:
        cycles: list[dict[str, Any]] = []
        for source_year in (1, 2, 3):
            source_track = next(
                track for track in parsed[source_year]["tracks"] if track["cohort_month"] == cohort_month
            )
            cycles.extend(source_track["cycles"])
        tracks.append({"cohort_month": cohort_month, "cycles": cycles})

    plan: dict[str, Any] = {
        "schema_version": 1,
        "plan_key": PLAN_KEY,
        "plan_name": "标准班三年学习计划",
        "version_label": VERSION_LABEL,
        "duration_cycles": 36,
        "status": "DRAFT",
        "source": {
            "year1_file": Path(year1_path).name,
            "year2_file": Path(year2_path).name,
            "year3_file": Path(year3_path).name,
            "formal_sheets": {
                str(year): {str(month): sheet for month, sheet in FORMAL_SHEETS[year].items()}
                for year in (1, 2, 3)
            },
            "credit_rules": [
                {"keyword": keyword, "credit_points": points, "canonical_key": canonical_key}
                for keyword, points, canonical_key in CREDIT_RULES
            ],
        },
        "review_checklist": [
            {
                "cohort_month": cohort_month,
                "cycle_indexes": [1, 6, 12, 13, 18, 24, 25, 30, 36],
                "status": "PENDING",
            }
            for cohort_month in COHORT_MONTHS
        ],
        "cohort_tracks": tracks,
    }
    plan["quality_report"] = summarize_plan(plan)
    return plan


def _iter_cycles(plan: dict[str, Any]) -> Iterable[tuple[int, dict[str, Any]]]:
    for track in plan.get("cohort_tracks", []):
        cohort_month = int(track.get("cohort_month"))
        for cycle in track.get("cycles", []):
            yield cohort_month, cycle


def summarize_plan(plan: dict[str, Any]) -> dict[str, Any]:
    cycles = list(_iter_cycles(plan))
    tasks = [task for _, cycle in cycles for task in cycle.get("tasks", [])]
    type_counts = Counter(str(task.get("task_type")) for task in tasks)
    credits_confirmed = sum(task.get("credit_points") is not None for task in tasks)
    missing_group = [
        {"cohort_month": cohort, "cycle_index": cycle.get("cycle_index")}
        for cohort, cycle in cycles
        if not any(task.get("task_type") == "GROUP_MEETING" for task in cycle.get("tasks", []))
    ]
    return {
        "track_count": len(plan.get("cohort_tracks", [])),
        "cycle_count": len(cycles),
        "cycles_per_track": {
            str(track.get("cohort_month")): len(track.get("cycles", []))
            for track in plan.get("cohort_tracks", [])
        },
        "task_count": len(tasks),
        "task_type_counts": dict(sorted(type_counts.items())),
        "confirmed_credit_task_count": credits_confirmed,
        "credit_pending_task_count": len(tasks) - credits_confirmed,
        "empty_cycle_count": sum(not cycle.get("tasks") for _, cycle in cycles),
        "missing_group_meeting_cycles": missing_group,
        "unknown_task_type_count": sum(task.get("task_type") not in TASK_TYPES for task in tasks),
        "errors": [],
        "warnings": [],
    }


def validate_plan(plan: dict[str, Any]) -> dict[str, Any]:
    report = summarize_plan(plan)
    errors: list[str] = []
    warnings: list[str] = []
    if plan.get("schema_version") != 1:
        errors.append("schema_version 必须为 1")
    if plan.get("plan_key") != PLAN_KEY:
        errors.append(f"plan_key 必须为 {PLAN_KEY}")
    if plan.get("version_label") != VERSION_LABEL:
        errors.append(f"version_label 必须为 {VERSION_LABEL}")
    if plan.get("duration_cycles") != 36:
        errors.append("duration_cycles 必须为 36")
    if plan.get("status") != "DRAFT":
        errors.append("标准化源必须保持 DRAFT，发布由受控流程完成")

    tracks = plan.get("cohort_tracks")
    if not isinstance(tracks, list) or len(tracks) != 4:
        errors.append("必须正好包含 4 条开班轨道")
        tracks = tracks if isinstance(tracks, list) else []
    seen_cohorts: set[int] = set()
    for track in tracks:
        cohort = track.get("cohort_month")
        if cohort not in COHORT_MONTHS:
            errors.append(f"非法开班批次: {cohort}")
            continue
        if cohort in seen_cohorts:
            errors.append(f"开班批次重复: {cohort}月")
        seen_cohorts.add(cohort)
        cycles = track.get("cycles")
        if not isinstance(cycles, list) or len(cycles) != 36:
            errors.append(f"{cohort}月轨道必须包含36个周期")
            cycles = cycles if isinstance(cycles, list) else []
        indexes = [cycle.get("cycle_index") for cycle in cycles]
        expected_indexes = list(range(1, 37))
        if indexes != expected_indexes:
            errors.append(f"{cohort}月轨道 cycle_index 断档或串轨")
        for cycle in cycles:
            index = cycle.get("cycle_index")
            year_index = cycle.get("year_index")
            year_cycle_index = cycle.get("year_cycle_index")
            expected_year = ((int(index) - 1) // 12) + 1 if isinstance(index, int) else None
            expected_year_cycle = ((int(index) - 1) % 12) + 1 if isinstance(index, int) else None
            if year_index != expected_year or year_cycle_index != expected_year_cycle:
                errors.append(f"{cohort}月轨道第{index}周期的学年索引不一致")
            for task in cycle.get("tasks", []):
                task_type = task.get("task_type")
                if task_type not in TASK_TYPES:
                    errors.append(f"{cohort}月轨道第{index}周期存在未知任务类型 {task_type}")
                if task.get("credit_points") is not None and task.get("credit_points") not in {20, 40}:
                    errors.append(f"{cohort}月轨道第{index}周期存在非规则学分")
                if not task.get("metadata", {}).get("source_sheet") or not task.get("metadata", {}).get("source_cell"):
                    errors.append(f"{cohort}月轨道第{index}周期任务缺少源单元格元数据")

    if len(seen_cohorts) != 4:
        errors.append("开班批次必须覆盖 1、4、7、10 月")
    if report["cycle_count"] != 144:
        errors.append("总周期必须为 144")
    if report["unknown_task_type_count"]:
        errors.append("存在无法标准分类的任务类型")
    if report["missing_group_meeting_cycles"]:
        errors.append(
            "小组学习会缺失周期: "
            + ", ".join(
                f"{item['cohort_month']}月/{item['cycle_index']}"
                for item in report["missing_group_meeting_cycles"][:12]
            )
        )
    if report["empty_cycle_count"]:
        warnings.append(f"存在 {report['empty_cycle_count']} 个无任务周期，需人工确认")
    report["errors"] = errors
    report["warnings"] = warnings
    return report


def assert_valid_plan(plan: dict[str, Any]) -> dict[str, Any]:
    report = validate_plan(plan)
    if report["errors"]:
        raise PlanValidationError(report)
    return report


def import_plan(
    connection: Any,
    plan: dict[str, Any],
    *,
    actor_user_id: int | None = None,
    replace_draft: bool = True,
) -> dict[str, Any]:
    """Insert or replace one DRAFT version; never overwrite PUBLISHED/ARCHIVED."""

    assert_valid_plan(plan)
    from app.db import execute
    from app.services.audit import write_audit

    now = datetime.now(UTC).isoformat()
    existing = execute(
        connection,
        "SELECT id, status FROM learning_plan_versions WHERE plan_key=? AND version_label=?",
        (plan["plan_key"], plan["version_label"]),
    ).fetchone()
    replaced = False
    if existing:
        existing_status = str(existing["status"])
        if existing_status != "DRAFT":
            raise ValueError(
                f"版本 {plan['plan_key']}/{plan['version_label']} 当前为 {existing_status}，禁止覆盖"
            )
        if not replace_draft:
            raise ValueError("DRAFT 版本已存在；如需幂等重建请显式允许 replace_draft")
        version_id = int(existing["id"])
        execute(
            connection,
            "DELETE FROM learning_plan_tasks WHERE plan_cycle_id IN "
            "(SELECT id FROM learning_plan_cycles WHERE plan_version_id=?)",
            (version_id,),
        )
        execute(connection, "DELETE FROM learning_plan_cycles WHERE plan_version_id=?", (version_id,))
        execute(
            connection,
            "UPDATE learning_plan_versions SET plan_name=?, duration_cycles=?, status='DRAFT', "
            "source_name=?, updated_at=? WHERE id=?",
            (
                plan["plan_name"],
                plan["duration_cycles"],
                "standard-3y-2026.json",
                now,
                version_id,
            ),
        )
        replaced = True
    else:
        cursor = execute(
            connection,
            "INSERT INTO learning_plan_versions(plan_key, plan_name, version_label, duration_cycles, "
            "status, source_name, created_at, updated_at) VALUES (?, ?, ?, ?, 'DRAFT', ?, ?, ?)",
            (
                plan["plan_key"],
                plan["plan_name"],
                plan["version_label"],
                plan["duration_cycles"],
                "standard-3y-2026.json",
                now,
                now,
            ),
        )
        version_id = int(cursor.lastrowid)

    cycle_count = 0
    task_count = 0
    for track in plan["cohort_tracks"]:
        for cycle in track["cycles"]:
            cursor = execute(
                connection,
                "INSERT INTO learning_plan_cycles(plan_version_id, cohort_month, cycle_index, "
                "year_index, cycle_label, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    version_id,
                    track["cohort_month"],
                    cycle["cycle_index"],
                    cycle["year_index"],
                    cycle["label"],
                    now,
                    now,
                ),
            )
            cycle_id = int(cursor.lastrowid)
            cycle_count += 1
            for task in cycle.get("tasks", []):
                execute(
                    connection,
                    "INSERT INTO learning_plan_tasks(plan_cycle_id, task_type, title, description, "
                    "credit_points, is_required, sort_order, metadata_json, created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        cycle_id,
                        task["task_type"],
                        task["title"],
                        task.get("description"),
                        task.get("credit_points"),
                        1 if task.get("is_required", True) else 0,
                        task.get("sort_order", 0),
                        json.dumps(
                            {**task.get("metadata", {}), "canonical_key": task.get("canonical_key")},
                            ensure_ascii=False,
                            sort_keys=True,
                        ),
                        now,
                        now,
                    ),
                )
                task_count += 1

    write_audit(
        connection,
        actor_user_id=actor_user_id,
        action="learning.plan.import_draft",
        resource_type="learning_plan_version",
        resource_id=str(version_id),
        purpose="导入已通过 B1 校验的 2026 三年学习计划草稿",
        after={
            "plan_key": plan["plan_key"],
            "version_label": plan["version_label"],
            "cohort_track_count": len(plan["cohort_tracks"]),
            "cycle_count": cycle_count,
            "task_count": task_count,
            "replaced_draft": replaced,
        },
    )
    return {
        "version_id": version_id,
        "replaced_draft": replaced,
        "track_count": len(plan["cohort_tracks"]),
        "cycle_count": cycle_count,
        "task_count": task_count,
    }
