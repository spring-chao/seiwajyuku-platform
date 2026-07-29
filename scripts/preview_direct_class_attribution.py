"""Read-only quality preview for the four Suzhou direct learning classes.

The workbook supplies a member's development center while its class/group fields
supply learning-operation scope.  This command never connects to the platform
database and never emits names, phone numbers, companies, or row identifiers.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


DIRECT_CLASSES = ("黄埔一班", "黄埔二班", "先锋班", "神仙班")
DIRECT_CLASS_CODES = {
    "黄埔一班": "SUZHOU_DIRECT_HUANGPU_1",
    "黄埔二班": "SUZHOU_DIRECT_HUANGPU_2",
    "先锋班": "SUZHOU_DIRECT_PIONEER",
    "神仙班": "SUZHOU_DIRECT_IMMORTAL",
}
CLASS_POLICIES = {
    "黄埔一班": "STANDARD_STUDY_CLASS",
    "黄埔二班": "STANDARD_STUDY_CLASS",
    "先锋班": "FLEXIBLE_NO_REQUIREMENTS",
    "神仙班": "CONTRIBUTOR_FLEXIBLE_NO_REQUIREMENTS",
}
DEVELOPMENT_CENTERS = (
    "园区分中心",
    "昆山分中心",
    "吴江分中心",
    "新吴分中心",
    "张家港分中心",
    "姑苏相城分中心",
)
REQUIRED_COLUMNS = ("是否在册", "所在分中心", "所属班级", "所属小组")
NOTE_ONLY_GROUP_VALUES = {"目前不读书"}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _counter_rows(counter: Counter[str], *, key: str) -> list[dict[str, Any]]:
    return [{key: name, "count": count} for name, count in sorted(counter.items())]


def _group_is_note_only(class_name: str, group_name: str) -> bool:
    return class_name == "先锋班" or group_name in NOTE_ONLY_GROUP_VALUES


def build_preview(rows: Iterable[Mapping[str, Any]]) -> dict[str, Any]:
    """Summarize direct-class mapping quality without retaining private values."""
    active_rows = [
        row for row in rows
        if _text(row.get("是否在册")) == "在册"
        and _text(row.get("所属班级")) in DIRECT_CLASSES
    ]
    classes = Counter(_text(row.get("所属班级")) for row in active_rows)
    centers = Counter(_text(row.get("所在分中心")) for row in active_rows)
    class_centers = Counter(
        (_text(row.get("所属班级")), _text(row.get("所在分中心")))
        for row in active_rows
    )
    groups = Counter(
        (_text(row.get("所属班级")), _text(row.get("所属小组")))
        for row in active_rows
        if _text(row.get("所属小组"))
    )
    unknown_centers = sorted(set(centers) - set(DEVELOPMENT_CENTERS))
    issues: list[dict[str, Any]] = []
    if unknown_centers:
        issues.append({
            "code": "UNKNOWN_DEVELOPMENT_CENTER",
            "count": sum(centers[name] for name in unknown_centers),
            "values": unknown_centers,
        })
    missing_centers = centers.get("", 0)
    if missing_centers:
        issues.append({"code": "MISSING_DEVELOPMENT_CENTER", "count": missing_centers})

    # These are status/duplicate labels, not reliable group nodes. Preserve the
    # original value in the member note rather than losing it or creating a group.
    notes_to_preserve = [
        {
            "class_name": class_name,
            "source_group_value": group_name,
            "target_note": f"原所属小组：{group_name}",
            "count": count,
        }
        for (class_name, group_name), count in sorted(groups.items())
        if _group_is_note_only(class_name, group_name)
    ]
    group_candidates = [
        {
            "class_name": class_name,
            "group_name": group_name,
            "count": count,
            "target_relation": "STUDY_GROUP",
        }
        for (class_name, group_name), count in sorted(groups.items())
        if not _group_is_note_only(class_name, group_name)
    ]
    study_group_relation_count = sum(item["count"] for item in group_candidates)
    core_relation_count = len(active_rows) * 3

    return {
        "mode": "READ_ONLY_PREVIEW",
        "automatic_production_write_allowed": False,
        "direct_class_member_count": len(active_rows),
        "classes": _counter_rows(classes, key="class_name"),
        "class_policies": [
            {"class_name": class_name, "policy": CLASS_POLICIES[class_name]}
            for class_name in DIRECT_CLASSES
        ],
        "organization_candidates": [
            {
                "code": DIRECT_CLASS_CODES[class_name],
                "name": class_name,
                "unit_type": "CLASS",
                "parent_lookup": "unit_code:SZ_ROOT（苏州塾根节点）",
                "reuse_only_when": "同名 CLASS 且父级唯一为 SZ_ROOT",
            }
            for class_name in DIRECT_CLASSES
        ],
        "development_centers": _counter_rows(centers, key="center"),
        "class_center_matrix": [
            {"class_name": class_name, "center": center, "count": count}
            for (class_name, center), count in sorted(class_centers.items())
        ],
        "nonempty_groups": [
            {"class_name": class_name, "group_name": group_name, "count": count}
            for (class_name, group_name), count in sorted(groups.items())
        ],
        "notes_to_preserve": notes_to_preserve,
        "group_candidates": group_candidates,
        "relationship_plan": {
            "PRIMARY_REGION": len(active_rows),
            "DEVELOPMENT_RELATION": len(active_rows),
            "STUDY_CLASS": len(active_rows),
            "STUDY_GROUP": study_group_relation_count,
            "total": core_relation_count + study_group_relation_count,
        },
        "write_gates": [
            "运行时工作簿 SHA256 必须与本次预览一致，否则停止。",
            "苏州塾根节点（SZ_ROOT）、六个发展分中心及四个直属班级必须各自唯一解析；任何缺失或重名停止。",
            "成员须按受保护的唯一标识与当前生产名册逐条匹配；不匹配记录只进差异清单。",
            "黄埔二班和先锋班各 1 名旧服务独有记录必须保留为差异，禁止自动停用或删除。",
            "生产写入前须生成事务内审计、写入前快照与可执行回滚清单，并取得当次明确确认。",
        ],
        "issues": issues,
    }


def preview_workbook(path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = sheet_name or workbook.sheetnames[0]
    if selected_sheet not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{selected_sheet}")
    sheet = workbook[selected_sheet]
    values = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values, ())]
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"缺少必需列：{', '.join(missing)}")
    rows = [
        dict(zip(headers, values_row))
        for values_row in values
        if any(value is not None and _text(value) for value in values_row)
    ]
    preview = build_preview(rows)
    preview.update({
        "source_file": path.name,
        "source_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "sheet_name": selected_sheet,
        "source_row_count": len(rows),
    })
    return preview


def main() -> None:
    parser = argparse.ArgumentParser(description="直属学习班发展归属只读预览")
    parser.add_argument("workbook", type=Path, help="直属班级名单 xlsx 文件")
    parser.add_argument("--sheet", help="工作表名称；默认使用第一个工作表")
    parser.add_argument("--output", type=Path, help="预览 JSON 输出路径；默认标准输出")
    args = parser.parse_args()
    preview = preview_workbook(args.workbook, args.sheet)
    rendered = json.dumps(preview, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered + "\n", encoding="utf-8")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
