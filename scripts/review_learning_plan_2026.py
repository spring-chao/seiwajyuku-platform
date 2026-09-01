from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from learning_plan_semantics import cohort_template_label, learning_cycle_label
from learning_plan_2026 import COHORT_MONTHS, assert_valid_plan


CHECKPOINT_CYCLES = (1, 6, 12, 13, 18, 24, 25, 30, 36)
REVIEW_STATUSES = {"PENDING", "CONFIRMED"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_file_variants(path: Path) -> set[str]:
    """Return raw and newline-normalized hashes for text artifacts.

    The confirmed 2026 manifest was created on Windows and stores the CRLF
    hash.  GitHub runners may checkout the same JSON with LF.  Both byte forms
    are accepted for verification; workbook hashes remain raw binary hashes.
    """

    data = path.read_bytes()
    variants = {data}
    if b"\r\n" in data or b"\n" in data:
        lf = data.replace(b"\r\n", b"\n")
        variants.add(lf)
        variants.add(lf.replace(b"\n", b"\r\n"))
    return {hashlib.sha256(item).hexdigest() for item in variants}


def sha256_file_matches(path: Path, expected: str) -> bool:
    return expected in sha256_file_variants(path)


def git_commit(repo_root: Path, revision: str) -> str:
    return subprocess.check_output(
        ["git", "-C", str(repo_root), "rev-parse", f"{revision}^{{commit}}"],
        text=True,
    ).strip()


def _compact(value: Any) -> str:
    return re.sub(r"[\s\ufeff\"“”'‘’（）()【】《》、，,；;：:。．/／+＋*]", "", str(value or ""))


def _source_text_matches(fragment: Any, source_cell: Any) -> bool:
    expected = _compact(fragment)
    actual = _compact(source_cell)
    if not expected or expected in actual:
        return True
    # Numbered course fragments such as “哲学手册编制2” are split from a
    # source cell written as “哲学手册编制1、2”.
    suffix = re.search(r"(\d+)$", expected)
    if suffix:
        base = expected[: suffix.start()]
        return bool(base and base in actual and suffix.group(1) in actual.split(base, 1)[1])
    return False


def verify_source_workbooks(
    plan: dict[str, Any],
    *,
    source_workbooks: dict[int, Path],
    checkpoint_only: bool = True,
) -> dict[str, Any]:
    """Compare JSON source references with the original workbook cell text."""

    from openpyxl import load_workbook

    workbooks = {
        year: load_workbook(path, read_only=True, data_only=True)
        for year, path in source_workbooks.items()
    }
    mismatches: list[dict[str, Any]] = []
    checked = 0
    try:
        for track in plan["cohort_tracks"]:
            for cycle in track["cycles"]:
                if checkpoint_only and cycle["cycle_index"] not in CHECKPOINT_CYCLES:
                    continue
                for task in cycle.get("tasks", []):
                    metadata = task.get("metadata") or {}
                    year = int(metadata.get("source_year"))
                    sheet_name = metadata.get("source_sheet")
                    cell = metadata.get("source_cell")
                    checked += 1
                    workbook = workbooks.get(year)
                    if workbook is None or sheet_name not in workbook.sheetnames:
                        mismatches.append(
                            {
                                "cohort_month": track["cohort_month"],
                                "cycle_index": cycle["cycle_index"],
                                "task_title": task.get("title"),
                                "reason": "source_sheet_missing",
                                "source_year": year,
                                "source_sheet": sheet_name,
                                "source_cell": cell,
                            }
                        )
                        continue
                    source_value = workbook[sheet_name][cell].value
                    if not _source_text_matches(metadata.get("source_text"), source_value):
                        mismatches.append(
                            {
                                "cohort_month": track["cohort_month"],
                                "cycle_index": cycle["cycle_index"],
                                "task_title": task.get("title"),
                                "reason": "source_text_mismatch",
                                "source_sheet": sheet_name,
                                "source_cell": cell,
                                "json_source_text": metadata.get("source_text"),
                                "workbook_cell_text": source_value,
                            }
                        )
    finally:
        for workbook in workbooks.values():
            workbook.close()
    return {
        "checked_task_count": checked,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
        "status": "PASS" if not mismatches else "FAIL",
    }


def _checkpoint(plan: dict[str, Any], cohort_month: int, cycle_index: int) -> dict[str, Any]:
    track = next(track for track in plan["cohort_tracks"] if track["cohort_month"] == cohort_month)
    cycle = next(cycle for cycle in track["cycles"] if cycle["cycle_index"] == cycle_index)
    tasks = cycle.get("tasks", [])
    source_refs: dict[tuple[str, str], dict[str, Any]] = {}
    for task in tasks:
        metadata = task.get("metadata") or {}
        key = (str(metadata.get("source_sheet") or ""), str(metadata.get("source_cell") or ""))
        if not all(key):
            continue
        source_refs.setdefault(key, {"source_sheet": key[0], "source_cell": key[1]})
    task_types = Counter(str(task.get("task_type")) for task in tasks)
    confirmed_credits = [
        {
            "title": task.get("title"),
            "credit_points": task.get("credit_points"),
            "source_sheet": (task.get("metadata") or {}).get("source_sheet"),
            "source_cell": (task.get("metadata") or {}).get("source_cell"),
        }
        for task in tasks
        if task.get("credit_points") is not None
    ]
    return {
        # Keep checkpoint_id stable for existing review records and API
        # consumers; expose the business meaning explicitly beside it.
        "checkpoint_id": f"{cohort_month}-{cycle_index}",
        "checkpoint_label": f"cohort_month={cohort_month}, learning_cycle_index={cycle_index}",
        "template_key": f"COHORT_MONTH_{cohort_month:02d}",
        "template_label": cohort_template_label(cohort_month),
        "cohort_month": cohort_month,
        "cycle_index": cycle_index,
        "learning_cycle_index": cycle_index,
        "learning_cycle_label": learning_cycle_label(cohort_month, cycle_index),
        "year_index": cycle.get("year_index"),
        "year_cycle_index": cycle.get("year_cycle_index"),
        "nominal_calendar_month": cycle.get("nominal_calendar_month"),
        "source_refs": sorted(source_refs.values(), key=lambda item: (item["source_sheet"], item["source_cell"])),
        "task_count": len(tasks),
        "task_type_counts": dict(sorted(task_types.items())),
        "confirmed_credits": confirmed_credits,
        "review_fields": [
            "class_meeting_content",
            "group_meeting_content",
            "online_course_split",
            "credit_points",
            "nominal_calendar_month",
        ],
        "status": "PENDING",
        "reviewed_by": None,
        "reviewed_at": None,
        "notes": None,
    }


def expected_checkpoint_ids() -> tuple[str, ...]:
    return tuple(
        f"{cohort_month}-{cycle_index}"
        for cohort_month in COHORT_MONTHS
        for cycle_index in CHECKPOINT_CYCLES
    )


def _checkpoint_summary(checkpoints: list[dict[str, Any]]) -> tuple[str, str | None, str | None]:
    """Derive the review summary; callers must not edit the top-level status directly."""

    expected_ids = expected_checkpoint_ids()
    actual_ids = [item.get("checkpoint_id") for item in checkpoints]
    if tuple(actual_ids) != expected_ids or len(set(actual_ids)) != len(expected_ids):
        raise ValueError("审核清单的36个 checkpoint_id 不完整或顺序不正确")
    for item in checkpoints:
        if item.get("status") not in REVIEW_STATUSES:
            raise ValueError(f"审核清单包含无效状态: {item.get('status')!r}")
        for field in ("reviewed_by", "reviewed_at", "notes"):
            if field not in item:
                raise ValueError(f"审核清单缺少逐项字段: {field}")
    if not all(
        item["status"] == "CONFIRMED"
        and str(item.get("reviewed_by") or "").strip()
        and str(item.get("reviewed_at") or "").strip()
        for item in checkpoints
    ):
        return "PENDING", None, None
    reviewers = sorted({str(item["reviewed_by"]).strip() for item in checkpoints})
    reviewed_at = max(str(item["reviewed_at"]).strip() for item in checkpoints)
    return "CONFIRMED", ",".join(reviewers), reviewed_at


def derive_manifest_status(manifest: dict[str, Any]) -> str:
    checkpoints = manifest.get("checkpoints")
    if not isinstance(checkpoints, list):
        raise ValueError("审核清单必须包含 checkpoints 数组")
    return _checkpoint_summary(checkpoints)[0]


def refresh_manifest_summary(manifest: dict[str, Any]) -> str:
    checkpoints = manifest.get("checkpoints")
    if not isinstance(checkpoints, list):
        raise ValueError("审核清单必须包含 checkpoints 数组")
    status, confirmed_by, confirmed_at = _checkpoint_summary(checkpoints)
    manifest["status"] = status
    manifest["confirmed_by"] = confirmed_by
    manifest["confirmed_at"] = confirmed_at
    return status


def build_review_manifest(
    plan: dict[str, Any],
    *,
    source_commit: str,
    source_json: Path,
    source_workbooks: dict[int, Path] | None = None,
) -> dict[str, Any]:
    assert_valid_plan(plan)
    checkpoints = [
        _checkpoint(plan, cohort_month, cycle_index)
        for cohort_month in COHORT_MONTHS
        for cycle_index in CHECKPOINT_CYCLES
    ]
    manifest = {
        "review_schema_version": 1,
        "plan_key": plan["plan_key"],
        "version_label": plan["version_label"],
        "source_commit": source_commit,
        "source_json": source_json.name,
        "source_json_sha256": sha256_file(source_json),
        "status": "PENDING",
        "required_checkpoint_count": len(checkpoints),
        "created_at": datetime.now(UTC).isoformat(),
        "confirmed_at": None,
        "confirmed_by": None,
        "checkpoints": checkpoints,
    }
    refresh_manifest_summary(manifest)
    if source_workbooks:
        manifest["source_workbooks"] = {
            str(year): {"file": path.name, "sha256": sha256_file(path)}
            for year, path in sorted(source_workbooks.items())
        }
    return manifest


def verify_review_manifest(
    manifest: dict[str, Any],
    *,
    plan: dict[str, Any],
    source_json: Path,
    expected_source_commit: str | None = None,
    source_workbooks: dict[int, Path] | None = None,
    require_confirmed: bool = False,
) -> None:
    assert_valid_plan(plan)
    if manifest.get("plan_key") != plan.get("plan_key") or manifest.get("version_label") != plan.get("version_label"):
        raise ValueError("审核清单与学习计划版本不一致")
    if manifest.get("source_json") != source_json.name:
        raise ValueError("审核清单的 source_json 文件名不一致")
    if not sha256_file_matches(source_json, str(manifest.get("source_json_sha256") or "")):
        raise ValueError("审核清单的 JSON SHA-256 与当前导入源不一致")
    if expected_source_commit and manifest.get("source_commit") != expected_source_commit:
        raise ValueError("审核清单的 source_commit 与固定审核提交不一致")
    checkpoints = manifest.get("checkpoints")
    expected_count = len(COHORT_MONTHS) * len(CHECKPOINT_CYCLES)
    if manifest.get("required_checkpoint_count") != expected_count or not isinstance(checkpoints, list) or len(checkpoints) != expected_count:
        raise ValueError("审核清单必须包含36个固定抽查点")
    derived_status, derived_confirmed_by, derived_confirmed_at = _checkpoint_summary(checkpoints)
    if manifest.get("status") != derived_status:
        raise ValueError("审核清单顶层 status 必须由36个逐项审核状态自动派生，不能手工修改")
    if manifest.get("confirmed_by") != derived_confirmed_by or manifest.get("confirmed_at") != derived_confirmed_at:
        raise ValueError("审核清单顶层确认信息必须由逐项审核记录自动派生")
    stored_workbooks = manifest.get("source_workbooks")
    if stored_workbooks is not None:
        if source_workbooks is None:
            raise ValueError("审核清单包含三份原始 Excel 指纹，必须显式提供 --year1、--year2、--year3")
        expected_years = {1, 2, 3}
        if set(source_workbooks) != expected_years or set(stored_workbooks) != {"1", "2", "3"}:
            raise ValueError("审核清单必须绑定第一、第二、第三年三份原始 Excel")
        for year, path in sorted(source_workbooks.items()):
            entry = stored_workbooks.get(str(year))
            if not isinstance(entry, dict) or entry.get("file") != path.name:
                raise ValueError(f"第{year}年原始 Excel 文件名与审核指纹不一致")
            if entry.get("sha256") != sha256_file(path):
                raise ValueError(f"第{year}年原始 Excel SHA-256 与审核指纹不一致，人工确认已失效")
    if require_confirmed:
        if derived_status != "CONFIRMED":
            raise ValueError("36项人工抽查尚未确认，禁止 B2 导入")


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="生成并核验 2026 学习计划的审核指纹与36点抽查清单")
    parser.add_argument(
        "--input",
        type=Path,
        default=repo_root / "data" / "learning-plans" / "standard-3y-2026.json",
    )
    parser.add_argument(
        "--source-commit",
        required=True,
        help="固定审核提交，例如 370dc94 或完整提交 SHA",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=repo_root / "data" / "learning-plans" / "standard-3y-2026.review.json",
    )
    parser.add_argument("--verify", action="store_true", help="只核验已有审核清单，不重新生成")
    parser.add_argument("--year1", type=Path, help="第一年原始工作簿（用于源单元格回读）")
    parser.add_argument("--year2", type=Path, help="第二年原始工作簿（用于源单元格回读）")
    parser.add_argument("--year3", type=Path, help="第三年原始工作簿（用于源单元格回读）")
    parser.add_argument(
        "--confirm-checkpoint",
        action="append",
        metavar="COHORT-CYCLE",
        help="确认一个抽查项，例如 1-1；可重复指定，不能直接修改顶层状态",
    )
    parser.add_argument("--reviewed-by", help="逐项审核人；与 --confirm-checkpoint 一起使用")
    parser.add_argument("--reviewed-at", help="逐项审核时间；默认当前 UTC 时间")
    parser.add_argument("--notes", help="逐项审核备注；与 --confirm-checkpoint 一起使用")
    args = parser.parse_args()
    plan = json.loads(args.input.read_text(encoding="utf-8"))
    resolved_commit = git_commit(repo_root, args.source_commit)
    source_workbooks = None
    if any(path is not None for path in (args.year1, args.year2, args.year3)):
        if not all(path is not None for path in (args.year1, args.year2, args.year3)):
            raise ValueError("--year1、--year2、--year3 必须同时提供")
        source_workbooks = {1: args.year1, 2: args.year2, 3: args.year3}
        source_report = verify_source_workbooks(plan, source_workbooks=source_workbooks)
        print(json.dumps({"source_verification": source_report}, ensure_ascii=False, indent=2))
        if source_report["mismatch_count"]:
            return 2
    if args.confirm_checkpoint:
        if not all(path is not None for path in (args.year1, args.year2, args.year3)):
            raise ValueError("--confirm-checkpoint 必须同时提供 --year1、--year2、--year3")
        if not args.reviewed_by or not args.reviewed_by.strip():
            raise ValueError("--confirm-checkpoint 必须提供 --reviewed-by")
        manifest = json.loads(args.output.read_text(encoding="utf-8"))
        verify_review_manifest(
            manifest,
            plan=plan,
            source_json=args.input,
            expected_source_commit=resolved_commit,
            source_workbooks=source_workbooks,
        )
        checkpoint_by_id = {item["checkpoint_id"]: item for item in manifest["checkpoints"]}
        unknown = sorted(set(args.confirm_checkpoint) - set(checkpoint_by_id))
        if unknown:
            raise ValueError(f"未知抽查项: {', '.join(unknown)}")
        reviewed_at = args.reviewed_at or datetime.now(UTC).isoformat()
        for checkpoint_id in args.confirm_checkpoint:
            checkpoint = checkpoint_by_id[checkpoint_id]
            checkpoint["status"] = "CONFIRMED"
            checkpoint["reviewed_by"] = args.reviewed_by.strip()
            checkpoint["reviewed_at"] = reviewed_at
            checkpoint["notes"] = args.notes
        refresh_manifest_summary(manifest)
        args.output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(json.dumps({"status": manifest["status"], "confirmed_checkpoint_count": sum(item["status"] == "CONFIRMED" for item in manifest["checkpoints"]), "output": str(args.output)}, ensure_ascii=False, indent=2))
        return 0
    if args.verify:
        manifest = json.loads(args.output.read_text(encoding="utf-8"))
        if source_workbooks is None and manifest.get("source_workbooks") is not None:
            raise ValueError("--verify 必须同时提供 --year1、--year2、--year3 以核验三份原始 Excel 指纹")
        verify_review_manifest(
            manifest,
            plan=plan,
            source_json=args.input,
            expected_source_commit=resolved_commit,
            source_workbooks=source_workbooks,
        )
        print(json.dumps({"status": manifest.get("status"), "source_commit": resolved_commit, "source_json_sha256": sha256_file(args.input)}, ensure_ascii=False, indent=2))
        return 0
    manifest = build_review_manifest(
        plan,
        source_commit=resolved_commit,
        source_json=args.input,
        source_workbooks=source_workbooks,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": manifest["status"], "source_commit": resolved_commit, "source_json_sha256": manifest["source_json_sha256"], "checkpoint_count": manifest["required_checkpoint_count"], "output": str(args.output)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
